import os
import re
import sqlite3
import uuid
import json
import datetime
import threading
import smtplib
import hmac
import hashlib
import secrets
import shutil
import tempfile
import zipfile
import csv
import io
import ipaddress
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from html import escape, unescape
from functools import wraps
from urllib.parse import quote
from flask import Flask, request, jsonify, send_from_directory, g, Response, abort, has_request_context
from flask_cors import CORS
from werkzeug.exceptions import RequestEntityTooLarge
import jwt
import bcrypt
import stripe
from security_utils import (
    PUBLIC_UPLOAD_MAX_BYTES,
    SAFE_IMAGE_EXTENSIONS,
    UploadSecurityError,
    clean_text,
    is_safe_public_upload_url,
    is_safe_static_image_url,
    is_safe_stored_filename,
    random_stored_name,
    secure_upload_headers,
    validate_upload,
)

try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    _has_limiter = True
except ImportError:
    _has_limiter = False

# Load .env file if present (dev convenience — never commit .env to git)
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
except ImportError:
    pass

try:
    from PIL import Image, ImageOps
    Image.MAX_IMAGE_PIXELS = int(os.environ.get('UPLOAD_IMAGE_MAX_PIXELS', '20000000') or 20000000)
    _has_pillow = True
except ImportError:
    Image = ImageOps = None
    _has_pillow = False

CLIENT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'client'))
app = Flask(__name__, static_folder=None)
# Public uploads are still individually limited and scanned below. The larger
# request cap allows strict-admin bulk product imports that include an image ZIP.
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MAX_CONTENT_LENGTH_MB', '160') or 160) * 1024 * 1024
CORS_ORIGINS = [
    'https://adhyashaktishop.com',
    'https://www.adhyashaktishop.com',
]
if os.environ.get('ALLOW_DEV_CORS', 'false').lower() == 'true':
    CORS_ORIGINS.extend(['http://localhost:5000', 'http://127.0.0.1:5000'])
CORS(app, origins=CORS_ORIGINS)
STRICT_CSP_REPORT_ONLY = os.environ.get('STRICT_CSP_REPORT_ONLY', 'true').lower() == 'true'

if _has_limiter:
    limiter = Limiter(get_remote_address, app=app, default_limits=[], storage_uri='memory://')
else:
    class _NoopLimiter:
        def limit(self, *a, **kw):
            def dec(f): return f
            return dec
    limiter = _NoopLimiter()

SECRET_KEY = os.environ.get('JWT_SECRET')
if not SECRET_KEY:
    raise RuntimeError("JWT_SECRET environment variable is not set. Set it before starting the server.")
AUTH_COOKIE_NAME = os.environ.get('AUTH_COOKIE_NAME', 'adhya_auth')
AUTH_COOKIE_SECURE = os.environ.get('AUTH_COOKIE_SECURE', 'true').lower() != 'false'
AUTH_COOKIE_MAX_AGE = 7 * 24 * 60 * 60
CSRF_COOKIE_NAME = os.environ.get('CSRF_COOKIE_NAME', 'adhya_csrf')
CSRF_HEADER_NAME = 'X-CSRF-Token'
LOGIN_GENERIC_ERROR = 'Invalid email or password'
CUSTOMER_LOGIN_FAILURE_LIMIT = int(os.environ.get('CUSTOMER_LOGIN_FAILURE_LIMIT', '8') or 8)
CUSTOMER_LOGIN_LOCKOUT_SECONDS = int(os.environ.get('CUSTOMER_LOGIN_LOCKOUT_SECONDS', str(15 * 60)) or (15 * 60))
STAFF_LOGIN_FAILURE_LIMIT = int(os.environ.get('STAFF_LOGIN_FAILURE_LIMIT', '5') or 5)
STAFF_LOGIN_IP_FAILURE_LIMIT = int(os.environ.get('STAFF_LOGIN_IP_FAILURE_LIMIT', '10') or 10)
STAFF_LOGIN_LOCKOUT_SECONDS = int(os.environ.get('STAFF_LOGIN_LOCKOUT_SECONDS', str(30 * 60)) or (30 * 60))
DB_PATH = os.path.join(os.path.dirname(__file__), 'ecommerce.db')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), '..', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
BULK_IMPORT_FOLDER = os.path.join(os.path.dirname(__file__), 'bulk_import_jobs')
os.makedirs(BULK_IMPORT_FOLDER, exist_ok=True)
ADMIN_BULK_UPLOAD_MAX_BYTES = int(os.environ.get('ADMIN_BULK_UPLOAD_MAX_MB', '150') or 150) * 1024 * 1024
DUMMY_PASSWORD_HASH = bcrypt.hashpw(secrets.token_urlsafe(48).encode(), bcrypt.gensalt()).decode()

# ─── Stripe ───────────────────────────────────────────────────────────────────
# Store these in environment variables in production — never commit live keys.
# On your server: set STRIPE_SECRET_KEY and STRIPE_PUBLISHABLE_KEY env vars.
STRIPE_SECRET_KEY      = os.environ.get('STRIPE_SECRET_KEY', '')
STRIPE_PUBLISHABLE_KEY = os.environ.get('STRIPE_PUBLISHABLE_KEY', '')
stripe.api_key = STRIPE_SECRET_KEY


# ─── Email (Zoho SMTP) ────────────────────────────────────────────────────────
# Set env vars before starting Flask — e.g.:
#   set CONTACT_MAIL_PASS=YourPassword
#   set ORDER_MAIL_PASS=YourPassword
# If passwords are not set the site works normally; emails are silently skipped.
CONTACT_MAIL_USER = os.environ.get('CONTACT_MAIL_USER', 'contact@adhyashaktishop.com')
CONTACT_MAIL_PASS = os.environ.get('CONTACT_MAIL_PASS', '')
ORDER_MAIL_USER   = os.environ.get('ORDER_MAIL_USER',   'order@adhyashaktishop.com')
ORDER_MAIL_PASS   = os.environ.get('ORDER_MAIL_PASS',   '')
_SMTP_HOST, _SMTP_PORT = 'smtp.zoho.com', 587


def h(s):
    """HTML-escape user-supplied strings before interpolating into email HTML."""
    return escape(str(s)) if s else ''


def _csrf_signature(raw):
    return hmac.new(SECRET_KEY.encode('utf-8'), raw.encode('utf-8'), hashlib.sha256).hexdigest()


def make_csrf_token():
    raw = secrets.token_urlsafe(32)
    return f'{raw}.{_csrf_signature(raw)}'


def make_password_reset_token():
    return secrets.token_urlsafe(48)


def valid_password_reset_token_format(token):
    return bool(re.fullmatch(r'[A-Za-z0-9_-]{40,200}', token or ''))


def password_reset_digest(token):
    return hmac.new(
        SECRET_KEY.encode('utf-8'),
        token.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()


def valid_csrf_token(token):
    if not token or '.' not in token:
        return False
    raw, signature = token.rsplit('.', 1)
    if not raw or not re.fullmatch(r'[A-Za-z0-9_-]+', raw):
        return False
    return hmac.compare_digest(signature, _csrf_signature(raw))


def set_csrf_cookie(response, token):
    response.set_cookie(
        CSRF_COOKIE_NAME,
        token,
        max_age=AUTH_COOKIE_MAX_AGE,
        httponly=False,
        secure=AUTH_COOKIE_SECURE,
        samesite='Lax',
        path='/'
    )
    return response


@app.before_request
def verify_csrf_token():
    if request.method in ('GET', 'HEAD', 'OPTIONS'):
        return None
    if not request.path.startswith('/api/'):
        return None
    if request.path == '/api/csp-report':
        return None

    header_token = request.headers.get(CSRF_HEADER_NAME, '')
    cookie_token = request.cookies.get(CSRF_COOKIE_NAME, '')
    if not header_token or not cookie_token:
        log_security_event('csrf_failed', 'warning', 'Missing CSRF token')
        return jsonify({'error': 'Security check failed. Please refresh the page and try again.'}), 403
    if not hmac.compare_digest(header_token, cookie_token) or not valid_csrf_token(header_token):
        log_security_event('csrf_failed', 'warning', 'Invalid CSRF token')
        return jsonify({'error': 'Security check failed. Please refresh the page and try again.'}), 403
    return None


def _content_security_policy(allow_inline=True, report_only=False):
    # CSP3 lets modern browsers distinguish script elements from legacy inline
    # event attributes. UI actions are delegated through /js/csp-actions.js, so
    # inline script elements and inline event attributes stay blocked.
    script_attr = "'none'"
    script_fallback_inline = ""
    style_attr = "'unsafe-inline'"
    directives = [
        "default-src 'self'",
        # Fallback for older CSP engines. CSP3 engines use script-src-elem/attr below.
        f"script-src 'self' https://js.stripe.com https://static.cloudflareinsights.com{script_fallback_inline}",
        "script-src-elem 'self' https://js.stripe.com https://static.cloudflareinsights.com",
        f"script-src-attr {script_attr}",
        "style-src 'self' https://cdnjs.cloudflare.com https://fonts.googleapis.com 'unsafe-inline'",
        "style-src-elem 'self' https://cdnjs.cloudflare.com https://fonts.googleapis.com",
        f"style-src-attr {style_attr}",
        "font-src 'self' https://cdnjs.cloudflare.com https://fonts.gstatic.com data:",
        "img-src 'self' data: blob: https:",
        "media-src 'self' blob:",
        "connect-src 'self' https://api.stripe.com https://checkout.stripe.com https://r.stripe.com https://m.stripe.network https://cloudflareinsights.com",
        "frame-src 'self' https://js.stripe.com https://hooks.stripe.com",
        "object-src 'none'",
        "base-uri 'self'",
        "form-action 'self'",
        "frame-ancestors 'self'",
        "manifest-src 'self'",
        "worker-src 'self' blob:",
        "upgrade-insecure-requests",
        "block-all-mixed-content",
    ]
    if report_only:
        directives.append("report-uri /api/csp-report")
    return '; '.join(directives) + ';'


@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-Permitted-Cross-Domain-Policies'] = 'none'
    response.headers['X-Download-Options'] = 'noopen'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Cross-Origin-Opener-Policy'] = 'same-origin-allow-popups'
    response.headers['Origin-Agent-Cluster'] = '?1'
    response.headers['Permissions-Policy'] = (
        'camera=(), microphone=(), geolocation=(), usb=(), '
        'payment=(self), interest-cohort=()'
    )
    has_route_csp = 'Content-Security-Policy' in response.headers
    if not has_route_csp:
        response.headers['Content-Security-Policy'] = _content_security_policy(allow_inline=True)
    if STRICT_CSP_REPORT_ONLY and not has_route_csp:
        response.headers['Content-Security-Policy-Report-Only'] = _content_security_policy(
            allow_inline=False,
            report_only=True
        )
    private_prefixes = (
        '/api/auth',
        '/api/admin',
        '/api/acc',
        '/api/orders',
        '/api/user',
        '/api/wishlist',
        '/api/upload',
        '/admin',
        '/accounts',
    )
    if request.path.startswith(private_prefixes):
        response.headers['Cache-Control'] = 'no-store, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    if request.path.startswith(('/api/', '/admin', '/accounts')):
        response.headers['X-Robots-Tag'] = 'noindex, nofollow'
    return response


@app.route('/api/csp-report', methods=['POST'])
@limiter.limit("60 per minute")
def csp_report():
    payload = request.get_json(silent=True) or {}
    try:
        app.logger.warning("CSP report: %s", json.dumps(payload)[:1200])
    except Exception:
        pass  # nosec B110
    return ('', 204)


def _email_html(content, preheader=''):
    safe_preheader = h(preheader)
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><style>
body{{font-family:Segoe UI,Arial,sans-serif;background:#f7f4ee;margin:0;padding:24px;color:#26332f}}
.wrap{{max-width:640px;margin:0 auto;background:#fff;border-radius:16px;overflow:hidden;border:1px solid #e7dfd2;box-shadow:0 6px 22px rgba(29,92,74,.10)}}
.preheader{{display:none!important;visibility:hidden;opacity:0;color:transparent;height:0;width:0;overflow:hidden;mso-hide:all}}
.hdr{{background:#0f3d31;padding:26px 32px;text-align:center}}
.hdr h1{{color:#fff;font-size:1.38rem;margin:0;font-family:Georgia,serif;letter-spacing:.2px}}
.hdr span{{color:#D6AE3D}}
.hdr p{{color:#d7eee6;font-size:.78rem;letter-spacing:.08em;text-transform:uppercase;margin:8px 0 0}}
.body{{padding:34px 34px 28px}}
.body h2{{color:#123f33;font-size:1.28rem;margin:0 0 14px;line-height:1.35}}
.body h3{{color:#123f33;font-size:1rem;margin:22px 0 10px}}
.body p{{color:#40514b;line-height:1.7;margin:0 0 14px}}
.muted{{color:#6b7280;font-size:.88rem}}
.cta{{display:inline-block;background:#1D5C4A;color:#fff!important;padding:12px 24px;border-radius:8px;text-decoration:none!important;font-weight:700;margin:8px 0}}
.cta.secondary{{background:#C49A22}}
.note{{background:#f3faf6;border:1px solid #cbe7d9;border-radius:12px;padding:14px 16px;color:#245346;margin:18px 0}}
.warning{{background:#fff8eb;border:1px solid #f0d49a;border-radius:12px;padding:14px 16px;color:#6f4b12;margin:18px 0}}
.trust{{background:#f8f6f0;border-top:1px solid #eee7dc;border-bottom:1px solid #eee7dc;padding:14px 34px;color:#52615c;font-size:.82rem;text-align:center}}
.ftr{{background:#fbfaf7;border-top:1px solid #eee7dc;padding:24px 32px;text-align:center}}
table.items{{width:100%;border-collapse:collapse;margin:16px 0;font-size:.9rem;border:1px solid #e9e3d8;border-radius:10px;overflow:hidden}}
table.items th{{background:#f4f1ea;padding:10px;text-align:left;border-bottom:1px solid #e9e3d8;color:#57645f;font-size:.78rem;text-transform:uppercase;letter-spacing:.04em}}
table.items td{{padding:10px;border-bottom:1px solid #eee8dd;color:#293833;vertical-align:top}}
table.items tr:last-child td{{border-bottom:0}}
@media(max-width:640px){{body{{padding:12px}}.body{{padding:26px 20px}}.hdr{{padding:22px 20px}}.trust{{padding:14px 20px}}.ftr{{padding:22px 20px}}}}
</style></head><body>
<div class="preheader">{safe_preheader}</div>
<div class="wrap">
<div class="hdr"><h1>Adhya <span>Shakti</span> Shop</h1><p>Secure checkout &bull; New Jersey, USA</p></div>
<div class="body">{content}</div>
<div class="trust">Secure checkout &bull; Order tracking &bull; Responsive support &bull; Handled by Adhya Shakti Shop</div>
<div class="ftr">
  <div style="font-family:Georgia,serif;font-size:1.05rem;font-weight:700;color:#1D5C4A;margin-bottom:4px">Adhya <span style="color:#C49A22">Shakti</span> Shop</div>
  <div style="font-size:.75rem;color:#8a8f8c;letter-spacing:.08em;text-transform:uppercase;margin-bottom:14px">Jewelry &nbsp;&bull;&nbsp; Custom Printing &nbsp;&bull;&nbsp; Personalized Gifts</div>
  <div style="margin-bottom:12px;font-size:.8rem;line-height:1.8">
    <a href="https://www.instagram.com/adhyashaktijewelry" style="color:#1D5C4A;text-decoration:none;font-weight:600;margin:0 6px">@adhyashaktijewelry</a>
    <span style="color:#ddd">&nbsp;|&nbsp;</span>
    <a href="https://www.instagram.com/adhyashaktiprinting" style="color:#1D5C4A;text-decoration:none;font-weight:600;margin:0 6px">@adhyashaktiprinting</a>
    <span style="color:#ddd">&nbsp;|&nbsp;</span>
    <a href="https://wa.me/c/18483363769" style="color:#16a34a;text-decoration:none;font-weight:600;margin:0 6px">WhatsApp</a>
  </div>
  <div style="font-size:.74rem;color:#9ca3af">New Jersey, USA &nbsp;&middot;&nbsp; <a href="mailto:contact@adhyashaktishop.com" style="color:#6b7280">contact@adhyashaktishop.com</a></div>
  <div style="font-size:.68rem;color:#b8b8b8;margin-top:8px">You are receiving this email because you interacted with Adhya Shakti Shop.</div>
  <div style="font-size:.68rem;color:#c7c7c7;margin-top:4px">&copy; 2026 Adhya Shakti Shop. All rights reserved.</div>
</div>
</div></body></html>"""


def _email_text_from_html(html):
    text = re.sub(r'(?i)<br\s*/?>', '\n', html or '')
    text = re.sub(r'(?i)</(p|div|h[1-6]|tr|li|table|blockquote)>', '\n', text)
    text = re.sub(r'(?s)<style.*?</style>', '', text)
    text = re.sub(r'(?s)<[^>]+>', '', text)
    text = unescape(text)
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n\s*\n+', '\n\n', text)
    return text.strip()


def _send_email_bg(from_addr, from_pass, to_addr, subject, html):
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'Adhya Shakti Shop <{from_addr}>'
        msg['To']      = to_addr
        msg['Reply-To'] = from_addr
        msg.attach(MIMEText(_email_text_from_html(html), 'plain', 'utf-8'))
        msg.attach(MIMEText(html, 'html', 'utf-8'))
        with smtplib.SMTP(_SMTP_HOST, _SMTP_PORT, timeout=10) as s:
            s.ehlo(); s.starttls(); s.login(from_addr, from_pass)
            s.sendmail(from_addr, [to_addr], msg.as_string())
        return True
    except Exception as exc:
        app.logger.error('Email failed to %s: %s', to_addr, exc)
        return False


def send_email(from_addr, from_pass, to_addr, subject, html):
    if not from_pass:
        return
    threading.Thread(target=_send_email_bg,
                     args=(from_addr, from_pass, to_addr, subject, html),
                     daemon=True).start()


WELCOME_COUPON_CODE = 'WELCOME10'


def _welcome_offer_email_block(code):
    return f"""
<div style="text-align:center;margin:28px 0">
  <div style="background:#f0fdf4;border:2px solid #16a34a;border-radius:14px;padding:24px 32px;display:inline-block;min-width:260px">
    <div style="font-size:.7rem;font-weight:800;letter-spacing:2.5px;text-transform:uppercase;color:#166534;margin-bottom:12px">First Order Welcome Code</div>
    <div style="font-size:2.2rem;font-weight:800;letter-spacing:5px;color:#1D5C4A;font-family:Georgia,serif;line-height:1">{h(code)}</div>
    <div style="width:40px;height:2px;background:#16a34a;margin:14px auto"></div>
    <div style="font-size:.88rem;color:#374151;font-weight:600">10% off your first order</div>
    <div style="font-size:.78rem;color:#6b7280;margin-top:6px">One-time use, tied to this email</div>
  </div>
</div>
<p>Enter <strong>{h(code)}</strong> at checkout with this same email address. This code works once, only on your first order.</p>"""


def email_welcome(name, to_email, welcome_code=None):
    offer_block = _welcome_offer_email_block(welcome_code) if welcome_code else ''
    preheader = 'Your Adhya Shakti Shop account is ready.'
    if welcome_code:
        preheader = f'Your account is ready. Use {welcome_code} for 10% off your first order.'
    html = _email_html(f"""
<h2>Welcome to Adhya Shakti Shop — You're In!</h2>
<p>Hi {h(name)},</p>
<p>We're so happy to have you here. Your account is all set. Here's what you can look forward to:</p>
<ul style="line-height:2;color:#444;padding-left:20px">
  <li><strong>Handcrafted Jewelry</strong> — unique pieces inspired by Indian heritage</li>
  <li><strong>Custom Printing</strong> — T-shirts, polo shirts, and hoodies</li>
  <li><strong>Personalized Gifts</strong> — made with love, right here in New Jersey</li>
</ul>
<p>Whether you're shopping for yourself or gifting someone special, we'll make sure every order feels personal.</p>
{offer_block}
<div style="text-align:center;margin:28px 0">
  <a href="https://adhyashaktishop.com/products" class="cta">Explore the Collection</a>
  <br>
  <a href="https://adhyashaktishop.com/dashboard" style="display:inline-block;margin-top:12px;font-size:.85rem;color:#1D5C4A">Visit My Account</a>
</div>
<p style="font-size:.88rem;color:#666;border-top:1px solid #eee;padding-top:16px;margin-top:8px">
  Follow us for new arrivals &amp; behind-the-scenes:<br>
  <a href="https://www.instagram.com/adhyashaktijewelry" style="color:#1D5C4A">@adhyashaktijewelry</a> &nbsp;&middot;&nbsp;
  <a href="https://www.instagram.com/adhyashaktiprinting" style="color:#1D5C4A">@adhyashaktiprinting</a>
</p>
<p style="font-size:.82rem;color:#999;margin-top:12px">Your account email: {h(to_email)}</p>""", preheader)
    send_email(CONTACT_MAIL_USER, CONTACT_MAIL_PASS, to_email,
               'Welcome to Adhya Shakti Shop - Your Account Is Ready', html)


def email_welcome_discount(to_email, code):
    html = _email_html(f"""
<h2>Your 10% Off Code Is Inside</h2>
<p>Thank you for joining the Adhya Shakti Shop family. Here is the first-order code you asked for:</p>
{_welcome_offer_email_block(code)}
<div style="text-align:center;margin:28px 0">
  <a href="https://adhyashaktishop.com/products" class="cta">Shop Now</a>
</div>
<p class="muted">If you already used this welcome offer or already placed your first order, the code cannot be used again.</p>
<p style="font-size:.85rem;color:#555;border-top:1px solid #eee;padding-top:16px;margin-top:8px">
  Follow us for new arrivals &amp; behind-the-scenes:<br>
  <a href="https://www.instagram.com/adhyashaktijewelry" style="color:#1D5C4A;font-weight:600">@adhyashaktijewelry</a>
  &nbsp;&middot;&nbsp;
  <a href="https://www.instagram.com/adhyashaktiprinting" style="color:#1D5C4A;font-weight:600">@adhyashaktiprinting</a>
</p>""", f'Use {code} for 10% off your first order.')
    send_email(CONTACT_MAIL_USER, CONTACT_MAIL_PASS, to_email,
               'Your 10% Off Welcome Code - Adhya Shakti Shop', html)


def _build_back_in_stock_email(name, product):
    product_id = product.get('id') or ''
    product_name = product.get('name') or 'Your saved product'
    product_url = f'https://adhyashaktishop.com/product/{quote(str(product_id))}'
    price = product.get('price')
    price_line = f'<p><strong>Price:</strong> ${float(price):.2f}</p>' if price not in (None, '') else ''
    image_url = ''
    images = product.get('images') or []
    if images:
        first = images[0]
        image_url = first if str(first).startswith('http') else f'https://adhyashaktishop.com{first}'
    image_block = f"""
<div style="text-align:center;margin:18px 0">
  <img src="{h(image_url)}" alt="{h(product_name)}" style="max-width:240px;width:100%;border-radius:14px;border:1px solid #eee8dd" />
</div>""" if image_url else ''
    hello = f'Hi {h(name)},' if name else 'Hi,'
    html = _email_html(f"""
<h2>{h(product_name)} Is Back In Stock</h2>
<p>{hello}</p>
<p>The product you asked about is available again. Inventory can move quickly, so please order soon if you still want it.</p>
{image_block}
<div class="note">
  <strong>{h(product_name)}</strong><br>
  {price_line}
  <span class="muted">Availability is not reserved until checkout is completed.</span>
</div>
<div style="text-align:center;margin:28px 0">
  <a href="{h(product_url)}" class="cta">View Product</a>
</div>
<p class="muted">You received this because you requested a back-in-stock notification for this product.</p>""", f'{product_name} is available again.')
    subject = f'{product_name} is back in stock - Adhya Shakti Shop'
    return subject, html


def email_back_in_stock(to_email, name, product):
    subject, html = _build_back_in_stock_email(name, product)
    send_email(CONTACT_MAIL_USER, CONTACT_MAIL_PASS, to_email, subject, html)


def email_password_reset(name, to_email, reset_link):
    html = _email_html(f"""
<h2>Reset Your Password</h2>
<p>Hi {h(name)},</p>
<p>We received a request to reset the password for your Adhya Shakti Shop account.</p>
<div class="note">
  This secure reset link expires in <strong>1 hour</strong>. Your password will not change unless you choose a new one.
</div>
<div style="text-align:center;margin:28px 0">
  <a href="{h(reset_link)}" class="cta">Reset My Password</a>
</div>
<p class="muted">If you did not request this, you can safely ignore this email. For your protection, do not forward this reset link to anyone.</p>
<p class="muted">Account email: {h(to_email)}</p>""", 'Your password reset link expires in 1 hour.')
    send_email(CONTACT_MAIL_USER, CONTACT_MAIL_PASS, to_email,
               'Reset Your Password — Adhya Shakti Shop', html)


def email_contact_notification(name, cust_email, phone, message, inquiry_type='', order_number=''):
    inq_line = f'<p><strong>Inquiry Type:</strong> {h(inquiry_type)}</p>' if inquiry_type else ''
    ord_line  = f'<p><strong>Order Number:</strong> {h(order_number)}</p>' if order_number else ''
    shop_html = _email_html(f"""
<h2>New Contact Message</h2>
<p><strong>From:</strong> {h(name)}<br>
<strong>Email:</strong> <a href="mailto:{h(cust_email)}">{h(cust_email)}</a><br>
<strong>Phone:</strong> {h(phone) or 'Not provided'}</p>
{inq_line}{ord_line}
<hr style="border:none;border-top:1px solid #eee;margin:16px 0">
<p style="white-space:pre-wrap">{h(message)}</p>""")
    send_email(CONTACT_MAIL_USER, CONTACT_MAIL_PASS, CONTACT_MAIL_USER,
               f'New Message from {h(name)} — Adhya Shakti Shop', shop_html)

    reply_html = _email_html(f"""
<h2>We Received Your Message</h2>
<p>Hi {h(name)},</p>
<p>Thank you for reaching out to Adhya Shakti Shop. We received your message and will reply within <strong>1-2 business days</strong>.</p>
<div class="note">
  If this is about an order, please keep your order number handy so we can help faster.
</div>
<blockquote style="border-left:3px solid #C49A22;padding:12px 16px;margin:20px 0;color:#555;font-style:italic;background:#fafafa;border-radius:0 6px 6px 0">
{h(message[:400])}{'…' if len(message) > 400 else ''}
</blockquote>
<p>While you wait, you can also reach us on WhatsApp for faster help.</p>
<div style="text-align:center;margin:24px 0 8px">
  <a href="https://wa.me/c/18483363769" class="cta">Message Us on WhatsApp</a>
  <a href="https://www.instagram.com/adhyashaktijewelry" class="cta secondary" style="margin-left:8px">Instagram</a>
</div>
<p class="muted">We aim to respond to all inquiries within 1-2 business days. If your matter is urgent, WhatsApp is usually the fastest route.</p>""", 'We received your message and will reply within 1-2 business days.')
    send_email(CONTACT_MAIL_USER, CONTACT_MAIL_PASS, cust_email,
               'Got Your Message — We\'ll Be in Touch Soon | Adhya Shakti Shop', reply_html)


def _money(value):
    try:
        return f"${float(value or 0):.2f}"
    except (TypeError, ValueError):
        return "$0.00"


def _order_option_html(item):
    lines = []
    variation = clean_text(item.get('variation'), 160)
    if variation:
        if ' / ' in variation:
            color, size = [p.strip() for p in variation.split(' / ', 1)]
            if color:
                lines.append(f"<span><strong>Color:</strong> {h(color)}</span>")
            if size:
                lines.append(f"<span><strong>Size:</strong> {h(size)}</span>")
        else:
            lines.append(f"<span><strong>Option:</strong> {h(variation)}</span>")
    custom_print = item.get('customPrint') if isinstance(item.get('customPrint'), dict) else None
    if custom_print:
        placement = clean_text(custom_print.get('placement'), 20)
        placement_label = {
            'front': 'Front only',
            'back': 'Back only',
            'both': 'Front and back',
        }.get(placement, 'Custom print')
        front_count = len(custom_print.get('front_images') or [])
        back_count = len(custom_print.get('back_images') or [])
        file_bits = []
        if front_count:
            file_bits.append(f"{front_count} front file{'s' if front_count != 1 else ''}")
        if back_count:
            file_bits.append(f"{back_count} back file{'s' if back_count != 1 else ''}")
        extra = custom_print.get('extra_charge')
        extra_text = f" ({_money(extra)} extra)" if extra else ''
        file_text = f" - {h(', '.join(file_bits))}" if file_bits else ''
        lines.append(f"<span><strong>Custom print:</strong> {h(placement_label)}{extra_text}{file_text}</span>")
    if not lines:
        return ''
    return "<div class='muted' style='font-size:.84rem;line-height:1.6;margin-top:4px'>" + "<br>".join(lines) + "</div>"


def _order_item_rows(items_data, include_total=True):
    rows = []
    for item in items_data or []:
        name = h(item.get('name', 'Item'))
        try:
            qty = max(1, int(item.get('qty') or item.get('quantity') or 1))
        except (TypeError, ValueError):
            qty = 1
        try:
            price = float(item.get('price') or 0)
        except (TypeError, ValueError):
            price = 0.0
        options = _order_option_html(item)
        total_cell = f"<td style='text-align:right'>{_money(price * qty)}</td>" if include_total else ''
        rows.append(
            f"<tr>"
            f"<td><strong>{name}</strong>{options}</td>"
            f"<td style='text-align:center'>{qty}</td>"
            f"<td style='text-align:right'>{_money(price)}</td>"
            f"{total_cell}"
            f"</tr>"
        )
    return ''.join(rows)


def _shipping_address_html(address):
    if isinstance(address, str):
        try:
            address = json.loads(address or '{}')
        except Exception:
            address = {}
    address = address or {}
    line1 = clean_text(address.get('line1') or address.get('address'), 160)
    landmark = clean_text(address.get('landmark'), 120)
    city = clean_text(address.get('city'), 80)
    state = clean_text(address.get('state'), 80)
    zip_code = clean_text(address.get('pin') or address.get('zip'), 20)
    first = h(line1) + (f", {h(landmark)}" if landmark else '')
    second = ', '.join([x for x in [city, state] if x])
    second = f"{h(second)} {h(zip_code)}".strip()
    if not first and not second:
        return "<span class='muted'>Not provided</span>"
    return f"{first}<br>{second}<br>United States"


def _support_email_block():
    return """
<div class="note" style="margin-top:18px">
  <strong>Need help?</strong><br>
  Reply to this email or contact <a href="mailto:contact@adhyashaktishop.com" style="color:#1D5C4A">contact@adhyashaktishop.com</a>.
  Please include your order number so we can help faster.
</div>"""


def email_admin_new_order(order_num, customer_name, customer_email, items_data, total, shipping_address=None):
    rows = _order_item_rows(items_data)
    html = _email_html(f"""
<h2>New Order Received</h2>
<div class="note">
  <strong>Order:</strong> {h(order_num)}<br>
  <strong>Customer:</strong> {h(customer_name)}<br>
  <strong>Email:</strong> <a href="mailto:{h(customer_email)}" style="color:#1D5C4A">{h(customer_email)}</a><br>
  <strong>Total:</strong> {_money(total)}
</div>
<h3>Items</h3>
<table class="items">
  <thead><tr><th>Item</th><th style="text-align:center">Qty</th><th style="text-align:right">Price</th><th style="text-align:right">Line Total</th></tr></thead>
  <tbody>{rows}</tbody>
</table>
<h3>Shipping Address</h3>
<p style="line-height:1.7">{_shipping_address_html(shipping_address)}</p>
<a href="https://adhyashaktishop.com/admin/orders" class="cta">View in Admin</a>""", f'New paid order {order_num} for {_money(total)}.')
    send_email(ORDER_MAIL_USER, ORDER_MAIL_PASS, ORDER_MAIL_USER,
               f'New Order - {order_num} | Adhya Shakti Shop', html)


def email_admin_return_request(order):
    reason = clean_text(order.get('return_reason'), 1000) if isinstance(order, dict) else ''
    html = _email_html(f"""
<h2>Return Request Submitted</h2>
<div class="warning">
  Review the returned package before processing a refund. Use the admin Process Return button after the package arrives.
</div>
<p><strong>Order:</strong> {h(order.get('order_number'))}<br>
<strong>Customer:</strong> {h(order.get('customer_name'))}<br>
<strong>Email:</strong> <a href="mailto:{h(order.get('customer_email'))}" style="color:#1D5C4A">{h(order.get('customer_email'))}</a><br>
<strong>Total:</strong> {_money(order.get('total'))}</p>
<h3>Customer Reason</h3>
<p style="white-space:pre-wrap">{h(reason) or 'No reason provided.'}</p>
<a href="https://adhyashaktishop.com/admin/orders?status=return_requested" class="cta">Review Return Requests</a>""", f'Return request for order {order.get("order_number")}.')
    send_email(ORDER_MAIL_USER, ORDER_MAIL_PASS, ORDER_MAIL_USER,
               f'Return Request - {order.get("order_number")} | Adhya Shakti Shop', html)


def email_bulk_order_notification(name, business, cust_email, phone, product_type, quantity, needed_by, message):
    shop_html = _email_html(f"""
<h2>New Bulk Order Request</h2>
<p><strong>Name:</strong> {h(name)}<br>
<strong>Business:</strong> {h(business) or 'Not provided'}<br>
<strong>Email:</strong> <a href="mailto:{h(cust_email)}">{h(cust_email)}</a><br>
<strong>Phone:</strong> {h(phone) or 'Not provided'}</p>
<p><strong>Product Type:</strong> {h(product_type) or 'Not specified'}<br>
<strong>Quantity:</strong> {h(quantity) or 'Not specified'}<br>
<strong>Needed By:</strong> {h(needed_by) or 'Not specified'}</p>
<hr style="border:none;border-top:1px solid #eee;margin:14px 0">
<p style="white-space:pre-wrap">{h(message) or 'No additional notes'}</p>""")
    send_email(CONTACT_MAIL_USER, CONTACT_MAIL_PASS, CONTACT_MAIL_USER,
               f'Bulk Order Request from {h(name)} — Adhya Shakti Shop', shop_html)

    business_row = f'<tr><td style="background:#f9f9f9;font-weight:600">Business</td><td>{h(business)}</td></tr>' if business else ''
    reply_html = _email_html(f"""
<h2>Bulk Inquiry Received</h2>
<p>Hi {h(name)},</p>
<p>Thank you for reaching out to Adhya Shakti Shop. We received your bulk inquiry and will review the details carefully before sending pricing and next steps.</p>
<div class="note">
  <strong>No payment is due right now.</strong><br>
  This message confirms your inquiry only. Your order is not final until you review and approve a quote.
</div>
<h3>Request Summary</h3>
<table class="items" style="margin:14px 0">
  <tbody>
    <tr><td style="width:40%;background:#f9f9f9;font-weight:600">Product</td><td>{h(product_type) or 'Not specified'}</td></tr>
    <tr><td style="background:#f9f9f9;font-weight:600">Quantity</td><td>{h(quantity) or 'Not specified'}</td></tr>
    <tr><td style="background:#f9f9f9;font-weight:600">Needed By</td><td>{h(needed_by) or 'Flexible'}</td></tr>
    {business_row}
  </tbody>
</table>
<h3>What Happens Next</h3>
<ol style="color:#444;line-height:1.9;padding-left:20px;margin:0 0 16px">
  <li>We will review your quantity, product type, timeline, and any design notes.</li>
  <li>We will reply within <strong>1-2 business days</strong> with pricing, timeline, and questions if anything is missing.</li>
  <li>No payment is required until you approve the final quote.</li>
</ol>
<div class="warning">
  <strong>Have artwork ready?</strong><br>
  Send it to <a href="mailto:contact@adhyashaktishop.com" style="color:#1D5C4A">contact@adhyashaktishop.com</a> — we accept PNG, JPG, PDF, AI, or EPS files at 150 DPI or higher. The sooner we have it, the faster we can finalize your quote.
</div>
<p>Questions in the meantime? Reply to this email or message us on WhatsApp.</p>
<div style="text-align:center;margin:24px 0 8px">
  <a href="https://wa.me/c/18483363769" class="cta">Message Us on WhatsApp</a>
</div>
<p class="muted">Adhya Shakti Shop is based in New Jersey. We handle bulk requests carefully so pricing, timing, and design expectations are clear before production begins.</p>""", 'Your bulk inquiry was received. We will reply within 1-2 business days.')
    send_email(CONTACT_MAIL_USER, CONTACT_MAIL_PASS, cust_email,
               'We\'ve Got Your Bulk Order Request — Here\'s What\'s Next | Adhya Shakti Shop', reply_html)


def email_order_confirmation(order_num, customer_name, customer_email, items_data,
                              subtotal, discount, shipping, total, shipping_address=None):
    rows = _order_item_rows(items_data)
    disc_row = (f"<tr><td colspan='3' style='color:#15803d'>Discount</td>"
                f"<td style='text-align:right;color:#15803d'>-{_money(discount)}</td></tr>") if discount else ''
    ship_cell = 'FREE' if shipping == 0 else _money(shipping)
    html = _email_html(f"""
<h2>Order Confirmed</h2>
<p>Hi {h(customer_name)},</p>
<p>Thank you for your order. Your payment was received securely, and we are getting everything ready.</p>
<div class="note">
  <strong>Order number:</strong> {h(order_num)}<br>
  <strong>Total charged:</strong> {_money(total)}<br>
  <strong>Processing time:</strong> 1-3 business days before shipping
</div>
<h3>Items Ordered</h3>
<table class="items">
  <thead><tr><th>Item</th><th style="text-align:center">Qty</th><th style="text-align:right">Price</th><th style="text-align:right">Line Total</th></tr></thead>
  <tbody>{rows}</tbody>
  <tfoot style="background:#f9f9f9">
    <tr><td colspan="3">Subtotal</td><td style="text-align:right">{_money(subtotal)}</td></tr>
    {disc_row}
    <tr><td colspan="3">Shipping</td><td style="text-align:right">{ship_cell}</td></tr>
    <tr style="font-weight:700"><td colspan="3">Total Charged</td><td style="text-align:right">{_money(total)}</td></tr>
  </tfoot>
</table>
<h3>Shipping Address</h3>
<p style="line-height:1.7">{_shipping_address_html(shipping_address)}</p>
<p>We will email you again when the order status changes. You can also check your order anytime from your account.</p>
<div style="text-align:center;margin:26px 0 10px">
  <a href="https://adhyashaktishop.com/dashboard" class="cta">View My Orders</a>
  <a href="https://adhyashaktishop.com/track-order" class="cta secondary" style="margin-left:8px">Track Order</a>
</div>
{_support_email_block()}
<p class="muted">If anything looks incorrect, contact us as soon as possible so we can help before the order ships.</p>""", f'Order {order_num} is confirmed. Total charged: {_money(total)}.')
    send_email(ORDER_MAIL_USER, ORDER_MAIL_PASS, customer_email,
               f'Order Confirmed - {order_num} | Adhya Shakti Shop', html)


_STATUS_COPY = {
    'processing':       ('Your Order Is Being Prepared', 'Your order is now being prepared. We will email you again when it ships.'),
    'shipped':          ('Your Order Has Shipped', 'Your order is on the way. Tracking may take a little time to update after the carrier receives the package.'),
    'delivered':        ('Your Order Was Delivered', 'Your order has been marked as delivered. We hope you love it.'),
    'cancelled':        ('Your Order Was Cancelled', 'Your order has been cancelled. If a refund applies, it normally appears on your card within 5-7 business days after processing.'),
    'return_requested': ('Return Request Received', 'We received your return request. Please ship the package back within 7 business days. We inspect returned items before processing refunds.'),
    'return_received':  ('Return Package Received', 'We received your returned package. We will complete the refund review after inspection.'),
    'refunded':         ('Refund Processed', 'Your refund has been issued. It normally appears on your card within 5-7 business days, depending on your bank.'),
}


def email_order_status(order_num, customer_name, customer_email, status, tracking_number=None):
    copy = _STATUS_COPY.get(status)
    if not copy:
        return
    subject_line, body_text = copy
    tracking_panel = (f"""
<div class="note">
  <strong>Tracking number:</strong> {h(tracking_number)}<br>
  Tracking may take a little time to appear after the carrier scan.
</div>""" if tracking_number and status == 'shipped' else '')
    html = _email_html(f"""
<h2>{subject_line}</h2>
<p>Hi {h(customer_name)},</p>
<div class="note">
  <strong>Order:</strong> {h(order_num)}<br>
  <strong>Status:</strong> {h(status.replace('_', ' ').title())}
</div>
{tracking_panel}
<p>{body_text}</p>
<div style="text-align:center;margin:24px 0 8px">
  <a href="https://adhyashaktishop.com/dashboard" class="cta">View My Orders</a>
  <a href="https://adhyashaktishop.com/track-order" class="cta secondary" style="margin-left:8px">Track Order</a>
</div>
{_support_email_block()}""", f'Order {order_num} status update: {subject_line}.')
    send_email(ORDER_MAIL_USER, ORDER_MAIL_PASS, customer_email,
               f'{subject_line} - {order_num} | Adhya Shakti Shop', html)


def _review_request_item_rows(items_data):
    rows = []
    seen = set()
    for item in items_data or []:
        product_id = clean_text(item.get('id'), 80)
        if not product_id or product_id in seen:
            continue
        seen.add(product_id)
        product_name = item.get('name') or 'Purchased item'
        review_url = f'https://adhyashaktishop.com/product/{quote(product_id)}'
        variation = clean_text(item.get('variation'), 100)
        rows.append(f"""
<tr>
  <td>
    <strong>{h(product_name)}</strong>
    {f'<div class="muted" style="font-size:.82rem;margin-top:3px">{h(variation)}</div>' if variation else ''}
  </td>
  <td style="text-align:right">
    <a href="{h(review_url)}" style="color:#1D5C4A;font-weight:700;text-decoration:none">Review</a>
  </td>
</tr>""")
    return ''.join(rows)


def email_review_request(order_num, customer_name, customer_email, items_data):
    item_rows = _review_request_item_rows(items_data)
    product_table = f"""
<h3>Review Your Item(s)</h3>
<table class="items">
  <thead><tr><th>Product</th><th style="text-align:right">Action</th></tr></thead>
  <tbody>{item_rows}</tbody>
</table>""" if item_rows else ''
    html = _email_html(f"""
<h2>How Was Your Order?</h2>
<p>Hi {h(customer_name)},</p>
<p>Your order <strong>{h(order_num)}</strong> was marked delivered. We hope everything arrived safely and you love your purchase.</p>
<div class="note">
  Your honest review helps other shoppers feel confident and helps our small shop improve.
</div>
{product_table}
<div style="text-align:center;margin:28px 0">
  <a href="https://adhyashaktishop.com/dashboard" class="cta">Open My Orders</a>
</div>
<p class="muted">If something is wrong with the order, please reply to this email instead of leaving a review first. We want a chance to make it right.</p>
{_support_email_block()}""", f'Order {order_num} was delivered. Share your review when you have a moment.')
    send_email(ORDER_MAIL_USER, ORDER_MAIL_PASS, customer_email,
               f'How was your order? - {order_num} | Adhya Shakti Shop', html)


def validate_password(password):
    if len(password) < 8:
        return 'Password must be at least 8 characters'
    if not re.search(r'[A-Z]', password):
        return 'Password must contain at least one uppercase letter'
    if not re.search(r'[a-z]', password):
        return 'Password must contain at least one lowercase letter'
    if not re.search(r'[0-9]', password):
        return 'Password must contain at least one number'
    if not re.search(r'[^A-Za-z0-9]', password):
        return 'Password must contain at least one special character (e.g. !@#$)'
    return None


RESERVED_NAME_WORDS = ('admin', 'staff')
EMAIL_RE = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,24}$')
PHONE_RE = re.compile(r'^[0-9+().\-\s]{7,30}$')


def validate_name(name):
    lname = name.lower()
    for word in RESERVED_NAME_WORDS:
        if word in lname:
            return f"Name cannot contain the word '{word}'"
    return None


def normalize_public_email(value):
    email = clean_text(value, 140).lower()
    return email if EMAIL_RE.fullmatch(email or '') else ''


def normalize_public_phone(value):
    phone = clean_text(value, 30)
    if not phone:
        return ''
    digits = re.sub(r'\D+', '', phone)
    if len(digits) < 7 or len(digits) > 15 or not PHONE_RE.fullmatch(phone):
        return ''
    return phone


def public_text_looks_spammy(*values):
    text = '\n'.join(clean_text(v, 2000, strip=False) for v in values if v)
    lowered = text.lower()
    if any(marker in lowered for marker in ('<script', '</script', 'javascript:', '<?php', '<iframe', '</iframe')):
        return True
    links = re.findall(r'(?:https?://|www\.)', lowered)
    return len(links) > 3


# ─── Database ────────────────────────────────────────────────────────────────

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
    return g.db


def request_ip():
    forwarded = (request.headers.get('X-Forwarded-For') or '').split(',')[0].strip()
    return forwarded or request.headers.get('X-Real-IP') or request.remote_addr or ''


def log_security_event(event_type, severity='info', message='', *, user_id=None, email=None, metadata=None):
    try:
        db = get_db()
        db.execute(
            """INSERT INTO security_events
               (id,event_type,severity,user_id,email,ip,user_agent,path,method,message,metadata)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (
                str(uuid.uuid4()),
                clean_text(event_type, 80),
                clean_text(severity, 20) or 'info',
                clean_text(user_id, 80) or None,
                clean_text(email, 140).lower() or None,
                clean_text(request_ip(), 80),
                clean_text(request.headers.get('User-Agent'), 300),
                clean_text(request.path, 200),
                clean_text(request.method, 12),
                clean_text(message, 500),
                json.dumps(metadata or {}, default=str)[:2000],
            )
        )
        db.commit()
    except Exception as exc:
        try:
            app.logger.warning('security event logging failed: %s', exc)
        except Exception:
            pass  # nosec B110


def public_abuse_guard(action, *, email='', fingerprint='', ip_limit=30, email_limit=None, fingerprint_limit=None, window_seconds=900):
    action = clean_text(action, 80)
    ip = clean_text(request_ip(), 80)
    email = normalize_public_email(email) if email else ''
    raw_fingerprint = clean_text(fingerprint, 200).lower()
    fingerprint_hash = hashlib.sha256(raw_fingerprint.encode('utf-8')).hexdigest() if raw_fingerprint else ''
    seconds = max(60, int(window_seconds or 900))
    window_arg = f'-{seconds} seconds'

    def too_many(db, field, value, limit):
        if not limit or not value:
            return False
        row = db.execute(
            f"SELECT COUNT(*) AS n FROM public_endpoint_attempts WHERE action=? AND {field}=? AND created_at >= datetime('now', ?)",  # nosec B608
            (action, value, window_arg)
        ).fetchone()
        return bool(row and row['n'] >= int(limit))

    try:
        db = get_db()
        db.execute("DELETE FROM public_endpoint_attempts WHERE created_at < datetime('now', '-2 days')")
        limited_by = None
        if too_many(db, 'ip', ip, ip_limit):
            limited_by = 'ip'
        elif too_many(db, 'email', email, email_limit):
            limited_by = 'email'
        elif too_many(db, 'fingerprint', fingerprint_hash, fingerprint_limit):
            limited_by = 'fingerprint'

        if limited_by:
            db.commit()
            log_security_event(
                'public_endpoint_rate_limited',
                'warning',
                'Public endpoint rate limit triggered',
                email=email or None,
                metadata={
                    'action': action,
                    'limited_by': limited_by,
                    'window_seconds': seconds,
                    'path': request.path,
                },
            )
            return jsonify({'error': 'Too many attempts. Please wait and try again.'}), 429

        db.execute(
            "INSERT INTO public_endpoint_attempts (id,action,ip,email,fingerprint) VALUES (?,?,?,?,?)",
            (str(uuid.uuid4()), action, ip, email or None, fingerprint_hash or None)
        )
        db.commit()
    except Exception as exc:
        try:
            app.logger.warning('public abuse guard failed open: %s', exc)
        except Exception:
            pass  # nosec B110
    return None


@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(exc):
    if request.path.startswith('/api/'):
        try:
            log_security_event(
                'upload_rejected',
                'warning',
                'Request body exceeded the upload limit',
                metadata={
                    'path': request.path,
                    'max_bytes': app.config.get('MAX_CONTENT_LENGTH'),
                    'error': str(exc)[:300],
                },
            )
        except Exception:
            pass  # nosec B110
        return jsonify({'error': 'File is too large. Maximum request size is 10 MB.'}), 413
    return ('File is too large', 413)


def _audit_json(value, limit=6000):
    if value in (None, ''):
        return None
    try:
        return json.dumps(value, default=str, sort_keys=True)[:limit]
    except Exception:
        return json.dumps(str(value))[:limit]


def _infer_audit_entity(action, metadata):
    metadata = metadata or {}
    entity_type = metadata.get('entity_type')
    entity_id = metadata.get('entity_id')
    if entity_type and entity_id:
        return clean_text(entity_type, 80), clean_text(entity_id, 120)
    for key, inferred in (
        ('order_id', 'order'),
        ('product_id', 'product'),
        ('category_id', 'category'),
        ('coupon_id', 'coupon'),
        ('user_id', 'user'),
        ('target_user_id', 'user'),
        ('customer_id', 'customer'),
        ('vendor_id', 'vendor'),
        ('sale_id', 'sale'),
        ('expense_id', 'expense'),
        ('purchase_id', 'purchase'),
        ('item_id', 'inventory_item'),
        ('attachment_id', 'attachment'),
        ('setting_key', 'setting'),
    ):
        if metadata.get(key):
            return inferred, clean_text(metadata.get(key), 120)
    return clean_text(entity_type or (action.split('_', 1)[0] if action else 'system'), 80), clean_text(entity_id or '', 120)


def log_audit_event(action, entity_type='', entity_id='', message='', *, before=None, after=None, metadata=None):
    """Durable admin/staff audit trail for business changes."""
    try:
        user = getattr(g, 'user', {}) or {}
        metadata = dict(metadata or {})
        inferred_type, inferred_id = _infer_audit_entity(action, metadata)
        entity_type = clean_text(entity_type or inferred_type, 80)
        entity_id = clean_text(entity_id or inferred_id, 120)
        db = get_db()
        db.execute(
            """INSERT INTO admin_audit_log
               (id,actor_id,actor_email,actor_role,action,entity_type,entity_id,message,
                before_json,after_json,metadata,ip,user_agent)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                str(uuid.uuid4()),
                clean_text(user.get('id'), 80) or None,
                clean_text(user.get('email'), 140).lower() or None,
                clean_text(user.get('role'), 30) or None,
                clean_text(action, 100),
                entity_type or None,
                entity_id or None,
                clean_text(message, 500),
                _audit_json(before),
                _audit_json(after),
                _audit_json({k: v for k, v in metadata.items() if k not in ('before', 'after')}, 3000) or '{}',
                clean_text(request_ip(), 80),
                clean_text(request.headers.get('User-Agent'), 300),
            )
        )
        db.commit()
    except Exception as exc:
        try:
            app.logger.warning('audit logging failed: %s', exc)
        except Exception:
            pass  # nosec B110


def log_admin_action(action, message='', metadata=None):
    user = getattr(g, 'user', {}) or {}
    log_security_event(
        action,
        'info',
        message,
        user_id=user.get('id'),
        email=user.get('email'),
        metadata=metadata or {},
    )
    log_audit_event(action, message=message, metadata=metadata or {})


def fetch_order_for_audit(db, oid):
    row = db.execute(
        "SELECT id,order_number,status,payment_status,tracking_number,total,review_requested_at,updated_at FROM orders WHERE id=?",
        (oid,)
    ).fetchone()
    return dict(row) if row else None


def fetch_user_for_audit(db, uid):
    row = db.execute(
        "SELECT id,name,email,phone,role,created_at FROM users WHERE id=?",
        (uid,)
    ).fetchone()
    return dict(row) if row else None


@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            phone TEXT,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'customer',
            address TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS categories (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS products (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT,
            price REAL NOT NULL,
            compare_price REAL,
            category_id TEXT,
            stock INTEGER DEFAULT 0,
            sku TEXT,
            images TEXT DEFAULT '[]',
            variations TEXT DEFAULT '[]',
            is_active INTEGER DEFAULT 1,
            allow_custom_print INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (category_id) REFERENCES categories(id)
        );

        CREATE TABLE IF NOT EXISTS coupons (
            id TEXT PRIMARY KEY,
            code TEXT UNIQUE NOT NULL,
            discount_type TEXT NOT NULL,
            discount_value REAL NOT NULL,
            min_order REAL DEFAULT 0,
            max_uses INTEGER,
            used_count INTEGER DEFAULT 0,
            expires_at TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS orders (
            id TEXT PRIMARY KEY,
            order_number TEXT UNIQUE NOT NULL,
            user_id TEXT,
            customer_name TEXT NOT NULL,
            customer_email TEXT NOT NULL,
            customer_phone TEXT,
            shipping_address TEXT NOT NULL,
            items TEXT NOT NULL,
            subtotal REAL NOT NULL,
            discount REAL DEFAULT 0,
            coupon_code TEXT,
            shipping_charge REAL DEFAULT 0,
            total REAL NOT NULL,
            payment_method TEXT DEFAULT 'cod',
            payment_status TEXT DEFAULT 'pending',
            payment_intent_id TEXT,
            status TEXT DEFAULT 'pending',
            tracking_number TEXT,
            notes TEXT,
            return_reason TEXT,
            review_requested_at TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS reviews (
            id TEXT PRIMARY KEY,
            product_id TEXT NOT NULL,
            user_id TEXT NOT NULL,
            rating INTEGER NOT NULL,
            comment TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS user_wishlist (
            user_id TEXT NOT NULL,
            product_id TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            PRIMARY KEY (user_id, product_id),
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS sliders (
            id TEXT PRIMARY KEY,
            title TEXT,
            subtitle TEXT,
            image_url TEXT,
            link TEXT,
            sort_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );

        CREATE TABLE IF NOT EXISTS product_variants (
            id TEXT PRIMARY KEY,
            product_id TEXT NOT NULL,
            color TEXT NOT NULL,
            size TEXT NOT NULL,
            stock INTEGER DEFAULT 0,
            FOREIGN KEY (product_id) REFERENCES products(id)
        );

        CREATE TABLE IF NOT EXISTS contact_messages (
            id TEXT PRIMARY KEY,
            name TEXT, email TEXT, phone TEXT, message TEXT,
            inquiry_type TEXT, order_number TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS bulk_orders (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            business_name TEXT,
            email TEXT NOT NULL,
            phone TEXT,
            product_type TEXT,
            quantity TEXT,
            needed_by TEXT,
            message TEXT,
            file_url TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS password_resets (
            id TEXT PRIMARY KEY,
            email TEXT NOT NULL,
            token TEXT NOT NULL UNIQUE,
            expires_at TEXT NOT NULL,
            used INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS auth_login_failures (
            id TEXT PRIMARY KEY,
            email TEXT NOT NULL,
            ip TEXT NOT NULL,
            portal TEXT DEFAULT 'customer',
            role TEXT,
            reason TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS public_endpoint_attempts (
            id TEXT PRIMARY KEY,
            action TEXT NOT NULL,
            ip TEXT NOT NULL,
            email TEXT,
            fingerprint TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS newsletter_subscribers (
            id TEXT PRIMARY KEY,
            email TEXT UNIQUE NOT NULL,
            subscribed_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS welcome_discounts (
            id TEXT PRIMARY KEY,
            email TEXT UNIQUE NOT NULL,
            user_id TEXT,
            coupon_code TEXT DEFAULT 'WELCOME10',
            issued_via TEXT,
            issued_at TEXT DEFAULT (datetime('now')),
            emailed_at TEXT,
            used_at TEXT,
            used_order_id TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS back_in_stock_requests (
            id TEXT PRIMARY KEY,
            product_id TEXT NOT NULL,
            email TEXT NOT NULL,
            name TEXT,
            user_id TEXT,
            status TEXT DEFAULT 'pending',
            request_count INTEGER DEFAULT 1,
            source TEXT DEFAULT 'product_page',
            created_at TEXT DEFAULT (datetime('now')),
            last_requested_at TEXT DEFAULT (datetime('now')),
            notified_at TEXT,
            FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL,
            UNIQUE (product_id, email)
        );

        CREATE TABLE IF NOT EXISTS abandoned_carts (
            id TEXT PRIMARY KEY,
            email TEXT NOT NULL UNIQUE,
            name TEXT,
            phone TEXT,
            user_id TEXT,
            items TEXT NOT NULL DEFAULT '[]',
            subtotal REAL DEFAULT 0,
            discount REAL DEFAULT 0,
            shipping REAL DEFAULT 0,
            total REAL DEFAULT 0,
            coupon_code TEXT,
            payment_intent_id TEXT,
            recovery_token TEXT,
            unsubscribed INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            reminded_at TEXT,
            converted_at TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS security_events (
            id TEXT PRIMARY KEY,
            event_type TEXT NOT NULL,
            severity TEXT DEFAULT 'info',
            user_id TEXT,
            email TEXT,
            ip TEXT,
            user_agent TEXT,
            path TEXT,
            method TEXT,
            message TEXT,
            metadata TEXT DEFAULT '{}',
            reviewed_at TEXT,
            reviewed_by TEXT,
            review_note TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS admin_audit_log (
            id TEXT PRIMARY KEY,
            actor_id TEXT,
            actor_email TEXT,
            actor_role TEXT,
            action TEXT NOT NULL,
            entity_type TEXT,
            entity_id TEXT,
            message TEXT,
            before_json TEXT,
            after_json TEXT,
            metadata TEXT DEFAULT '{}',
            ip TEXT,
            user_agent TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS admin_trusted_ips (
            id TEXT PRIMARY KEY,
            ip TEXT UNIQUE NOT NULL,
            label TEXT NOT NULL,
            note TEXT,
            is_active INTEGER DEFAULT 1,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
    ''')

    # Column migrations (safe — ignored if column already exists)
    for migration in [
        "ALTER TABLE products ADD COLUMN is_bestseller INTEGER DEFAULT 0",
        "ALTER TABLE products ADD COLUMN cost_price REAL DEFAULT 0",
        "ALTER TABLE products ADD COLUMN low_stock_threshold REAL DEFAULT 5",
        "ALTER TABLE reviews ADD COLUMN images TEXT DEFAULT '[]'",
    ]:
        try:
            db.execute(migration)
            db.commit()
        except Exception:
            pass  # nosec B110

    # Seed admin user
    admin_exists = db.execute("SELECT id FROM users WHERE role='admin'").fetchone()
    if not admin_exists:
        seed_pw = os.environ.get('DEFAULT_ADMIN_PASSWORD', '')
        if seed_pw:
            pw_error = validate_password(seed_pw)
            if pw_error:
                raise RuntimeError(f"DEFAULT_ADMIN_PASSWORD is not strong enough: {pw_error}")
            seed_email = os.environ.get('DEFAULT_ADMIN_EMAIL', 'admin@adhyashaktishop.com').strip().lower()
            seed_name = os.environ.get('DEFAULT_ADMIN_NAME', 'Admin').strip() or 'Admin'
            pwd = bcrypt.hashpw(seed_pw.encode(), bcrypt.gensalt()).decode()
            db.execute("INSERT INTO users (id,name,email,phone,password,role) VALUES (?,?,?,?,?,?)",
                       (str(uuid.uuid4()), seed_name, seed_email, '', pwd, 'admin'))
            db.commit()
        else:
            app.logger.warning(
                "No admin user exists and DEFAULT_ADMIN_PASSWORD is not set; "
                "create the first admin manually or set DEFAULT_ADMIN_PASSWORD before startup."
            )

    # Seed sample categories
    if not db.execute("SELECT id FROM categories LIMIT 1").fetchone():
        cats = [
            ('Jewelry',  'Handcrafted jewelry inspired by Indian heritage'),
            ('Clothing', 'Custom printed apparel and fashion'),
            ('Custom',   'Personalized custom-made items'),
        ]
        for name, desc in cats:
            db.execute("INSERT INTO categories (id,name,description) VALUES (?,?,?)",
                       (str(uuid.uuid4()), name, desc))
    else:
        # Ensure all required categories exist in existing databases
        required_cats = [
            ('Clothing',     'Ready-made printed clothing'),
            ('Custom',       'Personalized custom-made items'),
            ('T-Shirts',     'Custom printed t-shirts'),
            ('Polo Shirts',  'Custom printed polo shirts'),
            ('Hoodies',      'Custom printed hoodies'),
        ]
        for cat_name, cat_desc in required_cats:
            if not db.execute("SELECT id FROM categories WHERE name=?", (cat_name,)).fetchone():
                db.execute("INSERT INTO categories (id,name,description) VALUES (?,?,?)",
                           (str(uuid.uuid4()), cat_name, cat_desc))

    # Seed sample products
    if not db.execute("SELECT id FROM products LIMIT 1").fetchone():
        j_id = db.execute("SELECT id FROM categories WHERE name='Jewelry'").fetchone()['id']
        c_id = db.execute("SELECT id FROM categories WHERE name='Clothing'").fetchone()['id']
        products = [
            ('Lotus Jhumka Earrings', 'Elegant handcrafted jhumka earrings with lotus motif in oxidized silver finish', 29.99, 49.99, j_id, 25),
            ('Gold Lotus Necklace', 'Delicate gold-tone necklace with lotus pendant, perfect for everyday wear', 49.99, 79.99, j_id, 20),
            ('Lotus Print Tee', 'Premium cotton t-shirt with custom lotus print — available in multiple colors', 24.99, 34.99, c_id, 50),
            ('Custom Hoodie', 'Cozy heavyweight hoodie with personalized Adhya Shakti print', 44.99, 59.99, c_id, 30),
        ]
        for p in products:
            db.execute("INSERT INTO products (id,name,description,price,compare_price,category_id,stock) VALUES (?,?,?,?,?,?,?)",
                       (str(uuid.uuid4()),) + p)

    # Seed sliders
    if not db.execute("SELECT id FROM sliders LIMIT 1").fetchone():
        sliders = [
            ('Handcrafted Jewelry', 'Elegant pieces inspired by Indian heritage', '', '/products', 1),
            ('Custom Printed Clothing', 'Wear your story — unique prints, premium quality', '', '/products', 2),
            ('New Arrivals', 'Explore our latest collection', '', '/products', 3),
        ]
        for s in sliders:
            db.execute("INSERT INTO sliders (id,title,subtitle,image_url,link,sort_order) VALUES (?,?,?,?,?,?)",
                       (str(uuid.uuid4()),) + s)

    # Seed settings
    defaults = {'shop_name': 'Adhya Shakti Shop', 'shop_email': 'contact@adhyashaktishop.com',
                'shop_phone': '', 'shop_address': 'New Jersey, USA'}
    for k, v in defaults.items():
        db.execute("INSERT OR IGNORE INTO settings (key,value) VALUES (?,?)", (k, v))
    # This shop uses Stripe in the US. Remove stale Razorpay settings if they
    # exist from older builds so unused payment keys are not kept around.
    db.execute("DELETE FROM settings WHERE LOWER(key) LIKE 'razorpay%'")

    # Seed WELCOME10 coupon (10% off first order, no expiry, unlimited uses)
    if not db.execute("SELECT id FROM coupons WHERE code='WELCOME10'").fetchone():
        db.execute(
            "INSERT INTO coupons (id,code,discount_type,discount_value,min_order,is_active) VALUES (?,?,?,?,?,?)",
            (str(uuid.uuid4()), 'WELCOME10', 'percent', 10, 0, 1)
        )

    # ── Migrations: add new columns to existing databases safely ─────────────
    migrations = [
        "ALTER TABLE orders ADD COLUMN payment_intent_id TEXT",
        "ALTER TABLE orders ADD COLUMN review_requested_at TEXT",
        "ALTER TABLE orders ADD COLUMN return_reason TEXT",
        "ALTER TABLE products ADD COLUMN allow_custom_print INTEGER DEFAULT 0",
        "ALTER TABLE contact_messages ADD COLUMN inquiry_type TEXT",
        "ALTER TABLE contact_messages ADD COLUMN order_number TEXT",
        "ALTER TABLE users ADD COLUMN token_version INTEGER DEFAULT 0",
        "ALTER TABLE security_events ADD COLUMN reviewed_at TEXT",
        "ALTER TABLE security_events ADD COLUMN reviewed_by TEXT",
        "ALTER TABLE security_events ADD COLUMN review_note TEXT",
    ]
    for sql in migrations:
        try:
            db.execute(sql)
        except Exception:
            pass  # nosec B110
    try:
        db.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_orders_payment_intent_id
            ON orders(payment_intent_id)
            WHERE payment_intent_id IS NOT NULL AND payment_intent_id != ''
        """)
    except Exception as exc:
        app.logger.warning("Could not create payment-intent uniqueness index: %s", exc)
    try:
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_auth_login_failures_lookup
            ON auth_login_failures(email, ip, portal, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_auth_login_failures_ip
            ON auth_login_failures(ip, portal, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_public_endpoint_attempts_ip
            ON public_endpoint_attempts(action, ip, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_public_endpoint_attempts_email
            ON public_endpoint_attempts(action, email, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_public_endpoint_attempts_fingerprint
            ON public_endpoint_attempts(action, fingerprint, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_welcome_discounts_user
            ON welcome_discounts(user_id)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_welcome_discounts_used
            ON welcome_discounts(used_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_back_in_stock_product_status
            ON back_in_stock_requests(product_id, notified_at, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_back_in_stock_email
            ON back_in_stock_requests(email, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_abandoned_carts_pending
            ON abandoned_carts(converted_at, reminded_at, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_admin_audit_log_created
            ON admin_audit_log(created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_admin_audit_log_entity
            ON admin_audit_log(entity_type, entity_id, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_security_events_created
            ON security_events(created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_security_events_reviewed
            ON security_events(reviewed_at, severity, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_security_events_type_created
            ON security_events(event_type, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_security_events_ip_created
            ON security_events(ip, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_admin_trusted_ips_active
            ON admin_trusted_ips(is_active, ip)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_user_wishlist_user_created
            ON user_wishlist(user_id, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_orders_user_created
            ON orders(user_id, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_orders_status_created
            ON orders(status, created_at)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_products_category_active
            ON products(category_id, is_active)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_product_variants_product
            ON product_variants(product_id)
        """)
        db.execute("""
            CREATE INDEX IF NOT EXISTS idx_reviews_product_created
            ON reviews(product_id, created_at)
        """)
    except Exception as exc:
        app.logger.warning("Could not create security indexes: %s", exc)

    # ── Data cleanup ──────────────────────────────────────────────────────────
    db.execute("DELETE FROM categories WHERE name='Customized Clothing'")
    try:
        db.execute("""
            DELETE FROM product_variants
            WHERE NOT EXISTS (SELECT 1 FROM products p WHERE p.id=product_variants.product_id)
        """)
        db.execute("""
            DELETE FROM user_wishlist
            WHERE NOT EXISTS (SELECT 1 FROM users u WHERE u.id=user_wishlist.user_id)
               OR NOT EXISTS (SELECT 1 FROM products p WHERE p.id=user_wishlist.product_id)
        """)
        db.execute("""
            DELETE FROM reviews
            WHERE NOT EXISTS (SELECT 1 FROM users u WHERE u.id=reviews.user_id)
               OR NOT EXISTS (SELECT 1 FROM products p WHERE p.id=reviews.product_id)
        """)
        db.execute("""
            UPDATE products
            SET category_id=NULL
            WHERE category_id IS NOT NULL
              AND NOT EXISTS (SELECT 1 FROM categories c WHERE c.id=products.category_id)
        """)
        db.execute("""
            UPDATE categories
            SET parent_id=NULL
            WHERE parent_id IS NOT NULL
              AND NOT EXISTS (SELECT 1 FROM categories p WHERE p.id=categories.parent_id)
        """)
    except Exception as exc:
        app.logger.warning("Could not complete data consistency cleanup: %s", exc)

    db.commit()
    db.close()


# Call init_db at startup — works for both `python app.py` and WSGI (gunicorn/uWSGI)
init_db()


# ─── Auth helpers ─────────────────────────────────────────────────────────────

def make_auth_token(user_id, email, role, token_version):
    return jwt.encode(
        {
            'id': user_id,
            'email': email,
            'role': role,
            'tv': token_version or 0,
            'exp': datetime.datetime.utcnow() + datetime.timedelta(seconds=AUTH_COOKIE_MAX_AGE),
        },
        SECRET_KEY,
        algorithm='HS256'
    )


def auth_user_payload(user):
    return {'id': user['id'], 'name': user['name'], 'email': user['email'], 'role': user['role']}


def set_auth_cookie(response, token):
    if isinstance(token, bytes):
        token = token.decode('utf-8')
    response.set_cookie(
        AUTH_COOKIE_NAME,
        token,
        max_age=AUTH_COOKIE_MAX_AGE,
        httponly=True,
        secure=AUTH_COOKIE_SECURE,
        samesite='Lax',
        path='/'
    )
    return response


def clear_auth_cookie(response):
    response.delete_cookie(AUTH_COOKIE_NAME, path='/')
    return response


def request_auth_token():
    auth = request.headers.get('Authorization', '')
    if auth.lower().startswith('bearer '):
        return auth.split(' ', 1)[1].strip()
    return request.cookies.get(AUTH_COOKIE_NAME, '')


def normalized_login_portal(portal):
    portal = clean_text(portal or 'customer', 20).lower()
    return portal if portal in ('customer', 'staff') else 'customer'


def login_policy_for(user, portal):
    role = user['role'] if user else ''
    staff_like = portal == 'staff' or role in ('admin', 'staff')
    return {
        'staff_like': staff_like,
        'limit': STAFF_LOGIN_FAILURE_LIMIT if staff_like else CUSTOMER_LOGIN_FAILURE_LIMIT,
        'ip_limit': STAFF_LOGIN_IP_FAILURE_LIMIT if staff_like else None,
        'lockout_seconds': STAFF_LOGIN_LOCKOUT_SECONDS if staff_like else CUSTOMER_LOGIN_LOCKOUT_SECONDS,
    }


def recent_login_failures(db, *, email=None, ip=None, portal=None, seconds=900):
    where = ["created_at >= datetime('now', ?)"]
    params = [f'-{int(seconds)} seconds']
    if email is not None:
        where.append("email=?")
        params.append(email)
    if ip is not None:
        where.append("ip=?")
        params.append(ip)
    if portal is not None:
        where.append("portal=?")
        params.append(portal)
    row = db.execute(
        "SELECT COUNT(*) AS n FROM auth_login_failures WHERE " + " AND ".join(where),  # nosec B608
        params
    ).fetchone()
    return row['n'] if row else 0


def login_lockout_reason(db, email, user, portal):
    ip = clean_text(request_ip(), 80)
    policy = login_policy_for(user, portal)
    seconds = policy['lockout_seconds']
    exact_count = recent_login_failures(db, email=email, ip=ip, portal=portal, seconds=seconds)
    if exact_count >= policy['limit']:
        return 'email_ip'
    if policy['staff_like']:
        ip_count = recent_login_failures(db, ip=ip, portal='staff', seconds=seconds)
        if ip_count >= policy['ip_limit']:
            return 'staff_ip'
    return None


def record_login_failure(db, email, user, portal, reason):
    email = clean_text(email, 140).lower()
    ip = clean_text(request_ip(), 80)
    role = user['role'] if user else None
    db.execute(
        """INSERT INTO auth_login_failures (id,email,ip,portal,role,reason)
           VALUES (?,?,?,?,?,?)""",
        (str(uuid.uuid4()), email, ip, portal, role, clean_text(reason, 80))
    )
    db.execute("DELETE FROM auth_login_failures WHERE created_at < datetime('now', '-2 days')")
    db.commit()
    log_security_event(
        'login_failed',
        'warning',
        'Invalid login attempt',
        user_id=user['id'] if user else None,
        email=email,
        metadata={'portal': portal, 'role': role, 'reason': reason},
    )


def clear_login_failures(db, email, portal):
    db.execute(
        "DELETE FROM auth_login_failures WHERE email=? AND ip=? AND portal=?",
        (clean_text(email, 140).lower(), clean_text(request_ip(), 80), portal)
    )
    db.commit()


def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request_auth_token()
        if not token:
            if request.path.startswith('/api/'):
                log_security_event('auth_required_missing', 'warning', 'Protected API requested without a session')
            return jsonify({'error': 'Token required'}), 401
        try:
            data = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
            # Verify token version and use the current database role/email for
            # authorization decisions, so role changes take effect immediately.
            db = get_db()
            row = db.execute("SELECT email, role, token_version FROM users WHERE id=?", (data.get('id'),)).fetchone()
            if not row or row['token_version'] != data.get('tv', 0):
                log_security_event(
                    'auth_token_rejected',
                    'warning',
                    'Expired or revoked session token used',
                    user_id=data.get('id'),
                    email=data.get('email'),
                    metadata={'reason': 'token_version'},
                )
                return jsonify({'error': 'Session expired. Please log in again.'}), 401
            g.user = {
                'id': data.get('id'),
                'email': row['email'],
                'role': row['role'],
                'tv': row['token_version'],
                'exp': data.get('exp'),
            }
        except jwt.ExpiredSignatureError:
            log_security_event('auth_token_rejected', 'warning', 'Expired session token used', metadata={'reason': 'expired'})
            return jsonify({'error': 'Session expired. Please log in again.'}), 401
        except jwt.InvalidTokenError:
            log_security_event('auth_token_rejected', 'warning', 'Invalid session token used', metadata={'reason': 'invalid'})
            return jsonify({'error': 'Invalid token. Please log in again.'}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """Allows both admin and staff roles."""
    @wraps(f)
    @token_required
    def decorated(*args, **kwargs):
        if g.user.get('role') not in ('admin', 'staff'):
            log_security_event(
                'authorization_denied',
                'warning',
                'Non-staff account attempted staff/admin API access',
                user_id=g.user.get('id'),
                email=g.user.get('email'),
                metadata={'required': 'admin_or_staff', 'role': g.user.get('role')},
            )
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated


def admin_only_required(f):
    """Strictly admin-only — staff cannot access."""
    @wraps(f)
    @token_required
    def decorated(*args, **kwargs):
        if g.user.get('role') != 'admin':
            log_security_event(
                'authorization_denied',
                'warning',
                'Non-admin account attempted admin-only API access',
                user_id=g.user.get('id'),
                email=g.user.get('email'),
                metadata={'required': 'admin', 'role': g.user.get('role')},
            )
            return jsonify({'error': 'Administrator access required'}), 403
        return f(*args, **kwargs)
    return decorated


def row_to_dict(row):
    if row is None:
        return None
    return dict(row)


def rows_to_list(rows):
    return [dict(r) for r in rows]


def current_customer_id_from_request(db):
    token = request_auth_token()
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        row = db.execute(
            "SELECT token_version, role FROM users WHERE id=?",
            (payload.get('id'),)
        ).fetchone()
        if row and row['token_version'] == payload.get('tv', 0) and row['role'] == 'customer':
            return payload.get('id')
    except Exception:
        pass  # nosec B110
    return None


def _welcome_discount_row(db, email='', user_id=None):
    email = normalize_public_email(email) if email else ''
    if email:
        row = db.execute("SELECT * FROM welcome_discounts WHERE email=?", (email,)).fetchone()
        if row:
            return dict(row)
    if user_id:
        row = db.execute("SELECT * FROM welcome_discounts WHERE user_id=?", (user_id,)).fetchone()
        if row:
            return dict(row)
    return None


def _customer_has_paid_order(db, email='', user_id=None):
    email = normalize_public_email(email) if email else ''
    clauses = []
    params = []
    if email:
        clauses.append("LOWER(customer_email)=?")
        params.append(email)
    if user_id:
        clauses.append("user_id=?")
        params.append(user_id)
    if not clauses:
        return False
    sql = (
        "SELECT 1 FROM orders WHERE (" + " OR ".join(clauses) +
        ") AND payment_status IN ('paid','refund_pending','refunded') LIMIT 1"
    )
    return db.execute(sql, params).fetchone() is not None  # nosec B608


def ensure_welcome_discount(db, email, *, user_id=None, issued_via='newsletter', send=False):
    email = normalize_public_email(email)
    if not email:
        return {'issued': False, 'already_issued': False, 'emailed': False, 'code': WELCOME_COUPON_CODE}
    row = _welcome_discount_row(db, email=email, user_id=user_id)
    already_issued = row is not None
    if row:
        if user_id and not row.get('user_id'):
            db.execute("UPDATE welcome_discounts SET user_id=? WHERE id=?", (user_id, row['id']))
            row['user_id'] = user_id
    else:
        db.execute(
            """INSERT INTO welcome_discounts (id,email,user_id,coupon_code,issued_via)
               VALUES (?,?,?,?,?)""",
            (str(uuid.uuid4()), email, user_id, WELCOME_COUPON_CODE, clean_text(issued_via, 40))
        )
        row = {'email': email, 'user_id': user_id, 'emailed_at': None, 'used_at': None}

    emailed = False
    if send and not row.get('emailed_at'):
        email_welcome_discount(email, WELCOME_COUPON_CODE)
        db.execute("UPDATE welcome_discounts SET emailed_at=datetime('now') WHERE email=?", (email,))
        emailed = True
    return {
        'issued': not already_issued,
        'already_issued': already_issued,
        'emailed': emailed,
        'code': WELCOME_COUPON_CODE,
    }


def mark_welcome_discount_emailed(db, email, user_id=None):
    email = normalize_public_email(email) if email else ''
    if email:
        db.execute("UPDATE welcome_discounts SET emailed_at=datetime('now') WHERE email=?", (email,))
    elif user_id:
        db.execute("UPDATE welcome_discounts SET emailed_at=datetime('now') WHERE user_id=?", (user_id,))


def mark_welcome_discount_used(db, email, user_id, order_id):
    email = normalize_public_email(email) if email else ''
    clauses = []
    where_params = []
    if email:
        clauses.append("email=?")
        where_params.append(email)
    if user_id:
        clauses.append("user_id=?")
        where_params.append(user_id)
    if not clauses:
        return
    params = [clean_text(order_id, 80), user_id] + where_params
    sql = (
        "UPDATE welcome_discounts SET used_at=COALESCE(used_at, datetime('now')), "
        "used_order_id=COALESCE(used_order_id, ?), user_id=COALESCE(user_id, ?) "
        "WHERE " + " OR ".join(clauses)
    )
    db.execute(sql, params)  # nosec B608


def welcome_discount_status_payload(db, email='', user_id=None):
    if not email and user_id:
        row = db.execute("SELECT email FROM users WHERE id=?", (user_id,)).fetchone()
        email = row['email'] if row else ''
    email = normalize_public_email(email) if email else ''
    row = _welcome_discount_row(db, email=email, user_id=user_id)
    has_order = _customer_has_paid_order(db, email=email, user_id=user_id)
    used = bool(row and row.get('used_at'))
    issued = bool(row)
    available = bool(issued and not used and not has_order)
    if available:
        message = 'Your first-order welcome code is ready.'
    elif used:
        message = 'This welcome offer was already used.'
    elif has_order:
        message = 'WELCOME10 is only available before the first order.'
    else:
        message = 'No welcome offer is available for this account.'
    return {
        'available': available,
        'issued': issued,
        'used': used,
        'has_order': has_order,
        'code': WELCOME_COUPON_CODE if available else '',
        'message': message,
    }


def validate_welcome_coupon_for_checkout(db, coupon_code, *, customer_email='', user_id=None):
    if clean_text(coupon_code, 40).upper() != WELCOME_COUPON_CODE:
        return
    email = normalize_public_email(customer_email) if customer_email else ''
    if not email and user_id:
        row = db.execute("SELECT email FROM users WHERE id=?", (user_id,)).fetchone()
        email = row['email'] if row else ''
    email = normalize_public_email(email) if email else ''
    if not email:
        raise ValueError('Enter the email that received WELCOME10 before using this code.')
    row = _welcome_discount_row(db, email=email, user_id=user_id)
    if not row:
        raise ValueError('Please subscribe or create an account to receive WELCOME10.')
    if email and row.get('email') and normalize_public_email(row.get('email')) != email:
        raise ValueError('Use the email address that received WELCOME10.')
    if row.get('used_at'):
        raise ValueError('WELCOME10 was already used for this email.')
    if _customer_has_paid_order(db, email=email, user_id=user_id):
        raise ValueError('WELCOME10 is only for your first order.')


def generate_order_number():
    now = datetime.datetime.now()
    return f"ORD{now.strftime('%Y%m%d')}{str(uuid.uuid4())[:6].upper()}"


# ─── Auth Routes ─────────────────────────────────────────────────────────────

@app.route('/api/auth/csrf', methods=['GET'])
def csrf_token():
    token = request.cookies.get(CSRF_COOKIE_NAME, '')
    if not valid_csrf_token(token):
        token = make_csrf_token()
    response = jsonify({'csrf_token': token})
    response.headers['Cache-Control'] = 'no-store'
    return set_csrf_cookie(response, token)


@app.route('/api/auth/register', methods=['POST'])
@limiter.limit("5 per minute; 20 per hour")
def register():
    data = request.json or {}
    name = (data.get('name') or '').strip()
    raw_email = (data.get('email') or '').strip().lower()
    email = normalize_public_email(raw_email)
    password = data.get('password') or ''
    limited = public_abuse_guard('register', email=email or raw_email, ip_limit=8, email_limit=2, window_seconds=3600)
    if limited:
        return limited
    if not name:
        return jsonify({'error': 'Full name is required'}), 400
    name_error = validate_name(name)
    if name_error:
        return jsonify({'error': name_error}), 400
    if not email:
        return jsonify({'error': 'A valid email address is required'}), 400
    pw_error = validate_password(password)
    if pw_error:
        return jsonify({'error': pw_error}), 400
    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
    if existing:
        return jsonify({'error': 'Email already registered'}), 400
    pwd = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    uid = str(uuid.uuid4())
    db.execute("INSERT INTO users (id,name,email,phone,password,token_version) VALUES (?,?,?,?,?,0)",
               (uid, name, email, data.get('phone', ''), pwd))
    welcome = ensure_welcome_discount(db, email, user_id=uid, issued_via='register', send=False)
    welcome_emailed = False
    if welcome.get('issued'):
        email_welcome(name, email, WELCOME_COUPON_CODE)
        mark_welcome_discount_emailed(db, email, uid)
        welcome_emailed = True
    else:
        email_welcome(name, email)
    db.commit()
    user_payload = {'id': uid, 'name': name, 'email': email, 'role': 'customer'}
    token = make_auth_token(uid, email, 'customer', 0)
    return set_auth_cookie(jsonify({
        'user': user_payload,
        'welcome_discount': {
            'code': WELCOME_COUPON_CODE if welcome_emailed else '',
            'emailed': welcome_emailed,
            'already_issued': bool(welcome.get('already_issued')),
        },
    }), token)


@app.route('/api/auth/login', methods=['POST'])
@limiter.limit("10 per minute; 50 per hour")
def login():
    data = request.json or {}
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    portal = normalized_login_portal(data.get('portal'))
    if not email or not password:
        return jsonify({'error': 'Email and password are required'}), 400
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()

    lock_reason = login_lockout_reason(db, email, user, portal)
    if lock_reason:
        log_security_event(
            'login_locked',
            'warning',
            'Login temporarily locked after repeated failures',
            user_id=user['id'] if user else None,
            email=email,
            metadata={
                'portal': portal,
                'role': user['role'] if user else None,
                'lock_reason': lock_reason,
            },
        )
        return jsonify({'error': 'Too many login attempts. Please wait and try again.'}), 429

    password_ok = bcrypt.checkpw(password.encode(), (user['password'] if user else DUMMY_PASSWORD_HASH).encode())
    portal_allowed = bool(user)
    if user:
        is_staff_role = user['role'] in ('admin', 'staff')
        portal_allowed = not (
            (portal == 'customer' and is_staff_role) or
            (portal == 'staff' and not is_staff_role)
        )

    if not user or not password_ok or not portal_allowed:
        if not user:
            reason = 'unknown_email'
        elif not portal_allowed:
            reason = 'portal_mismatch'
        else:
            reason = 'bad_password'
        record_login_failure(db, email, user, portal, reason)
        return jsonify({'error': LOGIN_GENERIC_ERROR}), 401

    u = dict(user)
    tv = u.get('token_version') or 0
    token = make_auth_token(u['id'], u['email'], u['role'], tv)
    clear_login_failures(db, email, portal)
    log_security_event('login_success', 'info', 'User logged in', user_id=u['id'], email=u['email'], metadata={'role': u['role'], 'portal': portal})
    return set_auth_cookie(jsonify({'user': auth_user_payload(u)}), token)


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    return clear_auth_cookie(jsonify({'message': 'Logged out'}))


@app.route('/api/auth/me', methods=['GET'])
@token_required
def get_me():
    db = get_db()
    user = row_to_dict(db.execute("SELECT id,name,email,phone,role,address,created_at FROM users WHERE id=?", (g.user['id'],)).fetchone())
    return jsonify(user)


@app.route('/api/auth/forgot-password', methods=['POST'])
@limiter.limit("3 per minute; 10 per hour")
def forgot_password():
    data = request.json or {}
    raw_email = (data.get('email') or '').strip().lower()
    email = normalize_public_email(raw_email)
    limited = public_abuse_guard('forgot_password', email=email or raw_email, ip_limit=6, email_limit=3, window_seconds=3600)
    if limited:
        return limited
    if not email:
        return jsonify({'message': 'If that email is registered, a reset link has been sent.'})
    db = get_db()
    user = db.execute("SELECT id,name FROM users WHERE email=?", (email,)).fetchone()
    if user:
        token = make_password_reset_token()
        token_digest = password_reset_digest(token)
        expires = (datetime.datetime.utcnow() + datetime.timedelta(hours=1)).isoformat()
        db.execute("DELETE FROM password_resets WHERE email=?", (email,))
        db.execute("INSERT INTO password_resets (id,email,token,expires_at) VALUES (?,?,?,?)",
                   (str(uuid.uuid4()), email, token_digest, expires))
        db.commit()
        reset_link = f"https://adhyashaktishop.com/reset-password?token={token}"
        email_password_reset(user['name'], email, reset_link)
        log_security_event(
            'password_reset_requested',
            'info',
            'Password reset email sent',
            user_id=user['id'],
            email=email,
        )
    else:
        log_security_event(
            'password_reset_unknown_email',
            'warning',
            'Password reset requested for an unregistered email',
            email=email,
        )
    return jsonify({'message': 'If that email is registered, a reset link has been sent.'})


@app.route('/api/auth/reset-password', methods=['POST'])
@limiter.limit("5 per minute; 20 per hour")
def reset_password_route():
    data = request.json or {}
    token    = (data.get('token') or '').strip()
    new_pw   = data.get('password') or ''
    limited = public_abuse_guard('reset_password', fingerprint=token[:80], ip_limit=10, fingerprint_limit=5, window_seconds=900)
    if limited:
        return limited
    if not token or not new_pw:
        return jsonify({'error': 'Token and new password are required'}), 400
    if not valid_password_reset_token_format(token):
        log_security_event('password_reset_failed', 'warning', 'Invalid password reset token format')
        return jsonify({'error': 'Invalid or expired reset link'}), 400
    pw_error = validate_password(new_pw)
    if pw_error:
        return jsonify({'error': pw_error}), 400
    db = get_db()
    token_digest = password_reset_digest(token)
    reset = db.execute("SELECT * FROM password_resets WHERE token=? AND used=0", (token_digest,)).fetchone()
    if not reset:
        # Transitional support for reset links issued before reset tokens were hashed.
        reset = db.execute("SELECT * FROM password_resets WHERE token=? AND used=0", (token,)).fetchone()
    if not reset:
        log_security_event('password_reset_failed', 'warning', 'Invalid password reset token')
        return jsonify({'error': 'Invalid or expired reset link'}), 400
    if datetime.datetime.utcnow() > datetime.datetime.fromisoformat(reset['expires_at']):
        log_security_event('password_reset_failed', 'warning', 'Expired password reset token', email=reset['email'])
        return jsonify({'error': 'This reset link has expired. Please request a new one.'}), 400
    hashed = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
    db.execute("UPDATE users SET password=?, token_version=COALESCE(token_version,0)+1 WHERE email=?", (hashed, reset['email']))
    db.execute("UPDATE password_resets SET used=1 WHERE id=?", (reset['id'],))
    db.commit()
    log_security_event('password_reset_completed', 'info', 'Password reset completed', email=reset['email'])
    return clear_auth_cookie(jsonify({'message': 'Password updated. You can now log in with your new password.'}))


@app.route('/api/auth/change-password', methods=['POST'])
@limiter.limit("10 per minute; 30 per hour")
@token_required
def change_password():
    data = request.json or {}
    current_pw = data.get('current_password') or ''
    new_pw     = data.get('new_password') or ''
    if not current_pw or not new_pw:
        return jsonify({'error': 'Current and new password are required'}), 400
    pw_error = validate_password(new_pw)
    if pw_error:
        return jsonify({'error': pw_error}), 400
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (g.user['id'],)).fetchone()
    if not bcrypt.checkpw(current_pw.encode(), user['password'].encode()):
        return jsonify({'error': 'Current password is incorrect'}), 400
    hashed = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
    db.execute("UPDATE users SET password=?, token_version=COALESCE(token_version,0)+1 WHERE id=?", (hashed, g.user['id']))
    db.commit()
    log_security_event('password_changed', 'info', 'User changed password', user_id=g.user.get('id'), email=g.user.get('email'))
    return clear_auth_cookie(jsonify({'message': 'Password changed successfully. Please log in again.'}))


# ─── Products Routes ──────────────────────────────────────────────────────────

@app.route('/api/products', methods=['GET'])
def get_products():
    db = get_db()
    category = request.args.get('category')
    search = request.args.get('search')
    page = request.args.get('page', 1, type=int) or 1
    per_page = min(request.args.get('per_page', 12, type=int) or 12, 100)
    offset = (page - 1) * per_page

    query = """SELECT p.*, c.name as category_name,
        COALESCE(AVG(r.rating),0) as avg_rating,
        COUNT(r.id) as review_count
        FROM products p
        LEFT JOIN categories c ON p.category_id=c.id
        LEFT JOIN reviews r ON r.product_id=p.id
        WHERE p.is_active=1"""
    params = []
    category_filter = ''
    if category:
        cat_row = db.execute(
            "SELECT id,parent_id,kind,IFNULL(is_active,1) AS is_active FROM categories WHERE id=?",
            (category,)
        ).fetchone()
        category_ids = [category]
        if cat_row:
            if not cat_row['is_active']:
                category_ids = []
            else:
                category_ids = []
                stack = [category]
                while stack:
                    cur = stack.pop()
                    category_ids.append(cur)
                    stack.extend([r['id'] for r in db.execute(
                        "SELECT id FROM categories WHERE parent_id=? AND IFNULL(is_active,1)=1",
                        (cur,)
                    ).fetchall()])
        if category_ids:
            category_filter = " AND p.category_id IN (%s)" % ",".join("?" for _ in category_ids)
            query += category_filter
            params.extend(category_ids)
        else:
            category_filter = " AND 1=0"
            query += category_filter
    if search:
        query += " AND (p.name LIKE ? OR p.description LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%'])
    min_price = request.args.get('min_price', type=float)
    max_price = request.args.get('max_price', type=float)
    if min_price is not None:
        query += " AND p.price >= ?"
        params.append(min_price)
    if max_price is not None:
        query += " AND p.price <= ?"
        params.append(max_price)
    query += " GROUP BY p.id"

    total_where = ["p.is_active=1"]
    if category_filter:
        total_where.append(category_filter.replace(" AND ", "", 1))
    if search:
        total_where.append("(p.name LIKE ? OR p.description LIKE ?)")
    if min_price is not None:
        total_where.append("p.price >= ?")
    if max_price is not None:
        total_where.append("p.price <= ?")
    total = db.execute(
        "SELECT COUNT(*) FROM products p WHERE " + " AND ".join(total_where),  # nosec B608
        params
    ).fetchone()[0]
    sort = request.args.get('sort', 'newest')
    order_clause = {
        'newest':   'p.created_at DESC',
        'oldest':   'p.created_at ASC',
        'price_asc':  'p.price ASC',
        'price_desc': 'p.price DESC',
        'discount': '(p.compare_price - p.price) DESC',
    }.get(sort, 'p.created_at DESC')
    query += f" ORDER BY {order_clause} LIMIT {per_page} OFFSET {offset}"
    products = rows_to_list(db.execute(query, params).fetchall())
    for p in products:
        p['images'] = json.loads(p['images']) if p['images'] else []
        p['variations'] = json.loads(p['variations']) if p['variations'] else []
        variant_count = db.execute("SELECT COUNT(*) FROM product_variants WHERE product_id=?", (p['id'],)).fetchone()[0]
        p['has_variants'] = variant_count > 0
    return jsonify({'products': products, 'total': total, 'page': page, 'per_page': per_page})


@app.route('/api/products/<pid>', methods=['GET'])
def get_product(pid):
    db = get_db()
    p = row_to_dict(db.execute(
        "SELECT p.*, c.name as category_name FROM products p LEFT JOIN categories c ON p.category_id=c.id WHERE p.id=? AND p.is_active=1", (pid,)
    ).fetchone())
    if not p:
        return jsonify({'error': 'Not found'}), 404
    p['images'] = json.loads(p['images']) if p['images'] else []
    p['variations'] = json.loads(p['variations']) if p['variations'] else []
    p['variants'] = rows_to_list(db.execute(
        "SELECT color, size, stock FROM product_variants WHERE product_id=? ORDER BY color, size", (pid,)
    ).fetchall())
    reviews = rows_to_list(db.execute(
        "SELECT r.*, u.name as user_name FROM reviews r JOIN users u ON r.user_id=u.id WHERE r.product_id=?", (pid,)
    ).fetchall())
    for r in reviews:
        r['images'] = json.loads(r.get('images') or '[]')
    p['reviews'] = reviews
    return jsonify(p)


@app.route('/api/products/<pid>/back-in-stock', methods=['POST'])
@limiter.limit("8 per minute; 40 per hour")
def request_back_in_stock(pid):
    data = request.json or {}
    raw_email = (data.get('email') or '').strip().lower()
    email = normalize_public_email(raw_email)
    name = clean_text(data.get('name'), 120)
    limited = public_abuse_guard(
        'back_in_stock',
        email=email or raw_email,
        fingerprint=pid,
        ip_limit=20,
        email_limit=8,
        fingerprint_limit=160,
        window_seconds=3600,
    )
    if limited:
        return limited
    if not email:
        return jsonify({'error': 'Please enter a valid email address.'}), 400

    db = get_db()
    product = db.execute(
        "SELECT id,name,stock,is_active FROM products WHERE id=?",
        (clean_text(pid, 80),)
    ).fetchone()
    if not product or not product['is_active']:
        return jsonify({'error': 'Product not found'}), 404
    if int(product['stock'] or 0) > 0:
        return jsonify({
            'message': 'This product is back in stock now. You can add it to your cart.',
            'in_stock': True,
        })

    current_customer = _optional_customer_from_cookie(db)
    user_id = current_customer['id'] if current_customer and current_customer['email'].lower() == email else None
    db.execute(
        """
        INSERT INTO back_in_stock_requests
            (id,product_id,email,name,user_id,status,source)
        VALUES (?,?,?,?,?,'pending','product_page')
        ON CONFLICT(product_id,email) DO UPDATE SET
            name=COALESCE(NULLIF(excluded.name,''), back_in_stock_requests.name),
            user_id=COALESCE(excluded.user_id, back_in_stock_requests.user_id),
            status='pending',
            notified_at=NULL,
            last_requested_at=datetime('now'),
            request_count=IFNULL(back_in_stock_requests.request_count,1)+1
        """,
        (str(uuid.uuid4()), product['id'], email, name, user_id)
    )
    db.commit()
    log_security_event(
        'back_in_stock_requested',
        'info',
        'Customer requested a back-in-stock notification',
        user_id=user_id,
        email=email,
        metadata={'product_id': product['id'], 'product_name': product['name']},
    )
    return jsonify({
        'message': "You're on the list. We'll email you when this product is available again.",
        'in_stock': False,
    })


def _save_variants(db, pid, variants):
    db.execute("DELETE FROM product_variants WHERE product_id=?", (pid,))
    for v in (variants or []):
        db.execute(
            "INSERT INTO product_variants (id,product_id,color,size,stock) VALUES (?,?,?,?,?)",
            (str(uuid.uuid4()), pid, v['color'], v['size'], int(v.get('stock', 0)))
        )


def _clean_product_images(raw_images):
    if not isinstance(raw_images, list):
        return []
    cleaned = []
    for url in raw_images[:12]:
        text = clean_text(url, 260)
        if is_safe_static_image_url(text):
            cleaned.append(text)
    return cleaned


def _clean_product_variants(raw_variants):
    if not isinstance(raw_variants, list):
        return []
    cleaned = []
    for v in raw_variants[:240]:
        if not isinstance(v, dict):
            continue
        color = clean_text(v.get('color'), 40)
        size = clean_text(v.get('size'), 20)
        try:
            stock = max(0, min(999999, int(v.get('stock', 0))))
        except (TypeError, ValueError):
            stock = 0
        if color and size:
            cleaned.append({'color': color, 'size': size, 'stock': stock})
    return cleaned


def _clean_product_stock(value):
    try:
        return max(0, min(999999, int(value or 0)))
    except (TypeError, ValueError):
        return 0


def _optional_customer_from_cookie(db):
    token = request_auth_token()
    if not token:
        return None
    try:
        data = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        row = db.execute(
            "SELECT id,email,role,token_version FROM users WHERE id=?",
            (data.get('id'),)
        ).fetchone()
        if row and row['role'] == 'customer' and row['token_version'] == data.get('tv', 0):
            return row
    except jwt.InvalidTokenError:
        return None
    except Exception:
        return None
    return None


def _product_public_payload(row):
    if not row:
        return {}
    product = dict(row)
    try:
        product['images'] = json.loads(product.get('images') or '[]')
    except Exception:
        product['images'] = []
    return product


def notify_back_in_stock_for_product(db, product_id):
    product_row = db.execute(
        "SELECT id,name,price,images,stock,is_active FROM products WHERE id=?",
        (product_id,)
    ).fetchone()
    if not product_row or not product_row['is_active'] or int(product_row['stock'] or 0) <= 0:
        return 0
    requests = db.execute(
        """
        SELECT id,email,name
        FROM back_in_stock_requests
        WHERE product_id=? AND notified_at IS NULL
        ORDER BY created_at ASC
        """,
        (product_id,)
    ).fetchall()
    if not requests:
        return 0

    product = _product_public_payload(product_row)
    pending = [(r['id'], r['email'], r['name'] or '') for r in requests]
    # Deliver in the background and mark each request notified only after its
    # email actually sends, so a mail-server hiccup cannot silently drop
    # notifications — anything unsent stays pending and retries on the next restock.
    threading.Thread(
        target=_deliver_back_in_stock_notifications,
        args=(product_id, product, pending),
        daemon=True,
    ).start()

    if has_request_context():
        log_security_event(
            'back_in_stock_notified',
            'info',
            f'Queued {len(pending)} back-in-stock notification(s)',
            metadata={'product_id': product_id, 'product_name': product.get('name'), 'count': len(pending)},
        )
    return len(pending)


def _deliver_back_in_stock_notifications(product_id, product, pending):
    if not pending or not CONTACT_MAIL_PASS:
        return 0
    sent = 0
    conn = None
    try:
        conn = sqlite3.connect(DB_PATH, timeout=15)
        now = datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        for req_id, email, name in pending:
            # Claim this request atomically so a second delivery pass cannot
            # double-send it; revert to pending if the email does not go out.
            claimed = conn.execute(
                "UPDATE back_in_stock_requests SET status='notified', notified_at=? "
                "WHERE id=? AND notified_at IS NULL",
                (now, req_id),
            )
            conn.commit()
            if claimed.rowcount == 0:
                continue
            subject, html = _build_back_in_stock_email(name, product)
            if _send_email_bg(CONTACT_MAIL_USER, CONTACT_MAIL_PASS, email, subject, html):
                sent += 1
            else:
                conn.execute(
                    "UPDATE back_in_stock_requests SET status='pending', notified_at=NULL WHERE id=?",
                    (req_id,),
                )
                conn.commit()
    except Exception as exc:
        app.logger.error('back-in-stock delivery failed for product %s: %s', product_id, exc)
    finally:
        if conn is not None:
            conn.close()
    if sent < len(pending):
        app.logger.warning('back-in-stock: sent %s of %s for product %s (unsent remain pending)', sent, len(pending), product_id)
    return sent


def _clean_optional_money(value):
    if value in (None, ''):
        return None
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return None
    return amount if amount >= 0 else None


def product_audit_snapshot(db, pid):
    row = db.execute(
        """SELECT id,name,price,compare_price,cost_price,category_id,stock,low_stock_threshold,
                  sku,is_active,allow_custom_print,is_bestseller
           FROM products WHERE id=?""",
        (pid,)
    ).fetchone()
    return dict(row) if row else None


def _admin_product_health(db):
    def one(sql, params=()):
        row = db.execute(sql, params).fetchone()
        return int(row[0] or 0) if row else 0
    return {
        'all': one("SELECT COUNT(*) FROM products"),
        'active': one("SELECT COUNT(*) FROM products WHERE IFNULL(is_active,1)=1"),
        'inactive': one("SELECT COUNT(*) FROM products WHERE IFNULL(is_active,1)=0"),
        'out_of_stock': one("SELECT COUNT(*) FROM products WHERE IFNULL(stock,0)<=0"),
        'no_image': one("SELECT COUNT(*) FROM products WHERE images IS NULL OR images='' OR images='[]'"),
        'no_category': one("""
            SELECT COUNT(*) FROM products p
            WHERE p.category_id IS NULL OR p.category_id=''
               OR NOT EXISTS (SELECT 1 FROM categories c WHERE c.id=p.category_id)
        """),
        'no_cost': one("SELECT COUNT(*) FROM products WHERE IFNULL(cost_price,0)<=0"),
        'low_stock': one("""
            SELECT COUNT(*) FROM products
            WHERE IFNULL(is_active,1)=1
              AND IFNULL(stock,0)>0
              AND IFNULL(stock,0)<=IFNULL(low_stock_threshold,5)
        """),
        'duplicate_name': one("""
            SELECT COUNT(*) FROM products p
            WHERE LOWER(TRIM(IFNULL(p.name,''))) IN (
                SELECT LOWER(TRIM(IFNULL(name,''))) FROM products
                WHERE TRIM(IFNULL(name,''))!=''
                GROUP BY LOWER(TRIM(IFNULL(name,''))) HAVING COUNT(*)>1
            )
        """),
        'duplicate_sku': one("""
            SELECT COUNT(*) FROM products p
            WHERE TRIM(IFNULL(p.sku,''))!=''
              AND LOWER(TRIM(p.sku)) IN (
                SELECT LOWER(TRIM(sku)) FROM products
                WHERE TRIM(IFNULL(sku,''))!=''
                GROUP BY LOWER(TRIM(sku)) HAVING COUNT(*)>1
              )
        """),
        'notify_waiting': one("SELECT COUNT(*) FROM back_in_stock_requests WHERE notified_at IS NULL"),
    }


@app.route('/api/admin/products', methods=['GET'])
@admin_required
def admin_products_list():
    db = get_db()
    status = clean_text(request.args.get('status'), 40).lower() or 'active'
    search = clean_text(request.args.get('search'), 160)
    page = max(1, request.args.get('page', 1, type=int) or 1)
    per_page = min(max(5, request.args.get('per_page', 20, type=int) or 20), 100)
    offset = (page - 1) * per_page
    where = []
    params = []
    if status == 'active':
        where.append("IFNULL(p.is_active,1)=1")
    elif status == 'inactive':
        where.append("IFNULL(p.is_active,1)=0")
    elif status == 'out_of_stock':
        where.append("IFNULL(p.stock,0)<=0")
    elif status == 'no_image':
        where.append("(p.images IS NULL OR p.images='' OR p.images='[]')")
    elif status == 'no_category':
        where.append("""(
            p.category_id IS NULL OR p.category_id=''
            OR NOT EXISTS (SELECT 1 FROM categories cx WHERE cx.id=p.category_id)
        )""")
    elif status == 'no_cost':
        where.append("IFNULL(p.cost_price,0)<=0")
    elif status == 'low_stock':
        where.append("""
            IFNULL(p.is_active,1)=1
            AND IFNULL(p.stock,0)>0
            AND IFNULL(p.stock,0)<=IFNULL(p.low_stock_threshold,5)
        """)
    elif status == 'duplicate_name':
        where.append("""
            LOWER(TRIM(IFNULL(p.name,''))) IN (
                SELECT LOWER(TRIM(IFNULL(name,''))) FROM products
                WHERE TRIM(IFNULL(name,''))!=''
                GROUP BY LOWER(TRIM(IFNULL(name,''))) HAVING COUNT(*)>1
            )
        """)
    elif status == 'duplicate_sku':
        where.append("""
            TRIM(IFNULL(p.sku,''))!=''
            AND LOWER(TRIM(p.sku)) IN (
                SELECT LOWER(TRIM(sku)) FROM products
                WHERE TRIM(IFNULL(sku,''))!=''
                GROUP BY LOWER(TRIM(sku)) HAVING COUNT(*)>1
            )
        """)
    elif status == 'notify_waiting':
        where.append("""
            EXISTS (
                SELECT 1 FROM back_in_stock_requests bis
                WHERE bis.product_id=p.id AND bis.notified_at IS NULL
            )
        """)
    elif status not in ('all', ''):
        where.append("IFNULL(p.is_active,1)=1")
        status = 'active'
    if search:
        where.append("(p.name LIKE ? OR p.description LIKE ? OR p.sku LIKE ?)")
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    where_sql = (" WHERE " + " AND ".join(where)) if where else ""
    total = db.execute(f"SELECT COUNT(*) FROM products p{where_sql}", params).fetchone()[0]  # nosec B608
    rows = rows_to_list(db.execute(
        f"""SELECT p.*, c.name AS category_name,
                   (SELECT COUNT(*) FROM back_in_stock_requests bis
                    WHERE bis.product_id=p.id AND bis.notified_at IS NULL) AS back_in_stock_waiting,
                   (SELECT COUNT(*) FROM products pn
                    WHERE TRIM(IFNULL(pn.name,''))!=''
                      AND LOWER(TRIM(pn.name))=LOWER(TRIM(IFNULL(p.name,'')))) AS duplicate_name_count,
                   (SELECT COUNT(*) FROM products ps
                    WHERE TRIM(IFNULL(ps.sku,''))!=''
                      AND LOWER(TRIM(ps.sku))=LOWER(TRIM(IFNULL(p.sku,'')))) AS duplicate_sku_count
            FROM products p
            LEFT JOIN categories c ON c.id=p.category_id
            {where_sql}
            ORDER BY p.created_at DESC
            LIMIT ? OFFSET ?""",  # nosec B608
        [*params, per_page, offset]
    ).fetchall())
    for p in rows:
        p['images'] = json.loads(p['images']) if p.get('images') else []
        p['variations'] = json.loads(p['variations']) if p.get('variations') else []
        p['has_variants'] = db.execute(
            "SELECT COUNT(*) FROM product_variants WHERE product_id=?",
            (p['id'],)
        ).fetchone()[0] > 0
    return jsonify({
        'products': rows,
        'total': total,
        'page': page,
        'per_page': per_page,
        'status': status,
        'health': _admin_product_health(db),
    })


@app.route('/api/admin/products/<pid>', methods=['GET'])
@admin_required
def admin_get_product(pid):
    db = get_db()
    p = row_to_dict(db.execute(
        "SELECT p.*, c.name AS category_name FROM products p LEFT JOIN categories c ON c.id=p.category_id WHERE p.id=?",
        (pid,)
    ).fetchone())
    if not p:
        return jsonify({'error': 'Product not found'}), 404
    p['images'] = json.loads(p['images']) if p.get('images') else []
    p['variations'] = json.loads(p['variations']) if p.get('variations') else []
    p['variants'] = rows_to_list(db.execute(
        "SELECT color,size,stock FROM product_variants WHERE product_id=? ORDER BY color,size",
        (pid,)
    ).fetchall())
    return jsonify(p)


@app.route('/api/admin/products', methods=['POST'])
@admin_required
def create_product():
    data = request.json or {}
    name = clean_text(data.get('name'), 160)
    if not name:
        return jsonify({'error': 'Product name is required'}), 400
    try:
        price = float(data.get('price'))
    except (TypeError, ValueError):
        return jsonify({'error': 'A valid price is required'}), 400
    if price <= 0:
        return jsonify({'error': 'Price must be greater than 0'}), 400
    db = get_db()
    pid = str(uuid.uuid4())
    category_id = clean_text(data.get('category_id'), 80) or None
    if category_id and not db.execute("SELECT id FROM categories WHERE id=?", (category_id,)).fetchone():
        return jsonify({'error': 'Selected category was not found'}), 400
    cost_price = _clean_optional_money(data.get('cost_price')) or 0
    low_stock_threshold = _clean_product_stock(data.get('low_stock_threshold', 5))
    variants = _clean_product_variants(data.get('variants', []))
    total_stock = sum(int(v.get('stock', 0)) for v in variants) if variants else _clean_product_stock(data.get('stock', 0))
    db.execute(
        """INSERT INTO products
           (id,name,description,price,compare_price,cost_price,category_id,stock,low_stock_threshold,
            sku,images,variations,allow_custom_print,is_bestseller)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (pid, name, clean_text(data.get('description'), 5000), price, _clean_optional_money(data.get('compare_price')),
         cost_price, category_id, total_stock, low_stock_threshold, clean_text(data.get('sku'), 80),
         json.dumps(_clean_product_images(data.get('images', []))), json.dumps([]),
         1 if data.get('allow_custom_print') else 0,
         1 if data.get('is_bestseller') else 0)
    )
    _save_variants(db, pid, variants)
    db.commit()
    log_admin_action('product_created', f'Created product: {name}', {
        'product_id': pid,
        'entity_type': 'product',
        'entity_id': pid,
        'after': product_audit_snapshot(db, pid),
    })
    return jsonify({'id': pid, 'message': 'Product created'})


@app.route('/api/admin/products/<pid>', methods=['PUT'])
@admin_required
def update_product(pid):
    data = request.json or {}
    name = clean_text(data.get('name'), 160)
    if not name:
        return jsonify({'error': 'Product name is required'}), 400
    try:
        price = float(data.get('price'))
    except (TypeError, ValueError):
        return jsonify({'error': 'A valid price is required'}), 400
    if price <= 0:
        return jsonify({'error': 'Price must be greater than 0'}), 400
    db = get_db()
    before = product_audit_snapshot(db, pid)
    if not before:
        return jsonify({'error': 'Product not found'}), 404
    category_id = clean_text(data.get('category_id'), 80) or None
    if category_id and not db.execute("SELECT id FROM categories WHERE id=?", (category_id,)).fetchone():
        return jsonify({'error': 'Selected category was not found'}), 400
    cost_price = _clean_optional_money(data.get('cost_price')) or 0
    low_stock_threshold = _clean_product_stock(data.get('low_stock_threshold', 5))
    variants = _clean_product_variants(data.get('variants', []))
    total_stock = sum(int(v.get('stock', 0)) for v in variants) if variants else _clean_product_stock(data.get('stock', 0))
    db.execute(
        """UPDATE products SET name=?,description=?,price=?,compare_price=?,cost_price=?,category_id=?,
              stock=?,low_stock_threshold=?,sku=?,images=?,variations=?,is_active=?,allow_custom_print=?,is_bestseller=?
           WHERE id=?""",
        (name, clean_text(data.get('description'), 5000), price, _clean_optional_money(data.get('compare_price')),
         cost_price, category_id, total_stock, low_stock_threshold, clean_text(data.get('sku'), 80),
         json.dumps(_clean_product_images(data.get('images', []))), json.dumps([]),
         data.get('is_active', 1), 1 if data.get('allow_custom_print') else 0,
         1 if data.get('is_bestseller') else 0, pid)
    )
    _save_variants(db, pid, variants)
    db.commit()
    if int(before.get('stock') or 0) <= 0 and total_stock > 0:
        try:
            notify_back_in_stock_for_product(db, pid)
        except Exception as exc:
            app.logger.warning('back-in-stock notify after product update failed: %s', exc)
    log_admin_action('product_updated', f'Updated product: {name}', {
        'product_id': pid,
        'entity_type': 'product',
        'entity_id': pid,
        'before': before,
        'after': product_audit_snapshot(db, pid),
    })
    return jsonify({'message': 'Product updated'})


@app.route('/api/admin/products/<pid>', methods=['DELETE'])
@admin_required
def delete_product(pid):
    db = get_db()
    before = product_audit_snapshot(db, pid)
    if not before:
        return jsonify({'error': 'Product not found'}), 404
    db.execute("UPDATE products SET is_active=0 WHERE id=?", (pid,))
    db.commit()
    log_admin_action('product_deactivated', 'Product marked inactive', {
        'product_id': pid,
        'entity_type': 'product',
        'entity_id': pid,
        'before': before,
        'after': product_audit_snapshot(db, pid),
    })
    return jsonify({'message': 'Product deleted'})


# ─── Categories ───────────────────────────────────────────────────────────────

@app.route('/api/categories', methods=['GET'])
def get_categories():
    db = get_db()
    # Public flat read of the managed category tree. Hierarchical clients should
    # prefer /api/category-tree; this keeps older clients from seeing legacy rows.
    cats = rows_to_list(db.execute(
        "SELECT * FROM categories WHERE IFNULL(kind,'') IN ('catalog','clothing') "
        "AND IFNULL(is_active,1)=1 ORDER BY sort_order, name COLLATE NOCASE").fetchall())
    return jsonify(cats)


# (Legacy POST/DELETE /api/admin/categories removed — category management is now the
#  guarded tree API in accounts_module: /api/acc/categories/node. The unguarded delete
#  here could orphan products, and the create made flat kind=NULL rows invisible to the
#  managed tree. GET /api/categories above remains for legacy/flat reads.)


# ─── Stripe Payment ───────────────────────────────────────────────────────────

def _capture_abandoned_cart(db, *, email, name, phone, user_id, items, totals, payment_intent_id):
    """Snapshot a checkout attempt so an unfinished cart can be reminded later.
    Stores minimal item refs (id/qty/variation); product details are looked up
    at send time so the reminder always reflects current names, prices, images."""
    if not email:
        return
    refs = []
    for it in (items or [])[:50]:
        if not isinstance(it, dict):
            continue
        pid = clean_text(it.get('id'), 80)
        if not pid:
            continue
        try:
            qty = max(1, int(it.get('qty') or it.get('quantity') or 1))
        except Exception:
            qty = 1
        refs.append({'id': pid, 'qty': qty, 'variation': clean_text(it.get('variation'), 100)})
    if not refs:
        return
    db.execute(
        """
        INSERT INTO abandoned_carts
            (id,email,name,phone,user_id,items,subtotal,discount,shipping,total,
             coupon_code,payment_intent_id,recovery_token,created_at,reminded_at,converted_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'),NULL,NULL)
        ON CONFLICT(email) DO UPDATE SET
            name=excluded.name,
            phone=excluded.phone,
            user_id=COALESCE(excluded.user_id, abandoned_carts.user_id),
            items=excluded.items,
            subtotal=excluded.subtotal,
            discount=excluded.discount,
            shipping=excluded.shipping,
            total=excluded.total,
            coupon_code=excluded.coupon_code,
            payment_intent_id=excluded.payment_intent_id,
            recovery_token=COALESCE(abandoned_carts.recovery_token, excluded.recovery_token),
            created_at=datetime('now'),
            reminded_at=NULL,
            converted_at=NULL
        """,
        (str(uuid.uuid4()), email, name, phone, user_id, json.dumps(refs),
         float(totals.get('subtotal') or 0), float(totals.get('discount') or 0),
         float(totals.get('shipping') or 0), float(totals.get('total') or 0),
         totals.get('coupon_code') or '', payment_intent_id, uuid.uuid4().hex)
    )
    db.commit()


@app.route('/api/create-payment-intent', methods=['POST'])
@limiter.limit("15 per minute; 100 per hour")
def create_payment_intent():
    data = request.json or {}
    try:
        db = get_db()
        customer_name = clean_text(data.get('customer_name'), 100)
        raw_customer_email = clean_text(data.get('customer_email'), 140).lower()
        customer_email = normalize_public_email(raw_customer_email)
        raw_customer_phone = clean_text(data.get('customer_phone'), 30)
        customer_phone = normalize_public_phone(raw_customer_phone) if raw_customer_phone else ''
        limited = public_abuse_guard(
            'create_payment_intent',
            email=customer_email or raw_customer_email,
            ip_limit=24,
            email_limit=10,
            window_seconds=3600,
        )
        if limited:
            return limited
        if not customer_name or not customer_email:
            return jsonify({'error': 'Name and email are required before payment.'}), 400
        if raw_customer_phone and not customer_phone:
            return jsonify({'error': 'Please enter a valid phone number'}), 400
        if public_text_looks_spammy(customer_name, customer_phone):
            return jsonify({'error': 'Please remove unsupported characters and try again.'}), 400
        shipping_address = _normalize_shipping_address(data.get('shipping_address'))
        checkout_hash = _checkout_context_fingerprint(customer_name, customer_email, customer_phone, shipping_address)
        customer_user_id = current_customer_id_from_request(db)
        totals = _calculate_order_totals(
            db,
            data.get('items'),
            data.get('coupon_code'),
            customer_email=customer_email,
            user_id=customer_user_id,
        )
        if totals['total'] <= 0:
            return jsonify({'error': 'Invalid order total'}), 400
        if totals['total'] > 10000:
            return jsonify({'error': 'Order amount is too large. Please contact support.'}), 400
        amount_cents = int(round(totals['total'] * 100))  # Stripe needs cents
        intent = stripe.PaymentIntent.create(
            amount=amount_cents,
            currency='usd',
            automatic_payment_methods={'enabled': True},
            metadata={
                'source': 'adhya_checkout',
                'cart_hash': totals['cart_hash'],
                'coupon_code': totals['coupon_code'],
                'amount_cents': str(amount_cents),
                'checkout_hash': checkout_hash,
                'customer_email': customer_email[:140],
            },
        )
        # Snapshot this checkout attempt for abandoned-cart recovery. A capture
        # problem must never interfere with taking the customer's payment.
        try:
            _capture_abandoned_cart(
                db, email=customer_email, name=customer_name, phone=customer_phone,
                user_id=customer_user_id, items=data.get('items'),
                totals=totals, payment_intent_id=intent['id'],
            )
        except Exception as _cap_exc:
            app.logger.warning('abandoned-cart capture failed: %s', _cap_exc)

        return jsonify({
            'client_secret': intent['client_secret'],
            'subtotal': round(totals['subtotal'], 2),
            'discount': round(totals['discount'], 2),
            'shipping': round(totals['shipping'], 2),
            'total': round(totals['total'], 2),
        })
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except stripe.error.StripeError as e:
        return jsonify({'error': str(e.user_message)}), 400
    except Exception as e:
        return jsonify({'error': 'Payment setup failed. Please try again.'}), 500


@app.route('/api/cart-reminders/unsubscribe/<token>', methods=['GET'])
@limiter.limit("40 per hour")
def unsubscribe_cart_reminders(token):
    token = clean_text(token, 80)
    page = ('<!doctype html><html><head><meta charset="utf-8">'
            '<meta name="viewport" content="width=device-width,initial-scale=1">'
            '<title>Unsubscribed</title></head>'
            '<body style="font-family:Arial,Helvetica,sans-serif;max-width:520px;margin:60px auto;padding:0 20px;text-align:center;color:#333">'
            '<h2 style="color:#1D5C4A">You\'re unsubscribed</h2>'
            '<p>You won\'t receive any more cart reminder emails from Adhya Shakti Shop. '
            'You will still get emails about orders you place.</p>'
            '<p><a href="https://adhyashaktishop.com" style="color:#1D5C4A;font-weight:700;text-decoration:none">Return to the shop</a></p>'
            '</body></html>')
    if token:
        try:
            db = get_db()
            db.execute("UPDATE abandoned_carts SET unsubscribed=1 WHERE recovery_token=?", (token,))
            db.commit()
        except Exception:
            pass
    return Response(page, mimetype='text/html')


@app.route('/api/stripe-key', methods=['GET'])
def get_stripe_key():
    return jsonify({'publishable_key': STRIPE_PUBLISHABLE_KEY})


# ─── Orders ───────────────────────────────────────────────────────────────────

def _safe_json_list(value):
    try:
        parsed = json.loads(value or '[]')
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []


def _safe_product_image(images_json):
    for url in _safe_json_list(images_json):
        if is_safe_static_image_url(url):
            return url
    return ''


def _order_access_allowed(order):
    if not order:
        return False
    return order.get('user_id') == g.user.get('id') or g.user.get('role') in ('admin', 'staff')


def _log_order_access_denied(action, oid, order=None):
    log_security_event(
        'authorization_denied',
        'warning',
        'Order access denied',
        user_id=g.user.get('id'),
        email=g.user.get('email'),
        metadata={
            'action': action,
            'requested_order': clean_text(oid, 100),
            'target_order_id': order.get('id') if order else None,
            'target_order_number': order.get('order_number') if order else None,
            'target_user_id': order.get('user_id') if order else None,
            'role': g.user.get('role'),
        },
    )


def wishlist_product_payload(row):
    return {
        'id': row['id'],
        'name': row['name'],
        'price': row['price'] or 0,
        'compare_price': row['compare_price'] or 0,
        'image': _safe_product_image(row['images']),
    }


@app.route('/api/wishlist', methods=['GET'])
@token_required
def get_wishlist():
    db = get_db()
    rows = db.execute(
        """SELECT p.id,p.name,p.price,p.compare_price,p.images,w.created_at
           FROM user_wishlist w
           JOIN products p ON p.id=w.product_id
           WHERE w.user_id=? AND p.is_active=1
           ORDER BY w.created_at DESC""",
        (g.user['id'],)
    ).fetchall()
    return jsonify({'items': [wishlist_product_payload(r) for r in rows]})


@app.route('/api/wishlist', methods=['POST'])
@token_required
def add_wishlist():
    data = request.json or {}
    product_id = clean_text(data.get('product_id'), 80)
    if not product_id:
        return jsonify({'error': 'Product is required'}), 400
    db = get_db()
    product = db.execute(
        "SELECT id,name,price,compare_price,images FROM products WHERE id=? AND is_active=1",
        (product_id,)
    ).fetchone()
    if not product:
        return jsonify({'error': 'Product not found'}), 404
    db.execute(
        "INSERT OR IGNORE INTO user_wishlist (user_id,product_id) VALUES (?,?)",
        (g.user['id'], product_id)
    )
    db.commit()
    return jsonify({'item': wishlist_product_payload(product)})


@app.route('/api/wishlist/<product_id>', methods=['DELETE'])
@token_required
def remove_wishlist(product_id):
    db = get_db()
    db.execute(
        "DELETE FROM user_wishlist WHERE user_id=? AND product_id=?",
        (g.user['id'], clean_text(product_id, 80))
    )
    db.commit()
    return jsonify({'message': 'Removed from wishlist'})


def _normalize_custom_print(raw_custom, allow_custom_print):
    if not raw_custom:
        return None
    if not allow_custom_print:
        raise ValueError('Custom print upload is not available for one of the selected products.')
    if not isinstance(raw_custom, dict):
        raise ValueError('Invalid custom print data.')

    placement = clean_text(raw_custom.get('placement'), 20).lower()
    if placement not in ('front', 'back', 'both'):
        raise ValueError('Invalid custom print placement.')

    front = raw_custom.get('front_images') or []
    back = raw_custom.get('back_images') or []
    if not isinstance(front, list) or not isinstance(back, list):
        raise ValueError('Invalid custom print files.')
    front = [clean_text(u, 260) for u in front if is_safe_public_upload_url(u)]
    back = [clean_text(u, 260) for u in back if is_safe_public_upload_url(u)]
    if len(front) + len(back) == 0:
        raise ValueError('Please upload at least one custom print file.')
    if len(front) + len(back) > 6:
        raise ValueError('Please upload no more than 6 custom print files per item.')
    if placement == 'front' and not front:
        raise ValueError('Front design file is required.')
    if placement == 'back' and not back:
        raise ValueError('Back design file is required.')
    if placement == 'both' and (not front or not back):
        raise ValueError('Front and back design files are required.')

    return {
        'placement': placement,
        'front_images': front,
        'back_images': back,
        'extra_charge': 8.99 if placement == 'both' else 0,
    }


def _normalize_order_items(db, raw_items, *, check_stock=True):
    if not isinstance(raw_items, list) or not raw_items:
        raise ValueError('Your cart is empty')
    if len(raw_items) > 50:
        raise ValueError('Too many items in one order. Please split the order.')

    stored_items = []
    real_subtotal = 0.0
    for raw in raw_items:
        if not isinstance(raw, dict):
            raise ValueError('Invalid cart item.')
        pid = clean_text(raw.get('id'), 80)
        try:
            qty = int(raw.get('qty', 1))
        except (TypeError, ValueError):
            qty = 0
        if qty < 1 or qty > 99:
            raise ValueError('Invalid quantity in cart.')

        product = db.execute(
            "SELECT id,name,price,stock,images,allow_custom_print FROM products WHERE id=? AND is_active=1",
            (pid,)
        ).fetchone()
        if not product:
            raise ValueError('A product in your cart could not be found. Please refresh and try again.')

        variation = clean_text(raw.get('variation'), 80)
        variant_count = db.execute("SELECT COUNT(*) FROM product_variants WHERE product_id=?", (pid,)).fetchone()[0]
        if variant_count:
            if not variation or ' / ' not in variation:
                raise ValueError(f'Please select size/color for "{product["name"]}".')
            color, size = [clean_text(p, 40) for p in variation.split(' / ', 1)]
            variant = db.execute(
                "SELECT stock FROM product_variants WHERE product_id=? AND color=? AND size=?",
                (pid, color, size)
            ).fetchone()
            if not variant:
                raise ValueError(f'Please select an available size/color for "{product["name"]}".')
            available = variant['stock']
            if check_stock and available < qty:
                raise ValueError(f'"{product["name"]}" ({variation}) only has {available} left in stock. Please update your cart.')
            variation = f'{color} / {size}'
        elif check_stock and int(product['stock'] or 0) < qty:
            raise ValueError(f'"{product["name"]}" only has {product["stock"]} left in stock. Please update your cart.')

        custom_print = _normalize_custom_print(raw.get('customPrint'), bool(product['allow_custom_print']))
        unit_price = float(product['price'])
        if custom_print:
            unit_price += float(custom_print['extra_charge'])

        stored = {
            'id': product['id'],
            'name': product['name'],
            'price': unit_price,
            'qty': qty,
            'variation': variation,
            'image': _safe_product_image(product['images']),
            'customPrint': custom_print,
        }
        stored_items.append(stored)
        real_subtotal += unit_price * qty

    return stored_items, real_subtotal


def _normalize_shipping_address(raw_shipping):
    if not raw_shipping or not isinstance(raw_shipping, dict):
        raise ValueError('Shipping address is required')
    shipping_address = {
        'line1': clean_text(raw_shipping.get('line1'), 180),
        'city': clean_text(raw_shipping.get('city'), 80),
        'state': clean_text(raw_shipping.get('state'), 80),
        'pin': clean_text(raw_shipping.get('pin') or raw_shipping.get('zip'), 12),
        'landmark': clean_text(raw_shipping.get('landmark'), 120),
    }
    if not shipping_address['line1'] or not shipping_address['city'] or not shipping_address['state'] or not shipping_address['pin']:
        raise ValueError('Complete shipping address is required')
    if not re.fullmatch(r'\d{5}(?:-\d{4})?', shipping_address['pin']):
        raise ValueError('A valid ZIP code is required')
    return shipping_address


def _checkout_context_fingerprint(customer_name, customer_email, customer_phone, shipping_address):
    payload = {
        'customer_name': clean_text(customer_name, 100),
        'customer_email': clean_text(customer_email, 140).lower(),
        'customer_phone': clean_text(customer_phone, 30),
        'shipping_address': {
            'line1': clean_text((shipping_address or {}).get('line1'), 180),
            'city': clean_text((shipping_address or {}).get('city'), 80),
            'state': clean_text((shipping_address or {}).get('state'), 80),
            'pin': clean_text((shipping_address or {}).get('pin'), 12),
            'landmark': clean_text((shipping_address or {}).get('landmark'), 120),
        },
    }
    raw = json.dumps(payload, sort_keys=True, separators=(',', ':'))
    return hashlib.sha256(raw.encode()).hexdigest()


def _cart_fingerprint(stored_items, coupon_code=''):
    billable_items = []
    for item in stored_items:
        custom = item.get('customPrint') or None
        custom_sig = None
        if custom:
            custom_sig = {
                'placement': custom.get('placement'),
                'front_images': custom.get('front_images') or [],
                'back_images': custom.get('back_images') or [],
            }
        billable_items.append({
            'id': item.get('id'),
            'qty': int(item.get('qty') or 0),
            'variation': item.get('variation') or '',
            'customPrint': custom_sig,
        })
    payload = {
        'coupon_code': clean_text(coupon_code, 40).upper(),
        'items': billable_items,
    }
    raw = json.dumps(payload, sort_keys=True, separators=(',', ':'))
    return hashlib.sha256(raw.encode()).hexdigest()


def _calculate_order_totals(db, raw_items, raw_coupon_code='', *, check_stock=True, customer_email='', user_id=None):
    stored_items, real_subtotal = _normalize_order_items(db, raw_items, check_stock=check_stock)
    discount = 0.0
    coupon_code = clean_text(raw_coupon_code, 40).upper()
    coupon_obj = None
    if coupon_code:
        coupon = db.execute("SELECT * FROM coupons WHERE code=? AND is_active=1", (coupon_code,)).fetchone()
        if not coupon:
            raise ValueError('Invalid or expired coupon')
        coupon_obj = dict(coupon)
        if coupon_obj['expires_at'] and coupon_obj['expires_at'] < datetime.datetime.now().isoformat():
            raise ValueError('Coupon has expired')
        if coupon_obj['max_uses'] and coupon_obj['used_count'] >= coupon_obj['max_uses']:
            raise ValueError('Coupon usage limit reached')
        if real_subtotal < coupon_obj['min_order']:
            raise ValueError(f"Minimum order ${coupon_obj['min_order']:.2f} required")
        validate_welcome_coupon_for_checkout(
            db,
            coupon_code,
            customer_email=customer_email,
            user_id=user_id,
        )
        if coupon_obj['discount_type'] == 'percent':
            raw = (real_subtotal * coupon_obj['discount_value']) / 100
            discount = min(raw, real_subtotal)
        else:
            discount = min(coupon_obj['discount_value'], real_subtotal)

    shipping = 0.0 if real_subtotal >= 49 else 7.99
    total = real_subtotal - discount + shipping
    return {
        'stored_items': stored_items,
        'subtotal': real_subtotal,
        'discount': discount,
        'coupon_code': coupon_code,
        'coupon_obj': coupon_obj,
        'shipping': shipping,
        'total': total,
        'cart_hash': _cart_fingerprint(stored_items, coupon_code),
    }


@app.route('/api/cart/validate', methods=['POST'])
@limiter.limit("30 per minute; 200 per hour")
def validate_cart():
    data = request.json or {}
    db = get_db()
    raw_customer_email = clean_text(data.get('customer_email'), 140).lower()
    customer_email = normalize_public_email(raw_customer_email) if raw_customer_email else ''
    try:
        user_id = current_customer_id_from_request(db)
        totals = _calculate_order_totals(
            db,
            data.get('items'),
            data.get('coupon_code'),
            customer_email=customer_email,
            user_id=user_id,
        )
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    return jsonify({
        'items': totals['stored_items'],
        'subtotal': round(totals['subtotal'], 2),
        'discount': round(totals['discount'], 2),
        'shipping': round(totals['shipping'], 2),
        'total': round(totals['total'], 2),
        'coupon_code': totals['coupon_code'],
    })


def _reserve_order_stock(db, stored_items):
    for item in stored_items:
        pid = item['id']
        qty = int(item.get('qty') or 0)
        variation = item.get('variation', '') or ''
        if qty < 1:
            raise ValueError('Invalid quantity in cart.')

        if variation and ' / ' in variation:
            color, size = [p.strip() for p in variation.split(' / ', 1)]
            cur = db.execute(
                """UPDATE product_variants
                   SET stock = stock - ?
                   WHERE product_id=? AND color=? AND size=? AND stock >= ?""",
                (qty, pid, color, size, qty)
            )
            if cur.rowcount != 1:
                raise ValueError(f'"{item["name"]}" ({variation}) is no longer available in the requested quantity.')

        cur = db.execute(
            "UPDATE products SET stock = stock - ? WHERE id=? AND stock >= ?",
            (qty, pid, qty)
        )
        if cur.rowcount != 1:
            raise ValueError(f'"{item["name"]}" is no longer available in the requested quantity.')


def _restore_order_stock(db, stored_items):
    affected_product_ids = set()
    for item in stored_items:
        pid = item['id']
        qty = int(item.get('qty') or 0)
        variation = item.get('variation', '') or ''
        if qty < 1:
            continue
        db.execute("UPDATE products SET stock = stock + ? WHERE id=?", (qty, pid))
        affected_product_ids.add(pid)
        if variation and ' / ' in variation:
            color, size = [p.strip() for p in variation.split(' / ', 1)]
            db.execute(
                "UPDATE product_variants SET stock = stock + ? WHERE product_id=? AND color=? AND size=?",
                (qty, pid, color, size)
            )
    return list(affected_product_ids)


def public_order_payload(order):
    """Customer-safe order shape: excludes payment processor IDs and staff notes."""
    if not order:
        return None
    o = dict(order)
    raw_items = o.get('items') or []
    raw_shipping = o.get('shipping_address') or {}
    try:
        items = json.loads(raw_items) if isinstance(raw_items, str) else raw_items
    except Exception:
        items = []
    try:
        shipping_address = json.loads(raw_shipping) if isinstance(raw_shipping, str) else raw_shipping
    except Exception:
        shipping_address = {}
    return {
        'id': o.get('id'),
        'order_number': o.get('order_number'),
        'customer_name': o.get('customer_name'),
        'customer_email': o.get('customer_email'),
        'customer_phone': o.get('customer_phone'),
        'shipping_address': shipping_address,
        'items': items,
        'subtotal': o.get('subtotal') or 0,
        'discount': o.get('discount') or 0,
        'coupon_code': o.get('coupon_code') or '',
        'shipping_charge': o.get('shipping_charge') or 0,
        'total': o.get('total') or 0,
        'payment_method': o.get('payment_method') or '',
        'payment_status': o.get('payment_status') or '',
        'status': o.get('status') or '',
        'tracking_number': o.get('tracking_number') or '',
        'return_reason': o.get('return_reason') or '',
        'created_at': o.get('created_at'),
        'updated_at': o.get('updated_at'),
    }


def _refund_payment_intent(payment_intent_id, amount_cents, reason, email=None):
    if not payment_intent_id or amount_cents <= 0:
        return False
    try:
        stripe.Refund.create(payment_intent=payment_intent_id, amount=amount_cents)
        log_security_event(
            'payment_refunded',
            'info',
            reason,
            email=email,
            metadata={'payment_intent_id': payment_intent_id, 'amount_cents': amount_cents},
        )
        return True
    except Exception as exc:
        log_security_event(
            'payment_refund_failed',
            'critical',
            reason,
            email=email,
            metadata={'payment_intent_id': payment_intent_id, 'amount_cents': amount_cents, 'error': str(exc)},
        )
        return False


def _refund_succeeded_payment_intent(payment_intent_id, reason, email=None):
    """Refund a paid intent when checkout cannot safely create an order."""
    if not payment_intent_id:
        return False
    try:
        intent = stripe.PaymentIntent.retrieve(payment_intent_id)
        if intent.get('status') == 'succeeded':
            amount_cents = int(intent.get('amount') or 0)
            if amount_cents > 0:
                return _refund_payment_intent(payment_intent_id, amount_cents, reason, email=email)
    except Exception as exc:
        log_security_event(
            'payment_refund_failed',
            'critical',
            'Could not verify paid checkout for automatic refund',
            email=email,
            metadata={'payment_intent_id': payment_intent_id, 'reason': reason, 'error': str(exc)},
        )
    return False


def _payment_error_message(error, suffix):
    base = str(error).strip().rstrip('.')
    return f'{base}. {suffix}' if base else suffix


@app.route('/api/orders', methods=['POST'])
@limiter.limit("10 per minute; 50 per hour")
def create_order():
    data = request.json or {}
    customer_name = clean_text(data.get('customer_name'), 100)
    raw_customer_email = clean_text(data.get('customer_email'), 140).lower()
    customer_email = normalize_public_email(raw_customer_email)
    raw_customer_phone = clean_text(data.get('customer_phone'), 30)
    customer_phone = normalize_public_phone(raw_customer_phone) if raw_customer_phone else ''
    notes = clean_text(data.get('notes'), 1000)
    limited = public_abuse_guard(
        'create_order',
        email=customer_email or raw_customer_email,
        ip_limit=24,
        email_limit=10,
        window_seconds=3600,
    )
    if limited:
        return limited
    if not customer_name or not customer_email:
        return jsonify({'error': 'Name and email are required'}), 400
    if raw_customer_phone and not customer_phone:
        return jsonify({'error': 'Please enter a valid phone number'}), 400
    if public_text_looks_spammy(customer_name, customer_phone, notes):
        return jsonify({'error': 'Please remove unsupported characters and try again.'}), 400
    try:
        shipping_address = _normalize_shipping_address(data.get('shipping_address'))
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    checkout_hash = _checkout_context_fingerprint(customer_name, customer_email, customer_phone, shipping_address)
    # Orders must be paid: a verified Stripe payment is required (prevents fake/unpaid
    # orders that would still reserve stock and post to the books).
    payment_intent_id = clean_text(data.get('payment_intent_id'), 120)
    if not payment_intent_id:
        return jsonify({'error': 'Payment is required to place an order.'}), 400
    db = get_db()
    user_id = current_customer_id_from_request(db)
    if db.execute("SELECT id FROM orders WHERE payment_intent_id=?", (payment_intent_id,)).fetchone():
        log_security_event(
            'payment_reuse_blocked',
            'critical',
            'Duplicate payment intent rejected',
            email=customer_email,
            metadata={'payment_intent_id': payment_intent_id},
        )
        return jsonify({'error': 'This payment has already been used for an order. Please contact support if you need help.'}), 400
    try:
        totals = _calculate_order_totals(
            db,
            data.get('items'),
            data.get('coupon_code'),
            check_stock=False,
            customer_email=customer_email,
            user_id=user_id,
        )
    except ValueError as exc:
        refunded = _refund_succeeded_payment_intent(
            payment_intent_id,
            'Refunded paid checkout that failed cart validation before order save',
            email=customer_email,
        )
        if refunded:
            return jsonify({'error': _payment_error_message(exc, 'Your payment was refunded automatically.')}), 409
        return jsonify({'error': _payment_error_message(exc, 'Please refresh checkout and try again. If your card was charged, contact support.')}), 400
    stored_items = totals['stored_items']
    real_subtotal = totals['subtotal']
    discount = totals['discount']
    coupon_code = totals['coupon_code']
    coupon_obj = totals['coupon_obj']
    shipping = totals['shipping']
    total = totals['total']

    oid = str(uuid.uuid4())
    order_num = generate_order_number()

    # ── Server-side price verification — prevents client-side price manipulation ──
    try:
        intent = stripe.PaymentIntent.retrieve(payment_intent_id)
        if intent['status'] != 'succeeded':
            return jsonify({'error': 'Payment was not completed successfully. Please try again.'}), 400
        expected_cents = int(round(total * 100))
        if str(intent.get('currency', '')).lower() != 'usd':
            refunded = _refund_payment_intent(payment_intent_id, int(intent.get('amount') or 0), 'Refunded non-USD checkout attempt', email=customer_email)
            log_security_event(
                'payment_currency_mismatch',
                'critical',
                'Payment intent currency did not match USD checkout',
                email=customer_email,
                metadata={'payment_intent_id': payment_intent_id, 'currency': intent.get('currency')},
            )
            msg = 'Payment currency mismatch. The payment was refunded; please try again.' if refunded else 'Payment currency mismatch. Please contact support for refund help.'
            return jsonify({'error': msg}), 400
        if abs(intent['amount'] - expected_cents) > 2:
            app.logger.warning('Payment mismatch: intent=%d cents expected=%d cents', intent['amount'], expected_cents)
            refunded = _refund_payment_intent(payment_intent_id, int(intent['amount']), 'Refunded payment amount mismatch', email=customer_email)
            log_security_event(
                'payment_amount_mismatch',
                'critical',
                'Payment intent amount did not match server checkout total',
                email=customer_email,
                metadata={'payment_intent_id': payment_intent_id, 'stripe_amount': intent['amount'], 'expected_amount': expected_cents},
            )
            msg = 'Payment amount mismatch. The payment was refunded; please refresh checkout and try again.' if refunded else 'Payment amount mismatch. Please contact support for refund help.'
            return jsonify({'error': msg}), 400
        metadata = dict(intent.get('metadata') or {})
        if (
            metadata.get('source') != 'adhya_checkout'
            or metadata.get('cart_hash') != totals['cart_hash']
            or metadata.get('amount_cents') != str(expected_cents)
            or metadata.get('checkout_hash') != checkout_hash
        ):
            log_security_event(
                'payment_cart_mismatch',
                'critical',
                'Payment intent metadata did not match order checkout details',
                email=customer_email,
                metadata={'payment_intent_id': payment_intent_id},
            )
            refunded = _refund_payment_intent(payment_intent_id, int(intent['amount']), 'Refunded payment metadata/cart mismatch', email=customer_email)
            msg = 'Payment does not match this cart. The payment was refunded; please refresh checkout and try again.' if refunded else 'Payment does not match this cart. Please contact support for refund help.'
            return jsonify({'error': msg}), 400
    except stripe.error.StripeError:
        return jsonify({'error': 'Could not verify payment. Please contact support.'}), 400

    try:
        db.execute("BEGIN IMMEDIATE")
        if db.execute("SELECT id FROM orders WHERE payment_intent_id=?", (payment_intent_id,)).fetchone():
            db.rollback()
            log_security_event(
                'payment_reuse_blocked',
                'critical',
                'Duplicate payment intent rejected during order commit',
                email=customer_email,
                metadata={'payment_intent_id': payment_intent_id},
            )
            return jsonify({'error': 'This payment has already been used for an order. Please contact support if you need help.'}), 400

        locked_totals = _calculate_order_totals(
            db,
            data.get('items'),
            data.get('coupon_code'),
            check_stock=False,
            customer_email=customer_email,
            user_id=user_id,
        )
        locked_expected_cents = int(round(locked_totals['total'] * 100))
        if abs(intent['amount'] - locked_expected_cents) > 2 or locked_totals['cart_hash'] != totals['cart_hash']:
            db.rollback()
            _refund_payment_intent(payment_intent_id, int(intent['amount']), 'Refunded changed cart/payment mismatch', email=customer_email)
            return jsonify({'error': 'Your cart changed before the order could be saved. The payment was refunded; please refresh checkout and try again.'}), 409

        stored_items = locked_totals['stored_items']
        subtotal = locked_totals['subtotal']
        discount = locked_totals['discount']
        coupon_code = locked_totals['coupon_code']
        coupon_obj = locked_totals['coupon_obj']
        shipping = locked_totals['shipping']
        total = locked_totals['total']

        _reserve_order_stock(db, stored_items)

        # Only count a coupon as used after payment and stock have both been verified.
        if coupon_obj:
            db.execute("UPDATE coupons SET used_count=used_count+1 WHERE code=?", (coupon_code,))
        db.execute(
            """INSERT INTO orders (id,order_number,user_id,customer_name,customer_email,customer_phone,
               shipping_address,items,subtotal,discount,coupon_code,shipping_charge,total,
               payment_method,payment_status,payment_intent_id,notes)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (oid, order_num, user_id, customer_name, customer_email,
             customer_phone, json.dumps(shipping_address),
             json.dumps(stored_items), subtotal, discount, coupon_code,
             shipping, total, 'card',
             'paid', payment_intent_id, notes)
        )
        if coupon_code == WELCOME_COUPON_CODE:
            mark_welcome_discount_used(db, customer_email, user_id, oid)

        # Mirror this order into the bookkeeping module as an 'online' sale (non-fatal).
        try:
            if hasattr(app, 'acc_sync_order'):
                app.acc_sync_order(db, oid)
        except Exception as _e:
            app.logger.warning('accounts order-sync failed: %s', _e)

        db.commit()
    except ValueError as exc:
        db.rollback()
        refunded = _refund_payment_intent(payment_intent_id, int(intent['amount']), 'Refunded paid checkout that failed final stock validation', email=customer_email)
        if refunded:
            return jsonify({'error': _payment_error_message(exc, 'Your payment was refunded automatically.')}), 409
        return jsonify({'error': _payment_error_message(exc, 'Please contact support so we can refund this payment.')}), 409
    except sqlite3.Error:
        db.rollback()
        log_security_event(
            'order_commit_failed',
            'critical',
            'Database error while saving paid order',
            email=customer_email,
            metadata={'payment_intent_id': payment_intent_id},
        )
        return jsonify({'error': 'Could not save your order. Please contact support with your payment confirmation.'}), 500

    # This customer finished checking out — cancel any pending cart reminder.
    try:
        db.execute(
            "UPDATE abandoned_carts SET converted_at=datetime('now') WHERE email=? AND converted_at IS NULL",
            (customer_email,)
        )
        db.commit()
    except Exception as _ac_exc:
        app.logger.warning('abandoned-cart convert-mark failed: %s', _ac_exc)

    email_order_confirmation(
        order_num, customer_name, customer_email,
        stored_items, subtotal, discount,
        shipping, total, shipping_address
    )
    email_admin_new_order(
        order_num, customer_name, customer_email,
        stored_items, total, shipping_address
    )
    log_audit_event(
        'order_created',
        'order',
        oid,
        f'Checkout order created: {order_num}',
        after={'id': oid, 'order_number': order_num, 'total': total, 'payment_status': 'paid', 'status': 'pending'},
        metadata={'order_id': oid, 'order_number': order_num, 'total': total, 'customer_email': customer_email},
    )
    return jsonify({'id': oid, 'order_number': order_num, 'total': total, 'discount': discount})


@app.route('/api/orders/my', methods=['GET'])
@token_required
def my_orders():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM orders WHERE user_id=? ORDER BY created_at DESC", (g.user['id'],)
    ).fetchall()
    return jsonify([public_order_payload(o) for o in rows])


@app.route('/api/orders/<oid>', methods=['GET'])
@token_required
def get_order(oid):
    db = get_db()
    o = row_to_dict(db.execute("SELECT * FROM orders WHERE id=? OR order_number=?", (oid, oid)).fetchone())
    if not o:
        return jsonify({'error': 'Order not found'}), 404
    user = g.user
    if not _order_access_allowed(o):
        _log_order_access_denied('view_order', oid, o)
        return jsonify({'error': 'Order not found'}), 404
    o['items'] = json.loads(o['items'])
    o['shipping_address'] = json.loads(o['shipping_address'])
    if user.get('role') in ('admin', 'staff'):
        return jsonify(o)
    return jsonify(public_order_payload(o))


@app.route('/api/admin/orders', methods=['GET'])
@admin_required
def admin_orders():
    db = get_db()
    status = request.args.get('status')
    query = "SELECT * FROM orders"
    params = []
    if status:
        query += " WHERE status=?"
        params.append(status)
    query += " ORDER BY created_at DESC"
    orders = rows_to_list(db.execute(query, params).fetchall())
    for o in orders:
        o['items'] = json.loads(o['items'])
        o['shipping_address'] = json.loads(o['shipping_address'])
    return jsonify(orders)


@app.route('/api/orders/<oid>/cancel', methods=['POST'])
@limiter.limit("6 per minute; 30 per hour")
@token_required
def cancel_order(oid):
    db = get_db()
    order = row_to_dict(db.execute("SELECT * FROM orders WHERE id=? OR order_number=?", (oid, oid)).fetchone())
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    if not _order_access_allowed(order):
        _log_order_access_denied('cancel_order', oid, order)
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] not in ('pending', 'processing'):
        return jsonify({'error': 'This order has already been shipped and cannot be cancelled directly. Please use the Return Request option instead.'}), 400

    items = json.loads(order['items'])
    payment_intent_id = order.get('payment_intent_id', '')
    amount_cents = int(round(float(order['total']) * 100))

    try:
        db.execute("BEGIN IMMEDIATE")
        fresh = row_to_dict(db.execute("SELECT status,payment_status FROM orders WHERE id=?", (order['id'],)).fetchone())
        if not fresh or fresh['status'] not in ('pending', 'processing'):
            db.rollback()
            return jsonify({'error': 'This order can no longer be cancelled directly.'}), 400
        if fresh['payment_status'] in ('refunded', 'refund_pending'):
            db.rollback()
            return jsonify({'error': 'A refund is already completed or pending for this order.'}), 400
        restocked_product_ids = _restore_order_stock(db, items)
        db.execute("UPDATE orders SET status='cancelled', payment_status='refund_pending', updated_at=datetime('now') WHERE id=?",
                   (order['id'],))
        # Remove the mirrored bookkeeping sale for this cancelled order (non-fatal).
        try:
            if hasattr(app, 'acc_void_order_sale'):
                app.acc_void_order_sale(db, order['id'])
        except Exception as _e:
            app.logger.warning('accounts order-void failed: %s', _e)
        db.commit()
    except sqlite3.Error:
        db.rollback()
        log_security_event('order_cancel_failed', 'critical', 'Database error while cancelling order', email=order['customer_email'], metadata={'order_id': order['id']})
        return jsonify({'error': 'Could not cancel the order. Please contact support.'}), 500

    for product_id in restocked_product_ids:
        try:
            notify_back_in_stock_for_product(db, product_id)
        except Exception as exc:
            app.logger.warning('back-in-stock notify after cancellation failed: %s', exc)

    pay_status = 'refund_pending'
    if payment_intent_id and _refund_payment_intent(payment_intent_id, amount_cents, 'Order cancellation refund', email=order['customer_email']):
        pay_status = 'refunded'
        db.execute("UPDATE orders SET payment_status='refunded', updated_at=datetime('now') WHERE id=?", (order['id'],))
        db.commit()

    log_audit_event(
        'order_cancelled',
        'order',
        order['id'],
        f"Order cancelled: {order['order_number']}",
        before={'status': order['status'], 'payment_status': order['payment_status']},
        after={'status': 'cancelled', 'payment_status': pay_status},
        metadata={'order_id': order['id'], 'order_number': order['order_number'], 'cancelled_by_role': g.user.get('role')},
    )
    email_order_status(order['order_number'], order['customer_name'], order['customer_email'], 'cancelled')
    msg = 'Order cancelled. Refund will appear on your card within 5–7 business days.' if pay_status == 'refunded' \
          else 'Order cancelled. Our team will process your refund within 2 business days.'
    return jsonify({'message': msg})


@app.route('/api/orders/<oid>/request-return', methods=['POST'])
@limiter.limit("6 per minute; 30 per hour")
@token_required
def request_return(oid):
    data = request.json or {}
    reason = clean_text(data.get('reason'), 1000)
    db = get_db()
    order = row_to_dict(db.execute("SELECT * FROM orders WHERE id=? OR order_number=?", (oid, oid)).fetchone())
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    if order['user_id'] != g.user['id']:
        _log_order_access_denied('request_return', oid, order)
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] not in ('shipped', 'delivered'):
        return jsonify({'error': 'Return can only be requested for shipped or delivered orders.'}), 400
    if len(reason) < 10:
        return jsonify({'error': 'Please add a short reason for the return request.'}), 400
    if public_text_looks_spammy(reason):
        return jsonify({'error': 'Please remove links or unsupported content from the return reason.'}), 400
    db.execute(
        "UPDATE orders SET status='return_requested', return_reason=?, updated_at=datetime('now') WHERE id=?",
        (reason, order['id'])
    )
    db.commit()
    order['return_reason'] = reason
    log_audit_event(
        'return_requested',
        'order',
        order['id'],
        f"Return requested: {order['order_number']}",
        before={'status': order['status']},
        after={'status': 'return_requested'},
        metadata={'order_id': order['id'], 'order_number': order['order_number'], 'reason_length': len(reason)},
    )
    email_order_status(order['order_number'], order['customer_name'], order['customer_email'], 'return_requested')
    email_admin_return_request(order)
    return jsonify({'message': 'Return request submitted. Please ship the package back within 7 business days. We will email you the return address.'})


@app.route('/api/admin/orders/<oid>/process-return', methods=['POST'])
@limiter.limit("8 per minute; 40 per hour")
@admin_required
def process_return(oid):
    data = request.json or {}
    db = get_db()
    order = row_to_dict(db.execute("SELECT * FROM orders WHERE id=?", (oid,)).fetchone())
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] != 'return_requested':
        return jsonify({'error': 'Only orders with an active return request can be processed.'}), 400
    if order['payment_status'] in ('refunded', 'refund_pending'):
        return jsonify({'error': 'A refund is already completed or pending for this order.'}), 400
    refund_type = clean_text(data.get('refund_type'), 12).lower() or 'full'
    if refund_type not in ('full', 'half'):
        return jsonify({'error': 'Invalid refund type'}), 400
    payment_intent_id = order.get('payment_intent_id', '')
    total_cents = int(round(float(order['total']) * 100))
    amount = total_cents if refund_type == 'full' else total_cents // 2
    refund_label = 'Full refund' if refund_type == 'full' else '50% refund'

    try:
        db.execute("BEGIN IMMEDIATE")
        fresh = row_to_dict(db.execute("SELECT status,payment_status FROM orders WHERE id=?", (oid,)).fetchone())
        if not fresh or fresh['status'] != 'return_requested':
            db.rollback()
            return jsonify({'error': 'Only orders with an active return request can be processed.'}), 400
        if fresh['payment_status'] in ('refunded', 'refund_pending'):
            db.rollback()
            return jsonify({'error': 'A refund is already completed or pending for this order.'}), 400
        db.execute("UPDATE orders SET status='return_received', payment_status='refund_pending', updated_at=datetime('now') WHERE id=?", (oid,))
        db.commit()
    except sqlite3.Error:
        db.rollback()
        log_security_event('return_process_failed', 'critical', 'Database error while processing return', email=order['customer_email'], metadata={'order_id': oid})
        return jsonify({'error': 'Could not process the return. Please try again.'}), 500

    if not payment_intent_id:
        db.execute("UPDATE orders SET status='return_received', payment_status='refund_pending', updated_at=datetime('now') WHERE id=?", (oid,))
        db.commit()
        email_order_status(order['order_number'], order['customer_name'], order['customer_email'], 'return_received')
        log_admin_action('return_processed', 'Return marked as received for manual refund', {
            'order_id': oid,
            'entity_type': 'order',
            'entity_id': oid,
            'before': {'status': order['status'], 'payment_status': order['payment_status']},
            'after': {'status': 'return_received', 'payment_status': 'refund_pending'},
        })
        return jsonify({'message': 'Return marked as received. Process refund manually in Stripe.'})

    if _refund_payment_intent(payment_intent_id, amount, f'{refund_label} for returned order', email=order['customer_email']):
        db.execute("UPDATE orders SET status='return_received', payment_status='refunded', updated_at=datetime('now') WHERE id=?", (oid,))
        db.commit()
        email_order_status(order['order_number'], order['customer_name'], order['customer_email'], 'refunded')
        log_admin_action('return_processed', f'{refund_label} processed', {
            'order_id': oid,
            'entity_type': 'order',
            'entity_id': oid,
            'refund_type': refund_type,
            'amount_cents': amount,
            'before': {'status': order['status'], 'payment_status': order['payment_status']},
            'after': {'status': 'return_received', 'payment_status': 'refunded'},
        })
        return jsonify({'message': f'{refund_label} of ${amount/100:.2f} processed successfully.'})
    return jsonify({'error': 'Return marked received, but Stripe refund could not be completed. Process the refund manually in Stripe.'}), 400


def _allowed_admin_order_transition(old_status, new_status):
    if new_status == old_status:
        return True
    allowed = {
        'pending': {'processing', 'shipped'},
        'processing': {'shipped', 'delivered'},
        'shipped': {'delivered', 'return_requested'},
        'delivered': {'return_requested'},
        'return_requested': set(),
        'cancelled': set(),
        'return_received': set(),
    }
    return new_status in allowed.get(old_status or 'pending', set())


def _allowed_admin_payment_transition(old_payment_status, new_payment_status, old_status):
    if new_payment_status == old_payment_status:
        return True
    if new_payment_status in ('refunded', 'refund_pending'):
        return old_status in ('cancelled', 'return_received')
    if old_payment_status in ('refunded', 'refund_pending'):
        return new_payment_status in ('refunded', 'refund_pending')
    if old_payment_status == 'paid':
        return False
    return old_payment_status in ('pending', 'failed') and new_payment_status in ('pending', 'paid', 'failed')


def _order_items_for_email(order):
    try:
        return json.loads(order.get('items') or '[]')
    except Exception:
        return []


def _order_shipping_for_email(order):
    try:
        return json.loads(order.get('shipping_address') or '{}')
    except Exception:
        return {}


def send_review_request_for_order(db, order, *, force=False):
    if not order:
        return False, 'Order not found'
    if order.get('status') != 'delivered':
        return False, 'Review requests can only be sent after delivery.'
    if order.get('review_requested_at') and not force:
        return False, 'Review request was already sent.'
    items = _order_items_for_email(order)
    email_review_request(
        order['order_number'],
        order['customer_name'],
        order['customer_email'],
        items,
    )
    db.execute(
        "UPDATE orders SET review_requested_at=COALESCE(review_requested_at, datetime('now')), updated_at=datetime('now') WHERE id=?",
        (order['id'],)
    )
    db.commit()
    return True, 'Review request email sent.'


@app.route('/api/admin/orders/<oid>/email/<kind>', methods=['POST'])
@limiter.limit("20 per minute; 120 per hour")
@admin_required
def admin_send_order_email(oid, kind):
    kind = clean_text(kind, 40).lower()
    if kind not in ('confirmation', 'status', 'review'):
        return jsonify({'error': 'Email type not found'}), 404

    db = get_db()
    order = row_to_dict(db.execute("SELECT * FROM orders WHERE id=? OR order_number=?", (oid, oid)).fetchone())
    if not order:
        return jsonify({'error': 'Order not found'}), 404

    items = _order_items_for_email(order)
    shipping_address = _order_shipping_for_email(order)

    if kind == 'confirmation':
        email_order_confirmation(
            order['order_number'],
            order['customer_name'],
            order['customer_email'],
            items,
            order['subtotal'],
            order['discount'],
            order['shipping_charge'],
            order['total'],
            shipping_address,
        )
        message = 'Order confirmation email resent.'
    elif kind == 'status':
        email_order_status(
            order['order_number'],
            order['customer_name'],
            order['customer_email'],
            order['status'],
            order.get('tracking_number'),
        )
        message = 'Order status email sent.'
    else:
        sent, message = send_review_request_for_order(db, order, force=True)
        if not sent:
            return jsonify({'error': message}), 400

    log_admin_action('order_email_sent', message, {
        'order_id': order['id'],
        'entity_type': 'order',
        'entity_id': order['id'],
        'order_number': order['order_number'],
        'email_type': kind,
        'customer_email': order['customer_email'],
    })
    return jsonify({'message': message})


@app.route('/api/admin/orders/<oid>', methods=['PUT'])
@admin_required
def update_order(oid):
    data = request.json or {}
    allowed_statuses = {'pending', 'processing', 'shipped', 'delivered', 'cancelled', 'return_requested', 'return_received'}
    allowed_payment = {'pending', 'paid', 'failed', 'refunded', 'refund_pending'}
    status = clean_text(data.get('status'), 40)
    payment_status = clean_text(data.get('payment_status'), 40)
    if status not in allowed_statuses:
        return jsonify({'error': 'Invalid order status'}), 400
    if payment_status not in allowed_payment:
        return jsonify({'error': 'Invalid payment status'}), 400
    tracking = clean_text(data.get('tracking_number'), 80)
    notes = clean_text(data.get('notes'), 1000)
    db = get_db()
    order = row_to_dict(db.execute("SELECT * FROM orders WHERE id=?", (oid,)).fetchone())
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    old_status = order.get('status')
    old_payment_status = order.get('payment_status')
    before = fetch_order_for_audit(db, oid)
    if status == 'cancelled' and old_status != 'cancelled':
        return jsonify({'error': 'Use the Cancel & Refund action so stock, refund status, bookkeeping, and customer email are handled safely.'}), 400
    if status == 'return_received' and old_status != 'return_received':
        return jsonify({'error': 'Use the Process Return action so refund status and Stripe refund handling are safe.'}), 400
    if not _allowed_admin_order_transition(old_status, status):
        return jsonify({'error': f'Invalid order status flow: {old_status} to {status}'}), 400
    if payment_status in ('refunded', 'refund_pending') and old_payment_status not in ('refunded', 'refund_pending'):
        if old_status not in ('cancelled', 'return_received'):
            return jsonify({'error': 'Refund status can only be applied after a cancellation or processed return flow.'}), 400
    if not _allowed_admin_payment_transition(old_payment_status, payment_status, old_status):
        return jsonify({'error': f'Invalid payment status flow: {old_payment_status} to {payment_status}'}), 400
    if status == 'shipped' and not tracking:
        return jsonify({'error': 'Tracking number is required before marking an order shipped.'}), 400
    db.execute(
        "UPDATE orders SET status=?,payment_status=?,tracking_number=?,notes=?,updated_at=datetime('now') WHERE id=?",
        (status, payment_status, tracking, notes, oid)
    )
    db.commit()
    after = fetch_order_for_audit(db, oid)
    log_admin_action('order_updated', 'Order status/payment updated', {
        'order_id': oid,
        'entity_type': 'order',
        'entity_id': oid,
        'status': status,
        'payment_status': payment_status,
        'before': before,
        'after': after,
    })
    new_status = status
    if new_status and new_status != order.get('status'):
        email_order_status(
            order['order_number'], order['customer_name'], order['customer_email'],
            new_status, tracking or order.get('tracking_number')
        )
        if new_status == 'delivered':
            fresh_order = row_to_dict(db.execute("SELECT * FROM orders WHERE id=?", (oid,)).fetchone())
            sent_review, review_message = send_review_request_for_order(db, fresh_order, force=False)
            if sent_review:
                log_admin_action('review_request_sent', 'Review request sent after delivery', {
                    'order_id': oid,
                    'entity_type': 'order',
                    'entity_id': oid,
                    'order_number': order['order_number'],
                    'mode': 'automatic',
                })
            else:
                log_security_event(
                    'review_request_skipped',
                    'info',
                    review_message,
                    user_id=g.user.get('id'),
                    email=g.user.get('email'),
                    metadata={'order_id': oid, 'order_number': order['order_number'], 'mode': 'automatic'},
                )
    return jsonify({'message': 'Order updated'})


# ─── Coupons ──────────────────────────────────────────────────────────────────

@app.route('/api/coupons/validate', methods=['POST'])
@limiter.limit("30 per minute; 200 per hour")
def validate_coupon():
    data = request.json or {}
    code = clean_text(data.get('code'), 40).upper()
    limited = public_abuse_guard('coupon_validate', fingerprint=code, ip_limit=80, window_seconds=600)
    if limited:
        return limited
    if not code:
        return jsonify({'error': 'Coupon code is required'}), 400
    if not re.fullmatch(r'[A-Z0-9_-]{2,40}', code):
        public_abuse_guard('coupon_invalid', fingerprint=code, ip_limit=20, fingerprint_limit=8, window_seconds=900)
        return jsonify({'error': 'Invalid or expired coupon'}), 400
    db = get_db()
    if isinstance(data.get('items'), list) and data.get('items'):
        try:
            _stored_items, subtotal = _normalize_order_items(db, data.get('items'), check_stock=True)
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400
    else:
        try:
            subtotal = max(0.0, float(data.get('subtotal') or 0))
        except (TypeError, ValueError):
            subtotal = 0.0
    coupon = row_to_dict(db.execute(
        "SELECT * FROM coupons WHERE code=? AND is_active=1", (code,)
    ).fetchone())
    if not coupon:
        limited = public_abuse_guard('coupon_invalid', fingerprint=code, ip_limit=20, fingerprint_limit=8, window_seconds=900)
        if limited:
            return limited
        return jsonify({'error': 'Invalid or expired coupon'}), 400
    if coupon['expires_at'] and coupon['expires_at'] < datetime.datetime.now().isoformat():
        return jsonify({'error': 'Coupon expired'}), 400
    if coupon['max_uses'] and coupon['used_count'] >= coupon['max_uses']:
        return jsonify({'error': 'Coupon usage limit reached'}), 400
    if subtotal < coupon['min_order']:
        return jsonify({'error': f"Minimum order ${coupon['min_order']:.2f} required"}), 400
    try:
        validate_welcome_coupon_for_checkout(
            db,
            code,
            customer_email=data.get('customer_email') or '',
            user_id=current_customer_id_from_request(db),
        )
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    if coupon['discount_type'] == 'percent':
        discount_amount = min((subtotal * coupon['discount_value']) / 100, subtotal)
    else:
        discount_amount = min(coupon['discount_value'], subtotal)
    return jsonify({
        'code': coupon['code'],
        'discount_type': coupon['discount_type'],
        'discount_value': coupon['discount_value'],
        'min_order': coupon['min_order'],
        'subtotal': round(subtotal, 2),
        'discount_amount': round(discount_amount, 2),
    })


@app.route('/api/welcome-discount/status', methods=['GET'])
@token_required
def welcome_discount_status():
    if g.user.get('role') != 'customer':
        return jsonify({'available': False, 'issued': False, 'code': '', 'message': 'Staff accounts are not eligible for customer welcome offers.'})
    db = get_db()
    return jsonify(welcome_discount_status_payload(db, g.user.get('email'), g.user.get('id')))


@app.route('/api/admin/coupons', methods=['GET'])
@admin_only_required
def get_coupons():
    db = get_db()
    return jsonify(rows_to_list(db.execute("SELECT * FROM coupons ORDER BY created_at DESC").fetchall()))


@app.route('/api/admin/coupons', methods=['POST'])
@admin_only_required
def create_coupon():
    data = request.json or {}
    code = clean_text(data.get('code'), 40).upper()
    discount_type = clean_text(data.get('discount_type'), 20).lower()
    if not code or not re.fullmatch(r'[A-Z0-9_-]{2,40}', code):
        return jsonify({'error': 'Coupon code must use letters, numbers, hyphen, or underscore'}), 400
    if discount_type not in ('percent', 'fixed'):
        return jsonify({'error': 'Invalid discount type'}), 400
    try:
        discount_value = float(data.get('discount_value'))
        min_order = max(0.0, float(data.get('min_order') or 0))
    except (TypeError, ValueError):
        return jsonify({'error': 'Discount and minimum order must be valid amounts'}), 400
    if discount_value <= 0:
        return jsonify({'error': 'Discount must be greater than 0'}), 400
    if discount_type == 'percent' and discount_value > 100:
        return jsonify({'error': 'Percent discount cannot be more than 100'}), 400
    max_uses = data.get('max_uses')
    if max_uses in ('', None):
        max_uses = None
    else:
        try:
            max_uses = max(1, int(max_uses))
        except (TypeError, ValueError):
            return jsonify({'error': 'Max uses must be a whole number'}), 400
    expires_at = clean_text(data.get('expires_at'), 40) or None
    db = get_db()
    if db.execute("SELECT id FROM coupons WHERE code=?", (code,)).fetchone():
        return jsonify({'error': 'A coupon with that code already exists'}), 400
    cid = str(uuid.uuid4())
    db.execute(
        "INSERT INTO coupons (id,code,discount_type,discount_value,min_order,max_uses,expires_at) VALUES (?,?,?,?,?,?,?)",
        (cid, code, discount_type, discount_value, min_order, max_uses, expires_at)
    )
    db.commit()
    log_admin_action('coupon_created', 'Coupon created', {
        'coupon_id': cid,
        'entity_type': 'coupon',
        'entity_id': cid,
        'code': code,
        'after': {'id': cid, 'code': code, 'discount_type': discount_type, 'discount_value': discount_value, 'min_order': min_order, 'max_uses': max_uses, 'expires_at': expires_at},
    })
    return jsonify({'id': cid, 'message': 'Coupon created'})


@app.route('/api/admin/coupons/<cid>', methods=['DELETE'])
@admin_only_required
def delete_coupon(cid):
    db = get_db()
    before = row_to_dict(db.execute("SELECT * FROM coupons WHERE id=?", (cid,)).fetchone())
    if not before:
        return jsonify({'error': 'Coupon not found'}), 404
    db.execute("DELETE FROM coupons WHERE id=?", (cid,))
    db.commit()
    log_admin_action('coupon_deleted', 'Coupon deleted', {
        'coupon_id': cid,
        'entity_type': 'coupon',
        'entity_id': cid,
        'before': before,
    })
    return jsonify({'message': 'Coupon deleted'})


# ─── Users ────────────────────────────────────────────────────────────────────

@app.route('/api/admin/users', methods=['GET'])
@admin_only_required
def get_users():
    db = get_db()
    users = rows_to_list(db.execute(
        """SELECT u.id,u.name,u.email,u.phone,u.role,u.created_at,
                  COUNT(o.id) AS order_count,
                  COALESCE(SUM(CASE WHEN o.payment_status IN ('paid','refund_pending','refunded') THEN o.total ELSE 0 END),0) AS total_spent,
                  MAX(o.created_at) AS last_order_at,
                  COALESCE(SUM(CASE WHEN o.status='return_requested' THEN 1 ELSE 0 END),0) AS return_request_count
           FROM users u
           LEFT JOIN orders o
             ON o.user_id=u.id
             OR LOWER(o.customer_email)=LOWER(u.email)
           GROUP BY u.id
           ORDER BY u.created_at DESC"""
    ).fetchall())
    return jsonify(users)


@app.route('/api/admin/users', methods=['POST'])
@admin_only_required
def create_user():
    data = request.json or {}
    name = clean_text(data.get('name'), 100)
    email = clean_text(data.get('email'), 140).lower()
    password = data.get('password') or ''
    if not name or not email or not password:
        return jsonify({'error': 'Name, email and password are required'}), 400
    pw_error = validate_password(password)
    if pw_error:
        return jsonify({'error': pw_error}), 400
    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
    if existing:
        return jsonify({'error': 'A user with that email already exists'}), 400
    role = data.get('role', 'customer')
    if role not in ('customer', 'staff', 'admin'):
        role = 'customer'
    uid = str(uuid.uuid4())
    hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    db.execute("INSERT INTO users (id,name,email,phone,password,role) VALUES (?,?,?,?,?,?)",
               (uid, name, email, clean_text(data.get('phone'), 30), hashed, role))
    db.commit()
    log_admin_action('user_created', f'Created {role} user', {
        'target_user_id': uid,
        'entity_type': 'user',
        'entity_id': uid,
        'target_email': email,
        'role': role,
        'after': fetch_user_for_audit(db, uid),
    })
    return jsonify({'id': uid, 'message': 'User created successfully'})


@app.route('/api/admin/users/<uid>', methods=['PUT'])
@admin_only_required
def update_user(uid):
    data = request.json or {}
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    if not email or '@' not in email:
        return jsonify({'error': 'A valid email is required'}), 400
    db = get_db()
    user = db.execute("SELECT id, email, role FROM users WHERE id=?", (uid,)).fetchone()
    if not user:
        return jsonify({'error': 'User not found'}), 404
    before = fetch_user_for_audit(db, uid)
    # Prevent admin from accidentally demoting themselves
    if uid == g.user['id'] and data.get('role') != 'admin':
        return jsonify({'error': 'You cannot change your own role'}), 400
    role = data.get('role', user['role'])
    if role not in ('customer', 'staff', 'admin'):
        role = 'customer'
    session_revoked = role != user['role'] or email != user['email']
    db.execute("UPDATE users SET name=?,email=?,phone=?,role=? WHERE id=?",
               (name, email, clean_text(data.get('phone'), 30), role, uid))
    if session_revoked:
        db.execute("UPDATE users SET token_version=COALESCE(token_version,0)+1 WHERE id=?", (uid,))
    if data.get('password'):
        pw_error = validate_password(data['password'])
        if pw_error:
            return jsonify({'error': pw_error}), 400
        hashed = bcrypt.hashpw(data['password'].encode(), bcrypt.gensalt()).decode()
        db.execute("UPDATE users SET password=?, token_version=COALESCE(token_version,0)+1 WHERE id=?", (hashed, uid))
    db.commit()
    log_admin_action('user_updated', 'User account updated', {
        'target_user_id': uid,
        'entity_type': 'user',
        'entity_id': uid,
        'target_email': email,
        'role': role,
        'password_changed': bool(data.get('password')),
        'before': before,
        'after': fetch_user_for_audit(db, uid),
    })
    response = jsonify({'message': 'User updated successfully'})
    if uid == g.user['id'] and (data.get('password') or session_revoked):
        return clear_auth_cookie(response)
    return response


@app.route('/api/admin/users/<uid>', methods=['DELETE'])
@admin_only_required
def delete_user(uid):
    db = get_db()
    if uid == g.user['id']:
        return jsonify({'error': 'You cannot delete your own account'}), 400
    before = fetch_user_for_audit(db, uid)
    if not before:
        return jsonify({'error': 'User not found'}), 404
    db.execute("UPDATE orders SET user_id=NULL WHERE user_id=?", (uid,))
    db.execute("DELETE FROM user_wishlist WHERE user_id=?", (uid,))
    db.execute("DELETE FROM reviews WHERE user_id=?", (uid,))
    db.execute("DELETE FROM users WHERE id=?", (uid,))
    db.commit()
    log_admin_action('user_deleted', 'User account deleted', {
        'target_user_id': uid,
        'entity_type': 'user',
        'entity_id': uid,
        'before': before,
    })
    return jsonify({'message': 'User deleted'})


@app.route('/api/user/profile', methods=['PUT'])
@token_required
def update_profile():
    data = request.json
    name = clean_text(data.get('name'), 100)
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    name_error = validate_name(name)
    if name_error:
        return jsonify({'error': name_error}), 400
    raw_phone = clean_text(data.get('phone'), 30)
    phone = normalize_public_phone(raw_phone) if raw_phone else ''
    if raw_phone and not phone:
        return jsonify({'error': 'Please enter a valid phone number'}), 400
    addr = data.get('address') if isinstance(data.get('address'), dict) else {}
    clean_addr = {
        'line1': clean_text(addr.get('line1'), 180),
        'city': clean_text(addr.get('city'), 80),
        'state': clean_text(addr.get('state'), 80),
        'pin': clean_text(addr.get('pin') or addr.get('zip'), 12),
        'landmark': clean_text(addr.get('landmark'), 120),
    }
    if clean_addr['pin'] and not re.fullmatch(r'\d{5}(?:-\d{4})?', clean_addr['pin']):
        return jsonify({'error': 'Please enter a valid US ZIP code'}), 400
    db = get_db()
    db.execute("UPDATE users SET name=?,phone=?,address=? WHERE id=?",
               (name, phone, json.dumps(clean_addr), g.user['id']))
    db.commit()
    return jsonify({'message': 'Profile updated'})


# ─── Reviews ──────────────────────────────────────────────────────────────────

@app.route('/api/products/<pid>/reviews', methods=['POST'])
@token_required
def add_review(pid):
    data = request.json or {}
    limited = public_abuse_guard(
        'review_submit',
        email=g.user.get('email'),
        fingerprint=pid,
        ip_limit=20,
        email_limit=10,
        window_seconds=3600,
    )
    if limited:
        return limited
    try:
        rating = int(data.get('rating'))
    except (TypeError, ValueError):
        rating = 0
    if rating < 1 or rating > 5:
        return jsonify({'error': 'A rating of 1 to 5 stars is required'}), 400
    db = get_db()
    product = db.execute("SELECT id FROM products WHERE id=? AND is_active=1", (pid,)).fetchone()
    if not product:
        return jsonify({'error': 'Product not found'}), 404
    raw_images = data.get('images') or []
    if not isinstance(raw_images, list):
        raw_images = []
    images = [clean_text(u, 260) for u in raw_images[:3] if is_safe_public_upload_url(u)]
    images_json = json.dumps(images)
    comment = clean_text(data.get('comment'), 1200)
    if public_text_looks_spammy(comment):
        log_security_event(
            'review_rejected',
            'warning',
            'Review text looked unsafe or spammy',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
            metadata={'product_id': pid},
        )
        return jsonify({'error': 'Please remove unsupported text and try again.'}), 400
    existing = db.execute("SELECT id FROM reviews WHERE product_id=? AND user_id=?",
                          (pid, g.user['id'])).fetchone()
    if existing:
        db.execute("UPDATE reviews SET rating=?, comment=?, images=? WHERE id=?",
                   (rating, comment, images_json, existing['id']))
    else:
        rid = str(uuid.uuid4())
        db.execute("INSERT INTO reviews (id,product_id,user_id,rating,comment,images) VALUES (?,?,?,?,?,?)",
                   (rid, pid, g.user['id'], rating, comment, images_json))
    db.commit()
    return jsonify({'message': 'Review saved'})


@app.route('/api/admin/reviews', methods=['GET'])
@admin_required
def admin_list_reviews():
    db = get_db()
    rows = db.execute("""
        SELECT r.id, r.rating, r.comment, r.created_at,
               u.name AS user_name, u.email AS user_email,
               p.name AS product_name
        FROM reviews r
        JOIN users u ON u.id = r.user_id
        JOIN products p ON p.id = r.product_id
        ORDER BY r.created_at DESC
    """).fetchall()
    return jsonify(rows_to_list(rows))


@app.route('/api/admin/reviews/<rid>', methods=['DELETE'])
@admin_required
def admin_delete_review(rid):
    db = get_db()
    if not db.execute("SELECT id FROM reviews WHERE id=?", (rid,)).fetchone():
        return jsonify({'error': 'Review not found'}), 404
    db.execute("DELETE FROM reviews WHERE id=?", (rid,))
    db.commit()
    log_admin_action('review_deleted', 'Review deleted', {'review_id': rid})
    return jsonify({'message': 'Review deleted'})


@app.route('/api/reviews/featured', methods=['GET'])
def featured_reviews():
    db = get_db()
    rows = db.execute("""
        SELECT r.rating, r.comment, r.created_at, u.name AS user_name, p.name AS product_name
        FROM reviews r
        JOIN users u ON u.id = r.user_id
        JOIN products p ON p.id = r.product_id
        WHERE r.rating >= 4 AND r.comment IS NOT NULL AND trim(r.comment) != ''
        ORDER BY r.rating DESC, r.created_at DESC
        LIMIT 6
    """).fetchall()
    return jsonify(rows_to_list(rows))


@app.route('/api/stats/public', methods=['GET'])
def public_stats():
    db = get_db()
    orders_shipped = db.execute(
        "SELECT COUNT(*) FROM orders WHERE status IN ('shipped','delivered') AND payment_status='paid'"
    ).fetchone()[0]
    happy_customers = db.execute(
        "SELECT COUNT(DISTINCT customer_email) FROM orders WHERE status IN ('shipped','delivered') AND payment_status='paid'"
    ).fetchone()[0]
    return jsonify({
        'orders_shipped': max(orders_shipped, 200),
        'happy_customers': max(happy_customers, 150),
    })


@app.route('/api/newsletter/subscribe', methods=['POST'])
@limiter.limit("3 per minute; 10 per hour")
def newsletter_subscribe():
    data = request.json or {}
    raw_email = (data.get('email') or '').strip().lower()
    email = normalize_public_email(raw_email)
    limited = public_abuse_guard('newsletter_subscribe', email=email or raw_email, ip_limit=8, email_limit=3, window_seconds=3600)
    if limited:
        return limited
    if not email:
        return jsonify({'error': 'Valid email required'}), 400
    db = get_db()
    already = db.execute("SELECT 1 FROM newsletter_subscribers WHERE email=?", (email,)).fetchone()
    if not already:
        db.execute("INSERT INTO newsletter_subscribers (id,email) VALUES (?,?)",
                   (str(uuid.uuid4()), email))
    welcome = ensure_welcome_discount(db, email, issued_via='newsletter', send=True)
    db.commit()
    status = welcome_discount_status_payload(db, email=email)
    if welcome.get('emailed'):
        message = 'Check your email for your first-order 10% code.'
    elif status.get('used') or status.get('has_order'):
        message = 'You are already subscribed. The welcome discount is only for the first order.'
    else:
        message = 'That email already has a welcome code. Check your inbox or sign in to use it.'
    return jsonify({
        'message': message,
        'welcome_discount': {
            'code': WELCOME_COUPON_CODE if status.get('available') else '',
            'emailed': bool(welcome.get('emailed')),
            'already_issued': bool(welcome.get('already_issued')),
            'available': bool(status.get('available')),
        },
    })


# ─── SEO: Sitemap & Robots ────────────────────────────────────────────────────

@app.route('/sitemap.xml')
def sitemap():
    base = 'https://adhyashaktishop.com'
    pages = [
        ('/', '1.0', 'daily'),
        ('/products', '0.9', 'daily'),
        ('/jewelry', '0.9', 'weekly'),
        ('/clothing', '0.9', 'weekly'),
        ('/custom-printing', '0.8', 'weekly'),
        ('/about', '0.6', 'monthly'),
        ('/contact', '0.6', 'monthly'),
        ('/faq', '0.6', 'monthly'),
        ('/bulk-orders', '0.7', 'monthly'),
        ('/coming-soon', '0.5', 'monthly'),
        ('/track-order', '0.5', 'monthly'),
        ('/terms', '0.4', 'monthly'),
        ('/privacy', '0.4', 'monthly'),
        ('/refund', '0.4', 'monthly'),
    ]
    today = datetime.datetime.utcnow().strftime('%Y-%m-%d')
    entries = [
        (f'{base}{p}', today, freq, pri)
        for p, pri, freq in pages
    ]
    try:
        db = get_db()
        for row in db.execute("SELECT id,created_at AS lastmod FROM products WHERE is_active=1 ORDER BY created_at DESC LIMIT 500").fetchall():
            entries.append((f'{base}/product/{row["id"]}', (row['lastmod'] or today)[:10], 'weekly', '0.8'))
        for row in db.execute("""
            SELECT id,created_at AS lastmod
            FROM categories
            WHERE IFNULL(is_active,1)=1
              AND IFNULL(kind,'catalog') IN ('catalog','clothing')
            ORDER BY created_at DESC
            LIMIT 200
        """).fetchall():
            entries.append((f'{base}/products?category={row["id"]}', (row['lastmod'] or today)[:10], 'weekly', '0.6'))
    except Exception as exc:
        app.logger.warning('dynamic sitemap entries failed: %s', exc)
    urls = '\n'.join(
        f'  <url><loc>{escape(str(loc), quote=True)}</loc><lastmod>{escape(str(lastmod), quote=True)}</lastmod>'
        f'<changefreq>{escape(str(freq), quote=True)}</changefreq><priority>{escape(str(pri), quote=True)}</priority></url>'
        for loc, lastmod, freq, pri in entries
    )
    xml = f'<?xml version="1.0" encoding="UTF-8"?>\n<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n{urls}\n</urlset>'
    return Response(xml, mimetype='application/xml')


@app.route('/robots.txt')
def robots():
    content = (
        'User-agent: *\n'
        'Allow: /\n'
        'Disallow: /api/\n'
        'Disallow: /admin\n'
        'Disallow: /dashboard\n'
        'Disallow: /checkout\n'
        'Disallow: /cart\n'
        'Disallow: /wishlist\n'
        'Disallow: /login\n'
        'Disallow: /register\n'
        'Disallow: /forgot-password\n'
        'Disallow: /reset-password\n'
        'Disallow: /order-success\n'
        'Disallow: /track-order\n\n'
        'Sitemap: https://adhyashaktishop.com/sitemap.xml\n'
    )
    return Response(content, mimetype='text/plain')


@app.route('/google-shopping-feed.xml')
@limiter.limit("60 per hour")
def google_shopping_feed():
    """Google Merchant Center product feed (RSS 2.0 + g: namespace).
    Register this URL as a scheduled feed in Merchant Center; Google refetches
    it on its own, so it always reflects current products, prices, and stock."""
    base = 'https://adhyashaktishop.com'
    db = get_db()
    cat_names = {}
    try:
        for c in db.execute("SELECT id,name FROM categories").fetchall():
            cat_names[c['id']] = c['name'] or ''
    except Exception:
        pass
    rows = db.execute(
        """SELECT id,name,description,price,compare_price,sku,stock,images,category_id
           FROM products WHERE IFNULL(is_active,1)=1 ORDER BY created_at DESC LIMIT 1000"""
    ).fetchall()

    def x(v):
        return escape(str(v if v is not None else ''), quote=True)

    def strip_html(s):
        s = re.sub(r'<[^>]+>', ' ', str(s or ''))
        return re.sub(r'\s+', ' ', s).strip()

    def abs_img(im):
        im = str(im)
        return im if im.startswith('http') else base + im

    items = []
    for r in rows:
        price = float(r['price'] or 0)
        if price <= 0:
            continue  # Google rejects zero-price items
        try:
            images = [str(im) for im in json.loads(r['images'] or '[]') if im]
        except Exception:
            images = []
        if not images:
            continue  # Google requires at least one image
        name = (r['name'] or 'Product')[:150]
        desc = (strip_html(r['description']) or name)[:4900]
        compare = float(r['compare_price'] or 0)
        avail = 'in_stock' if int(r['stock'] or 0) > 0 else 'out_of_stock'
        cat_name = cat_names.get(r['category_id'], '')
        low = ('%s %s' % (cat_name, name)).lower()
        gcat = ('Apparel & Accessories > Clothing'
                if any(k in low for k in ('shirt', 'hoodie', 'cloth', 'apparel', 'co-ord', 'polo', 'tee'))
                else 'Apparel & Accessories > Jewelry')
        p = ['    <item>',
             '      <g:id>%s</g:id>' % x(r['id']),
             '      <g:title>%s</g:title>' % x(name),
             '      <g:description>%s</g:description>' % x(desc),
             '      <g:link>%s/product/%s</g:link>' % (base, x(r['id'])),
             '      <g:image_link>%s</g:image_link>' % x(abs_img(images[0]))]
        for im in images[1:11]:
            p.append('      <g:additional_image_link>%s</g:additional_image_link>' % x(abs_img(im)))
        p.append('      <g:availability>%s</g:availability>' % avail)
        p.append('      <g:condition>new</g:condition>')
        if compare > price:
            p.append('      <g:price>%.2f USD</g:price>' % compare)
            p.append('      <g:sale_price>%.2f USD</g:sale_price>' % price)
        else:
            p.append('      <g:price>%.2f USD</g:price>' % price)
        p.append('      <g:brand>Adhya Shakti Shop</g:brand>')
        sku = (r['sku'] or '').strip()
        if sku:
            p.append('      <g:mpn>%s</g:mpn>' % x(sku))
        else:
            p.append('      <g:identifier_exists>no</g:identifier_exists>')
        p.append('      <g:google_product_category>%s</g:google_product_category>' % x(gcat))
        if cat_name:
            p.append('      <g:product_type>%s</g:product_type>' % x(cat_name))
        p.append('    </item>')
        items.append('\n'.join(p))

    xml = ('<?xml version="1.0" encoding="UTF-8"?>\n'
           '<rss version="2.0" xmlns:g="http://base.google.com/ns/1.0">\n'
           '  <channel>\n'
           '    <title>Adhya Shakti Shop</title>\n'
           '    <link>https://adhyashaktishop.com</link>\n'
           '    <description>Handcrafted jewelry and custom-printed clothing from New Jersey, USA.</description>\n'
           '%s\n'
           '  </channel>\n'
           '</rss>') % '\n'.join(items)
    return Response(xml, mimetype='application/xml')


# ─── Sliders & Settings ───────────────────────────────────────────────────────

@app.route('/api/sliders', methods=['GET'])
def get_sliders():
    db = get_db()
    return jsonify(rows_to_list(db.execute("SELECT * FROM sliders WHERE is_active=1 ORDER BY sort_order").fetchall()))


@app.route('/api/admin/sliders', methods=['POST'])
@admin_only_required
def create_slider():
    raw = request.get_json(silent=True)
    data = {} if raw is None else raw
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid slider data'}), 400
    title = clean_text(data.get('title'), 160)
    if not title:
        return jsonify({'error': 'Slider title is required'}), 400
    try:
        sort_order = int(data.get('sort_order') or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'Sort order must be a whole number'}), 400
    db = get_db()
    sid = str(uuid.uuid4())
    db.execute("INSERT INTO sliders (id,title,subtitle,image_url,link,sort_order) VALUES (?,?,?,?,?,?)",
               (sid, title, clean_text(data.get('subtitle'), 260), clean_text(data.get('image_url'), 260),
                clean_text(data.get('link'), 260), sort_order))
    db.commit()
    log_admin_action('slider_created', 'Homepage slider created', {'slider_id': sid, 'title': title})
    return jsonify({'id': sid, 'message': 'Slider created'})


@app.route('/api/settings', methods=['GET'])
def get_settings():
    db = get_db()
    public_keys = ('shop_name', 'shop_email', 'shop_phone', 'shop_address')
    rows = db.execute(
        "SELECT key,value FROM settings WHERE key IN (?,?,?,?)",
        public_keys
    ).fetchall()
    return jsonify({r['key']: r['value'] for r in rows})


@app.route('/api/admin/settings', methods=['PUT'])
@admin_only_required
def update_settings():
    raw = request.get_json(silent=True)
    data = {} if raw is None else raw
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid settings data'}), 400
    forbidden = [k for k in data.keys() if clean_text(k, 120).lower().startswith('razorpay')]
    if forbidden:
        return jsonify({'error': 'Razorpay settings are disabled for this Stripe shop.'}), 400
    db = get_db()
    keys = [clean_text(k, 120) for k in data.keys()]
    before = {
        r['key']: r['value']
        for r in db.execute(
            "SELECT key,value FROM settings WHERE key IN (%s)" % ",".join("?" for _ in keys),  # nosec B608
            keys
        ).fetchall()
    } if keys else {}
    for k, v in data.items():
        db.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)", (k, str(v)))
    db.commit()
    after = {k: str(v) for k, v in data.items()}
    log_admin_action('settings_updated', 'Store settings updated', {
        'entity_type': 'settings',
        'entity_id': 'store',
        'keys': list(data.keys())[:50],
        'before': before,
        'after': after,
    })
    return jsonify({'message': 'Settings updated'})


SECURITY_LOW_RISK_REVIEW_EVENTS = (
    'login_success',
    'password_reset_requested',
    'password_reset_completed',
    'data_exported',
    'backup_created',
    'backup_verified',
    'backup_downloaded',
    'backup_retention_cleaned',
    'backup_restore_drill',
    'security_events_reviewed',
    'security_events_unreviewed',
)

SECURITY_NEEDS_ATTENTION_EVENTS = (
    'login_failed',
    'login_locked',
    'password_reset_unknown_email',
    'password_reset_failed',
    'csrf_failed',
    'auth_token_rejected',
    'authorization_denied',
    'public_endpoint_rate_limited',
    'upload_rejected',
    'backup_restore_drill_failed',
    'bulk_order_rejected',
    'contact_rejected',
    'public_honeypot_triggered',
)

SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS = (
    'auth_required_missing',
    'authorization_denied',
    'csrf_failed',
)


def normalize_security_ip(value):
    raw = clean_text(value, 80)
    if not raw:
        return ''
    try:
        return str(ipaddress.ip_address(raw))
    except ValueError:
        return ''


def _security_placeholders(values):
    return ','.join('?' for _ in values)


def _security_bucket(event_type, severity, trusted_label='', reviewed_at=''):
    if reviewed_at:
        return 'reviewed'
    if severity == 'critical':
        return 'needs_attention'
    if event_type in SECURITY_LOW_RISK_REVIEW_EVENTS:
        return 'routine'
    if trusted_label and event_type in SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS:
        return 'routine'
    if event_type in SECURITY_NEEDS_ATTENTION_EVENTS:
        return 'needs_attention'
    if severity == 'warning':
        return 'needs_attention'
    return 'routine'


def _security_explanation(event_type):
    explanations = {
        'login_success': 'A valid account logged in. Usually normal.',
        'login_failed': 'Someone tried to log in but the credentials did not match.',
        'login_locked': 'Login protection temporarily locked an account/IP after repeated failures.',
        'password_reset_requested': 'A password reset email was requested for a known account.',
        'password_reset_unknown_email': 'Someone requested a reset for an email that is not registered.',
        'password_reset_failed': 'A reset link was invalid, expired, malformed, or reused.',
        'password_reset_completed': 'A password was reset successfully.',
        'csrf_failed': 'A form/API request failed the browser security token check.',
        'auth_required_missing': 'A protected API was requested without an active login session.',
        'auth_token_rejected': 'An expired, revoked, or invalid session token was used.',
        'authorization_denied': 'A logged-in user tried to access an area their role cannot access.',
        'public_endpoint_rate_limited': 'A public form or reset endpoint hit the anti-abuse limit.',
        'upload_rejected': 'An uploaded file was blocked by file safety checks.',
        'bulk_order_rejected': 'A bulk inquiry looked unsafe or spam-like.',
        'contact_rejected': 'A contact form message looked unsafe or spam-like.',
        'public_honeypot_triggered': 'A hidden anti-spam field was filled.',
        'data_exported': 'An admin downloaded a CSV export.',
        'backup_created': 'An admin created a site backup.',
        'backup_verified': 'An admin verified a backup archive.',
        'backup_downloaded': 'An admin downloaded a backup archive.',
        'backup_retention_cleaned': 'Old backup archives were removed by the 7-day retention policy.',
        'backup_restore_drill': 'An admin safely tested a backup restore in a temporary folder.',
        'backup_restore_drill_failed': 'A backup restore drill found a problem with the backup archive.',
        'security_events_reviewed': 'An admin marked security events reviewed.',
        'security_events_unreviewed': 'An admin reopened reviewed security events.',
    }
    return explanations.get(event_type, 'Security activity recorded by the system.')


@app.route('/api/admin/trusted-ips', methods=['GET'])
@admin_only_required
def admin_trusted_ips():
    db = get_db()
    rows = rows_to_list(db.execute("""
        SELECT id, ip, label, note, is_active, created_by, created_at, updated_at
        FROM admin_trusted_ips
        WHERE is_active=1
        ORDER BY updated_at DESC, created_at DESC
    """).fetchall())
    return jsonify({'trusted_ips': rows})


@app.route('/api/admin/trusted-ips', methods=['POST'])
@admin_only_required
def admin_trusted_ips_save():
    data = request.json or {}
    ip = normalize_security_ip(data.get('ip'))
    label = clean_text(data.get('label'), 80)
    note = clean_text(data.get('note'), 240)
    if not ip:
        return jsonify({'error': 'Enter a valid IP address'}), 400
    if not label:
        label = 'Trusted admin IP'
    db = get_db()
    db.execute("""
        INSERT INTO admin_trusted_ips (id,ip,label,note,is_active,created_by)
        VALUES (?,?,?,?,1,?)
        ON CONFLICT(ip) DO UPDATE SET
          label=excluded.label,
          note=excluded.note,
          is_active=1,
          updated_at=datetime('now')
    """, (str(uuid.uuid4()), ip, label, note or None, g.user.get('id')))
    db.commit()
    log_audit_event('trusted_ip_saved', 'Trusted admin IP saved', metadata={
        'entity_type': 'trusted_ip',
        'entity_id': ip,
        'ip': ip,
        'label': label,
    })
    return jsonify({'message': 'Trusted IP saved', 'ip': ip, 'label': label})


@app.route('/api/admin/trusted-ips/<path:ip>', methods=['DELETE'])
@admin_only_required
def admin_trusted_ips_delete(ip):
    normalized = normalize_security_ip(ip)
    if not normalized:
        return jsonify({'error': 'Enter a valid IP address'}), 400
    db = get_db()
    cur = db.execute(
        "UPDATE admin_trusted_ips SET is_active=0, updated_at=datetime('now') WHERE ip=?",
        (normalized,)
    )
    db.commit()
    log_audit_event('trusted_ip_removed', 'Trusted admin IP removed', metadata={
        'entity_type': 'trusted_ip',
        'entity_id': normalized,
        'ip': normalized,
    })
    return jsonify({'message': 'Trusted IP removed', 'updated': cur.rowcount or 0})


@app.route('/api/admin/security-events', methods=['GET'])
@admin_only_required
def admin_security_events():
    db = get_db()
    limit = request.args.get('limit', 150, type=int) or 150
    limit = max(25, min(limit, 500))
    event_type = clean_text(request.args.get('event_type'), 80)
    severity = clean_text(request.args.get('severity'), 20)
    email = clean_text(request.args.get('email'), 140).lower()
    ip = clean_text(request.args.get('ip'), 80)
    query = clean_text(request.args.get('q'), 160)
    reviewed = clean_text(request.args.get('reviewed'), 20).lower()
    scope = clean_text(request.args.get('scope'), 40).lower()
    days = request.args.get('days', type=int)
    days = max(1, min(days, 90)) if days else None

    where = []
    params = []

    def add_event_type_filter(types):
        placeholders = ','.join('?' for _ in types)
        where.append(f"event_type IN ({placeholders})")  # nosec B608
        params.extend(types)

    if event_type:
        where.append("event_type=?")
        params.append(event_type)
    if severity:
        where.append("severity=?")
        params.append(severity)
    if email:
        where.append("email LIKE ?")
        params.append(f"%{email}%")
    if ip:
        where.append("ip=?")
        params.append(ip)
    if days:
        where.append("created_at >= datetime('now', ?)")
        params.append(f'-{days} days')
    if reviewed == 'reviewed':
        where.append("reviewed_at IS NOT NULL")
    elif reviewed == 'unreviewed':
        where.append("reviewed_at IS NULL")

    if scope == 'needs_attention':
        needs_ph = _security_placeholders(SECURITY_NEEDS_ATTENTION_EVENTS)
        trusted_noise_ph = _security_placeholders(SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS)
        low_ph = _security_placeholders(SECURITY_LOW_RISK_REVIEW_EVENTS)
        where.append(f"""(
            severity='critical'
            OR event_type IN ({needs_ph})
            OR (
                severity='warning'
                AND event_type NOT IN ({low_ph})
                AND NOT (
                    event_type IN ({trusted_noise_ph})
                    AND ip IN (SELECT ip FROM admin_trusted_ips WHERE is_active=1)
                )
            )
        )""")  # nosec B608
        params.extend([*SECURITY_NEEDS_ATTENTION_EVENTS, *SECURITY_LOW_RISK_REVIEW_EVENTS, *SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS])
    elif scope == 'routine':
        low_ph = _security_placeholders(SECURITY_LOW_RISK_REVIEW_EVENTS)
        trusted_noise_ph = _security_placeholders(SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS)
        where.append(f"""(
            severity='info'
            OR event_type IN ({low_ph})
            OR (
                event_type IN ({trusted_noise_ph})
                AND ip IN (SELECT ip FROM admin_trusted_ips WHERE is_active=1)
            )
        )""")  # nosec B608
        params.extend([*SECURITY_LOW_RISK_REVIEW_EVENTS, *SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS])
    elif scope == 'trusted':
        where.append("ip IN (SELECT ip FROM admin_trusted_ips WHERE is_active=1)")
    elif scope == 'untrusted_risk':
        where.append("severity IN ('warning','critical')")
        where.append("""(
            ip IS NULL
            OR ip NOT IN (SELECT ip FROM admin_trusted_ips WHERE is_active=1)
        )""")
    elif scope == 'risky':
        where.append("severity IN ('warning','critical')")
        where.append("created_at >= datetime('now','-24 hours')")
    elif scope == 'critical':
        where.append("severity='critical'")
    elif scope == 'failed_logins':
        add_event_type_filter(('login_failed', 'login_locked'))
    elif scope == 'password_reset':
        add_event_type_filter((
            'password_reset_requested',
            'password_reset_unknown_email',
            'password_reset_failed',
            'password_reset_completed',
        ))
    elif scope == 'suspicious':
        where.append("""(
            severity IN ('warning','critical')
            OR event_type IN (
                'login_failed','login_locked','csrf_failed','auth_token_rejected',
                'auth_required_missing','public_endpoint_rate_limited',
                'upload_rejected','bulk_order_rejected','contact_rejected',
                'public_honeypot_triggered'
            )
        )""")
    elif scope == 'uploads':
        add_event_type_filter(('upload_rejected',))

    if query:
        like = f"%{query}%"
        where.append("""(
            event_type LIKE ?
            OR severity LIKE ?
            OR COALESCE(email,'') LIKE ?
            OR COALESCE(ip,'') LIKE ?
            OR COALESCE(path,'') LIKE ?
            OR COALESCE(message,'') LIKE ?
            OR COALESCE(user_id,'') LIKE ?
        )""")
        params.extend([like, like, like, like, like, like, like])

    sql = "SELECT * FROM security_events"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY created_at DESC LIMIT ?"
    params.append(limit)

    events = rows_to_list(db.execute(sql, params).fetchall())
    trusted_rows = rows_to_list(db.execute("""
        SELECT ip, label
        FROM admin_trusted_ips
        WHERE is_active=1
    """).fetchall())
    trusted_by_ip = {r['ip']: r['label'] for r in trusted_rows}
    for event in events:
        try:
            event['metadata'] = json.loads(event.get('metadata') or '{}')
        except Exception:
            event['metadata'] = {}
        trusted_label = trusted_by_ip.get(event.get('ip') or '') or ''
        event['trusted_ip_label'] = trusted_label
        event['bucket'] = _security_bucket(
            event.get('event_type') or '',
            event.get('severity') or 'info',
            trusted_label,
            event.get('reviewed_at') or '',
        )
        event['explanation'] = _security_explanation(event.get('event_type') or '')

    summary = rows_to_list(db.execute("""
        SELECT event_type, severity, COUNT(*) AS count
        FROM security_events
        WHERE created_at >= datetime('now','-7 days')
        GROUP BY event_type, severity
        ORDER BY count DESC
        LIMIT 20
    """).fetchall())

    risky = db.execute("""
        SELECT COUNT(*) AS count
        FROM security_events
        WHERE severity IN ('warning','critical')
          AND created_at >= datetime('now','-24 hours')
    """).fetchone()['count']

    needs_ph = _security_placeholders(SECURITY_NEEDS_ATTENTION_EVENTS)
    low_ph = _security_placeholders(SECURITY_LOW_RISK_REVIEW_EVENTS)
    trusted_noise_ph = _security_placeholders(SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS)
    attention_sql = f"""(
        se.severity='critical'
        OR se.event_type IN ({needs_ph})
        OR (
            se.severity='warning'
            AND se.event_type NOT IN ({low_ph})
            AND NOT (tip.ip IS NOT NULL AND se.event_type IN ({trusted_noise_ph}))
        )
    )"""
    routine_sql = f"""(
        se.severity='info'
        OR se.event_type IN ({low_ph})
        OR (tip.ip IS NOT NULL AND se.event_type IN ({trusted_noise_ph}))
    )"""
    stats_params = [
        *SECURITY_NEEDS_ATTENTION_EVENTS,
        *SECURITY_LOW_RISK_REVIEW_EVENTS,
        *SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS,
        *SECURITY_LOW_RISK_REVIEW_EVENTS,
        *SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS,
    ]
    stats = db.execute(f"""
        SELECT
          SUM(CASE WHEN se.severity IN ('warning','critical') AND se.reviewed_at IS NULL THEN 1 ELSE 0 END) AS unreviewed_risky,
          SUM(CASE WHEN se.severity='critical' AND se.reviewed_at IS NULL THEN 1 ELSE 0 END) AS unreviewed_critical,
          SUM(CASE WHEN se.event_type IN ('login_failed','login_locked') AND se.created_at >= datetime('now','-24 hours') THEN 1 ELSE 0 END) AS failed_logins_24h,
          SUM(CASE WHEN se.event_type LIKE 'password_reset%' AND se.created_at >= datetime('now','-24 hours') THEN 1 ELSE 0 END) AS password_resets_24h,
          SUM(CASE WHEN se.reviewed_at IS NULL THEN 1 ELSE 0 END) AS unreviewed_total,
          SUM(CASE WHEN se.reviewed_at IS NULL AND {attention_sql} THEN 1 ELSE 0 END) AS unreviewed_needs_attention,
          SUM(CASE WHEN se.reviewed_at IS NULL AND {routine_sql} THEN 1 ELSE 0 END) AS unreviewed_routine,
          SUM(CASE WHEN tip.ip IS NOT NULL THEN 1 ELSE 0 END) AS trusted_ip_events,
          SUM(CASE WHEN se.severity IN ('warning','critical') AND tip.ip IS NULL THEN 1 ELSE 0 END) AS untrusted_risk_events
        FROM security_events se
        LEFT JOIN admin_trusted_ips tip ON tip.ip=se.ip AND tip.is_active=1
    """, stats_params).fetchone()  # nosec B608

    top_ips = rows_to_list(db.execute("""
        SELECT
          se.ip,
          tip.label AS trusted_label,
          COUNT(*) AS total,
          SUM(CASE WHEN se.severity IN ('warning','critical') THEN 1 ELSE 0 END) AS risky,
          MAX(se.created_at) AS last_seen
        FROM security_events se
        LEFT JOIN admin_trusted_ips tip ON tip.ip=se.ip AND tip.is_active=1
        WHERE COALESCE(se.ip,'') != ''
          AND se.created_at >= datetime('now','-7 days')
        GROUP BY se.ip, tip.label
        HAVING risky > 0 OR total >= 5
        ORDER BY CASE WHEN tip.label IS NULL THEN 0 ELSE 1 END ASC, risky DESC, total DESC, last_seen DESC
        LIMIT 10
    """).fetchall())

    event_types = rows_to_list(db.execute("""
        SELECT event_type, COUNT(*) AS count
        FROM security_events
        GROUP BY event_type
        ORDER BY count DESC, event_type ASC
        LIMIT 60
    """).fetchall())

    return jsonify({
        'events': events,
        'summary': summary,
        'risky_24h': risky,
        'stats': dict(stats) if stats else {},
        'top_ips': top_ips,
        'event_types': event_types,
        'trusted_ips': trusted_rows,
        'filters': {
            'event_type': event_type,
            'severity': severity,
            'email': email,
            'ip': ip,
            'q': query,
            'reviewed': reviewed,
            'scope': scope,
            'days': days,
            'limit': limit,
        },
    })


@app.route('/api/admin/security-events/reviewed', methods=['POST'])
@admin_only_required
def admin_security_events_reviewed():
    data = request.json or {}
    raw_ids = data.get('ids') or []
    if not isinstance(raw_ids, list):
        return jsonify({'error': 'Event IDs are required'}), 400
    ids = [clean_text(v, 80) for v in raw_ids if clean_text(v, 80)]
    ids = list(dict.fromkeys(ids))[:300]
    if not ids:
        return jsonify({'error': 'Select at least one event'}), 400

    reviewed = bool(data.get('reviewed', True))
    note = clean_text(data.get('note'), 240)
    placeholders = ','.join('?' for _ in ids)
    db = get_db()
    if reviewed:
        params = [g.user.get('id'), note or None, *ids]
        cur = db.execute(
            f"""UPDATE security_events
                SET reviewed_at=datetime('now'), reviewed_by=?, review_note=?
                WHERE id IN ({placeholders})""",  # nosec B608
            params,
        )
        action = 'security_events_reviewed'
        message = 'Security events marked reviewed'
    else:
        cur = db.execute(
            f"""UPDATE security_events
                SET reviewed_at=NULL, reviewed_by=NULL, review_note=NULL
                WHERE id IN ({placeholders})""",  # nosec B608
            ids,
        )
        action = 'security_events_unreviewed'
        message = 'Security events marked unreviewed'
    db.commit()
    updated = cur.rowcount if cur.rowcount is not None else len(ids)
    log_audit_event(action, message=message, metadata={
        'entity_type': 'security_events',
        'entity_id': 'bulk',
        'count': updated,
        'requested_count': len(ids),
        'note': note,
        'sample_ids': ids[:10],
    })
    return jsonify({'message': message, 'updated': updated})


@app.route('/api/admin/security-events/review-low-risk', methods=['POST'])
@admin_only_required
def admin_security_events_review_low_risk():
    data = request.json or {}
    days = data.get('days', 30)
    try:
        days = int(days)
    except (TypeError, ValueError):
        days = 30
    days = max(1, min(days, 90))
    include_trusted_noise = bool(data.get('include_trusted_noise', True))

    low_ph = _security_placeholders(SECURITY_LOW_RISK_REVIEW_EVENTS)
    where = [
        "reviewed_at IS NULL",
        "created_at >= datetime('now', ?)",
        f"event_type IN ({low_ph})",
        "severity != 'critical'",
    ]
    params = [f'-{days} days', *SECURITY_LOW_RISK_REVIEW_EVENTS]

    if include_trusted_noise:
        trusted_noise_ph = _security_placeholders(SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS)
        where = [
            "reviewed_at IS NULL",
            "created_at >= datetime('now', ?)",
            f"""(
                (event_type IN ({low_ph}) AND severity != 'critical')
                OR (
                    event_type IN ({trusted_noise_ph})
                    AND ip IN (SELECT ip FROM admin_trusted_ips WHERE is_active=1)
                    AND severity='warning'
                )
            )""",
        ]
        params = [f'-{days} days', *SECURITY_LOW_RISK_REVIEW_EVENTS, *SECURITY_TRUSTED_ROUTINE_WARNING_EVENTS]

    db = get_db()
    cur = db.execute(
        f"""UPDATE security_events
            SET reviewed_at=datetime('now'), reviewed_by=?, review_note=?
            WHERE {' AND '.join(where)}""",  # nosec B608
        [g.user.get('id'), 'Bulk reviewed low-risk/routine events', *params],
    )
    db.commit()
    updated = cur.rowcount if cur.rowcount is not None else 0
    log_audit_event('security_low_risk_reviewed', 'Low-risk security events marked reviewed', metadata={
        'entity_type': 'security_events',
        'entity_id': 'low_risk',
        'count': updated,
        'days': days,
        'include_trusted_noise': include_trusted_noise,
    })
    return jsonify({'message': 'Low-risk security events marked reviewed', 'updated': updated})


@app.route('/api/admin/audit-log', methods=['GET'])
@admin_only_required
def admin_audit_log():
    db = get_db()
    limit = request.args.get('limit', 100, type=int) or 100
    limit = max(25, min(limit, 300))
    entity_type = clean_text(request.args.get('entity_type'), 80)
    action = clean_text(request.args.get('action'), 100)
    where = []
    params = []
    if entity_type:
        where.append("entity_type=?")
        params.append(entity_type)
    if action:
        where.append("action=?")
        params.append(action)
    sql = "SELECT * FROM admin_audit_log"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY created_at DESC LIMIT ?"
    params.append(limit)
    rows = rows_to_list(db.execute(sql, params).fetchall())
    for row in rows:
        for key in ('before_json', 'after_json', 'metadata'):
            try:
                row[key] = json.loads(row.get(key) or '{}') if row.get(key) else None
            except Exception:
                row[key] = None
    return jsonify({'entries': rows})


def _integrity_sample(db, sql, params=(), limit=5):
    try:
        rows = db.execute(sql + " LIMIT ?", (*params, limit)).fetchall()
        return [dict(r) for r in rows]
    except Exception as exc:
        return [{'error': str(exc)[:180]}]


def _integrity_scalar(db, sql, params=()):
    try:
        row = db.execute(sql, params).fetchone()
        return int(row[0] or 0) if row else 0
    except Exception:
        return None


def _integrity_entry(key, label, count, *, severity='warning', ok_text='OK', recommendation='', sample=None):
    status = 'ok' if not count else severity
    return {
        'key': key,
        'label': label,
        'status': status,
        'count': 0 if count is None else count,
        'message': ok_text if not count else recommendation,
        'recommendation': recommendation,
        'sample': sample or [],
    }


def build_integrity_report(db):
    checks = []
    try:
        integrity = db.execute('PRAGMA integrity_check').fetchone()[0]
    except Exception as exc:
        integrity = f'error: {exc}'
    checks.append({
        'key': 'sqlite_integrity',
        'label': 'SQLite integrity check',
        'status': 'ok' if integrity == 'ok' else 'critical',
        'count': 0 if integrity == 'ok' else 1,
        'message': integrity,
        'recommendation': '' if integrity == 'ok' else 'Create a backup and investigate database integrity before further changes.',
        'sample': [],
    })

    try:
        fk_rows = [dict(r) for r in db.execute('PRAGMA foreign_key_check').fetchall()]
    except Exception as exc:
        fk_rows = [{'error': str(exc)[:180]}]
    checks.append({
        'key': 'foreign_keys',
        'label': 'Foreign key consistency',
        'status': 'ok' if not fk_rows else 'critical',
        'count': len(fk_rows),
        'message': 'OK' if not fk_rows else 'Foreign key issues found',
        'recommendation': '' if not fk_rows else 'Review the listed orphaned references before changing related records.',
        'sample': fk_rows[:5],
    })

    count = _integrity_scalar(db, "SELECT COUNT(*) FROM products WHERE stock < 0")
    checks.append(_integrity_entry(
        'negative_product_stock', 'Products with negative stock', count, severity='critical',
        recommendation='Fix stock quantities before accepting more sales for these products.',
        sample=_integrity_sample(db, "SELECT id,name,stock FROM products WHERE stock < 0 ORDER BY stock ASC")
    ))

    count = _integrity_scalar(db, "SELECT COUNT(*) FROM product_variants WHERE stock < 0")
    checks.append(_integrity_entry(
        'negative_variant_stock', 'Variants with negative stock', count, severity='critical',
        recommendation='Fix color/size stock quantities before accepting more sales.',
        sample=_integrity_sample(db, "SELECT product_id,color,size,stock FROM product_variants WHERE stock < 0 ORDER BY stock ASC")
    ))

    mismatch_sql = """
        SELECT p.id,p.name,p.stock,IFNULL(SUM(v.stock),0) AS variant_stock
        FROM products p JOIN product_variants v ON v.product_id=p.id
        GROUP BY p.id
        HAVING p.stock != IFNULL(SUM(v.stock),0)
    """
    count = _integrity_scalar(db, f"SELECT COUNT(*) FROM ({mismatch_sql})")
    checks.append(_integrity_entry(
        'variant_stock_mismatch', 'Product stock differs from variant stock total', count, severity='warning',
        recommendation='Open these products and save the variant matrix so total stock matches available sizes/colors.',
        sample=_integrity_sample(db, mismatch_sql + " ORDER BY p.name")
    ))

    count = _integrity_scalar(db, """
        SELECT COUNT(*) FROM products p
        WHERE p.is_active=1 AND p.category_id IS NOT NULL
          AND NOT EXISTS (SELECT 1 FROM categories c WHERE c.id=p.category_id AND IFNULL(c.is_active,1)=1)
    """)
    checks.append(_integrity_entry(
        'invalid_product_categories', 'Active products in missing/inactive categories', count, severity='warning',
        recommendation='Move these products into a visible category or reactivate the category.',
        sample=_integrity_sample(db, """
            SELECT p.id,p.name,p.category_id FROM products p
            WHERE p.is_active=1 AND p.category_id IS NOT NULL
              AND NOT EXISTS (SELECT 1 FROM categories c WHERE c.id=p.category_id AND IFNULL(c.is_active,1)=1)
            ORDER BY p.name
        """)
    ))

    count = _integrity_scalar(db, "SELECT COUNT(*) FROM products WHERE is_active=1 AND (images IS NULL OR images='' OR images='[]')")
    checks.append(_integrity_entry(
        'products_without_images', 'Active products without images', count, severity='warning',
        recommendation='Add at least one clear product image so the website looks complete.',
        sample=_integrity_sample(db, "SELECT id,name,sku FROM products WHERE is_active=1 AND (images IS NULL OR images='' OR images='[]') ORDER BY name")
    ))

    count = _integrity_scalar(db, "SELECT COUNT(*) FROM products WHERE is_active=1 AND IFNULL(stock,0) <= 0")
    checks.append(_integrity_entry(
        'products_zero_stock', 'Active products with zero stock', count, severity='warning',
        recommendation='Restock these products or mark them inactive if they should not appear for sale.',
        sample=_integrity_sample(db, "SELECT id,name,sku,stock FROM products WHERE is_active=1 AND IFNULL(stock,0) <= 0 ORDER BY name")
    ))

    count = _integrity_scalar(db, "SELECT COUNT(*) FROM products WHERE is_active=1 AND IFNULL(cost_price,0) <= 0")
    checks.append(_integrity_entry(
        'products_without_cost', 'Active products without cost price', count, severity='warning',
        recommendation='Add cost price for accurate profit, inventory value, and bookkeeping reports.',
        sample=_integrity_sample(db, "SELECT id,name,sku,cost_price FROM products WHERE is_active=1 AND IFNULL(cost_price,0) <= 0 ORDER BY name")
    ))

    duplicate_categories_sql = """
        SELECT IFNULL(parent_id,'ROOT') AS parent_id, lower(name) AS normalized_name, COUNT(*) AS duplicates
        FROM categories WHERE IFNULL(kind,'') IN ('catalog','clothing')
        GROUP BY IFNULL(parent_id,'ROOT'), lower(name)
        HAVING COUNT(*) > 1
    """
    count = _integrity_scalar(db, f"SELECT COUNT(*) FROM ({duplicate_categories_sql})")
    checks.append(_integrity_entry(
        'duplicate_category_names', 'Duplicate category names under the same parent', count, severity='warning',
        recommendation='Merge, rename, or deactivate duplicate categories so product placement stays clear.',
        sample=_integrity_sample(db, duplicate_categories_sql + " ORDER BY normalized_name")
    ))

    orphan_wishlist_sql = """
        SELECT w.user_id,w.product_id FROM user_wishlist w
        WHERE NOT EXISTS (SELECT 1 FROM users u WHERE u.id=w.user_id)
           OR NOT EXISTS (SELECT 1 FROM products p WHERE p.id=w.product_id)
    """
    count = _integrity_scalar(db, f"SELECT COUNT(*) FROM ({orphan_wishlist_sql})")
    checks.append(_integrity_entry(
        'orphan_wishlist_rows', 'Wishlist rows without valid user/product', count, severity='warning',
        recommendation='Startup cleanup removes these; if they persist, review delete flows.',
        sample=_integrity_sample(db, orphan_wishlist_sql)
    ))

    orphan_reviews_sql = """
        SELECT r.id,r.user_id,r.product_id FROM reviews r
        WHERE NOT EXISTS (SELECT 1 FROM users u WHERE u.id=r.user_id)
           OR NOT EXISTS (SELECT 1 FROM products p WHERE p.id=r.product_id)
    """
    count = _integrity_scalar(db, f"SELECT COUNT(*) FROM ({orphan_reviews_sql})")
    checks.append(_integrity_entry(
        'orphan_review_rows', 'Reviews without valid user/product', count, severity='warning',
        recommendation='Startup cleanup removes these; if they persist, review delete flows.',
        sample=_integrity_sample(db, orphan_reviews_sql)
    ))

    sale_mismatch_sql = """
        SELECT s.id,s.ref_no,s.total,IFNULL(x.line_total,0) AS line_total
        FROM acc_sales s
        LEFT JOIN (
            SELECT sale_id, SUM(qty*unit_price) AS line_total
            FROM acc_sale_items GROUP BY sale_id
        ) x ON x.sale_id=s.id
        WHERE ABS(IFNULL(s.subtotal,0) - IFNULL(x.line_total,0)) > 0.02
    """
    count = _integrity_scalar(db, f"SELECT COUNT(*) FROM ({sale_mismatch_sql})")
    checks.append(_integrity_entry(
        'bookkeeping_sale_line_mismatch', 'Bookkeeping sales with mismatched line totals', count, severity='warning',
        recommendation='Open and resave these sales so subtotal, cost, and profit recalculate.',
        sample=_integrity_sample(db, sale_mismatch_sql + " ORDER BY s.ref_no")
    ))

    purchase_mismatch_sql = """
        SELECT p.id,p.ref_no,p.total,IFNULL(x.line_total,0) AS line_total
        FROM acc_purchases p
        LEFT JOIN (
            SELECT purchase_id, SUM(qty*unit_cost) AS line_total
            FROM acc_purchase_items GROUP BY purchase_id
        ) x ON x.purchase_id=p.id
        WHERE ABS(IFNULL(p.total,0) - IFNULL(x.line_total,0)) > 0.02
    """
    count = _integrity_scalar(db, f"SELECT COUNT(*) FROM ({purchase_mismatch_sql})")
    checks.append(_integrity_entry(
        'bookkeeping_purchase_line_mismatch', 'Bookkeeping purchases with mismatched line totals', count, severity='warning',
        recommendation='Open and resave these purchases so totals and stock records line up.',
        sample=_integrity_sample(db, purchase_mismatch_sql + " ORDER BY p.ref_no")
    ))

    bad_order_json = []
    try:
        for row in db.execute("SELECT id,order_number,items,shipping_address,subtotal FROM orders ORDER BY created_at DESC").fetchall():
            try:
                items = json.loads(row['items'] or '[]')
                shipping = json.loads(row['shipping_address'] or '{}')
                if not isinstance(items, list) or not isinstance(shipping, dict):
                    raise ValueError('wrong JSON shape')
                line_total = sum(float(i.get('price') or 0) * int(i.get('qty') or 0) for i in items if isinstance(i, dict))
                if items and abs(line_total - float(row['subtotal'] or 0)) > 0.02:
                    bad_order_json.append({'id': row['id'], 'order_number': row['order_number'], 'issue': 'subtotal_mismatch'})
            except Exception as exc:
                bad_order_json.append({'id': row['id'], 'order_number': row['order_number'], 'issue': str(exc)[:80]})
    except Exception as exc:
        bad_order_json.append({'error': str(exc)[:180]})
    checks.append(_integrity_entry(
        'order_json_or_total_issues', 'Orders with bad JSON or subtotal mismatch', len(bad_order_json), severity='warning',
        recommendation='Review these orders before using them for reporting or customer service.',
        sample=bad_order_json[:5]
    ))

    private_bills_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'private_bills'))
    missing_files = []
    try:
        for row in db.execute("SELECT id,stored_name,original_name,parent_type,parent_id FROM acc_attachments").fetchall():
            stored = row['stored_name']
            path = os.path.abspath(os.path.join(private_bills_dir, stored or ''))
            if not stored or not is_safe_stored_filename(stored) or os.path.commonpath([private_bills_dir, path]) != private_bills_dir or not os.path.isfile(path):
                missing_files.append(dict(row))
    except Exception as exc:
        missing_files.append({'error': str(exc)[:180]})
    checks.append(_integrity_entry(
        'missing_bill_files', 'Bill attachment records missing files', len(missing_files), severity='warning',
        recommendation='Re-upload missing receipt files or remove broken attachment records.',
        sample=missing_files[:5]
    ))

    unsafe_uploads = []
    try:
        if os.path.isdir(UPLOAD_FOLDER):
            for name in os.listdir(UPLOAD_FOLDER):
                if not is_safe_public_upload_url(f'/uploads/{name}'):
                    unsafe_uploads.append({'filename': name})
    except Exception as exc:
        unsafe_uploads.append({'error': str(exc)[:180]})
    checks.append(_integrity_entry(
        'unsafe_upload_filenames', 'Unexpected files in public uploads folder', len(unsafe_uploads), severity='critical',
        recommendation='Remove unexpected public upload files after confirming they are not needed.',
        sample=unsafe_uploads[:5]
    ))

    used_uploads = set()
    unused_uploads = []
    try:
        for row in db.execute("SELECT images FROM products WHERE images IS NOT NULL AND images!=''").fetchall():
            try:
                for url in json.loads(row['images'] or '[]'):
                    text = clean_text(url, 260)
                    if text.startswith('/uploads/'):
                        used_uploads.add(os.path.basename(text))
            except Exception:
                continue
        if os.path.isdir(UPLOAD_FOLDER):
            for name in os.listdir(UPLOAD_FOLDER):
                if is_safe_stored_filename(name) and name not in used_uploads:
                    unused_uploads.append({'filename': name})
    except Exception as exc:
        unused_uploads.append({'error': str(exc)[:180]})
    checks.append(_integrity_entry(
        'unused_upload_files', 'Safe upload files not used by products', len(unused_uploads), severity='warning',
        recommendation='Review before deleting. These may be old product photos or abandoned uploads.',
        sample=unused_uploads[:5]
    ))

    critical = sum(1 for c in checks if c['status'] == 'critical')
    warnings = sum(1 for c in checks if c['status'] == 'warning')
    return {
        'generated_at': datetime.datetime.utcnow().isoformat(timespec='seconds') + 'Z',
        'critical_count': critical,
        'warning_count': warnings,
        'ok_count': sum(1 for c in checks if c['status'] == 'ok'),
        'checks': checks,
    }


@app.route('/api/admin/integrity', methods=['GET'])
@admin_only_required
def admin_integrity():
    return jsonify(build_integrity_report(get_db()))


def _csv_safe(value):
    if value is None:
        return ''
    text = str(value).replace('\r', ' ').replace('\n', ' ').strip()
    if text[:1] in ('=', '+', '-', '@'):
        return "'" + text
    return text


ADMIN_EXPORTS = {
    'customers': {
        'filename': 'customers',
        'sql': "SELECT id,name,email,phone,role,created_at FROM users ORDER BY created_at DESC",
        'columns': ['id', 'name', 'email', 'phone', 'role', 'created_at'],
    },
    'orders': {
        'filename': 'orders',
        'sql': """SELECT id,order_number,customer_name,customer_email,customer_phone,status,payment_status,
                 subtotal,discount,shipping_charge,total,coupon_code,tracking_number,return_reason,created_at,updated_at,notes
                 FROM orders ORDER BY created_at DESC""",
        'columns': ['id', 'order_number', 'customer_name', 'customer_email', 'customer_phone', 'status', 'payment_status',
                    'subtotal', 'discount', 'shipping_charge', 'total', 'coupon_code', 'tracking_number', 'return_reason', 'created_at', 'updated_at', 'notes'],
    },
    'products': {
        'filename': 'products',
        'sql': """SELECT p.id,p.name,p.sku,p.price,p.compare_price,p.cost_price,p.stock,p.low_stock_threshold,
                 c.name AS category,p.is_active,p.allow_custom_print,p.is_bestseller,p.created_at
                 FROM products p LEFT JOIN categories c ON c.id=p.category_id ORDER BY p.name COLLATE NOCASE""",
        'columns': ['id', 'name', 'sku', 'price', 'compare_price', 'cost_price', 'stock', 'low_stock_threshold',
                    'category', 'is_active', 'allow_custom_print', 'is_bestseller', 'created_at'],
    },
    'back_in_stock': {
        'filename': 'back_in_stock_requests',
        'sql': """SELECT b.id,p.name AS product,p.sku,b.email,b.name,b.status,b.request_count,
                 b.created_at,b.last_requested_at,b.notified_at
                 FROM back_in_stock_requests b
                 LEFT JOIN products p ON p.id=b.product_id
                 ORDER BY COALESCE(b.notified_at,'') ASC, b.last_requested_at DESC""",
        'columns': ['id', 'product', 'sku', 'email', 'name', 'status', 'request_count', 'created_at', 'last_requested_at', 'notified_at'],
    },
    'inventory': {
        'filename': 'inventory',
        'sql': """SELECT id,name,sku,stock,cost_price,price,
                 ROUND(IFNULL(stock,0)*IFNULL(cost_price,0),2) AS cost_value,
                 ROUND(IFNULL(stock,0)*IFNULL(price,0),2) AS retail_value,
                 low_stock_threshold,is_active
                 FROM products ORDER BY name COLLATE NOCASE""",
        'columns': ['id', 'name', 'sku', 'stock', 'cost_price', 'price', 'cost_value', 'retail_value', 'low_stock_threshold', 'is_active'],
    },
    'categories': {
        'filename': 'categories',
        'sql': "SELECT id,name,parent_id,kind,is_active,sort_order FROM categories ORDER BY sort_order,name COLLATE NOCASE",
        'columns': ['id', 'name', 'parent_id', 'kind', 'is_active', 'sort_order'],
    },
    'reviews': {
        'filename': 'reviews',
        'sql': """SELECT r.id,p.name AS product,u.name AS customer,u.email,r.rating,r.comment,r.created_at
                 FROM reviews r LEFT JOIN products p ON p.id=r.product_id LEFT JOIN users u ON u.id=r.user_id
                 ORDER BY r.created_at DESC""",
        'columns': ['id', 'product', 'customer', 'email', 'rating', 'comment', 'created_at'],
    },
    'coupons': {
        'filename': 'coupons',
        'sql': "SELECT id,code,discount_type,discount_value,min_order,max_uses,used_count,expires_at,is_active,created_at FROM coupons ORDER BY created_at DESC",
        'columns': ['id', 'code', 'discount_type', 'discount_value', 'min_order', 'max_uses', 'used_count', 'expires_at', 'is_active', 'created_at'],
    },
    'vendors': {
        'filename': 'vendors',
        'sql': "SELECT id,name,contact_name,phone,email,address,notes,is_active,created_at FROM acc_vendors ORDER BY name COLLATE NOCASE",
        'columns': ['id', 'name', 'contact_name', 'phone', 'email', 'address', 'notes', 'is_active', 'created_at'],
    },
    'sales': {
        'filename': 'sales',
        'sql': "SELECT id,ref_no,sale_date,channel,customer_name,payment_method,subtotal,discount,total,cost_total,profit,created_at FROM acc_sales ORDER BY sale_date DESC, created_at DESC",
        'columns': ['id', 'ref_no', 'sale_date', 'channel', 'customer_name', 'payment_method', 'subtotal', 'discount', 'total', 'cost_total', 'profit', 'created_at'],
    },
    'purchases': {
        'filename': 'purchases',
        'sql': """SELECT p.id,p.ref_no,p.purchase_date,IFNULL(v.name,p.supplier) AS vendor,p.payment_method,p.total,p.notes,p.created_at
                 FROM acc_purchases p LEFT JOIN acc_vendors v ON v.id=p.vendor_id ORDER BY p.purchase_date DESC, p.created_at DESC""",
        'columns': ['id', 'ref_no', 'purchase_date', 'vendor', 'payment_method', 'total', 'notes', 'created_at'],
    },
    'expenses': {
        'filename': 'expenses',
        'sql': """SELECT e.id,e.expense_date,e.category,e.payee,IFNULL(v.name,'') AS vendor,e.payment_method,e.amount,e.notes,e.created_at
                 FROM acc_expenses e LEFT JOIN acc_vendors v ON v.id=e.vendor_id ORDER BY e.expense_date DESC, e.created_at DESC""",
        'columns': ['id', 'expense_date', 'category', 'payee', 'vendor', 'payment_method', 'amount', 'notes', 'created_at'],
    },
    'security_events': {
        'filename': 'security_events',
        'sql': "SELECT id,event_type,severity,email,ip,path,method,message,created_at FROM security_events ORDER BY created_at DESC LIMIT 5000",
        'columns': ['id', 'event_type', 'severity', 'email', 'ip', 'path', 'method', 'message', 'created_at'],
    },
    'audit_log': {
        'filename': 'audit_log',
        'sql': "SELECT id,actor_email,actor_role,action,entity_type,entity_id,message,ip,created_at FROM admin_audit_log ORDER BY created_at DESC LIMIT 5000",
        'columns': ['id', 'actor_email', 'actor_role', 'action', 'entity_type', 'entity_id', 'message', 'ip', 'created_at'],
    },
}


@app.route('/api/admin/export/<kind>', methods=['GET'])
@admin_only_required
def admin_export(kind):
    kind = clean_text(kind, 40).lower()
    cfg = ADMIN_EXPORTS.get(kind)
    if not cfg:
        return jsonify({'error': 'Export type not found'}), 404
    db = get_db()
    try:
        rows = db.execute(cfg['sql']).fetchall()
    except Exception as exc:
        log_security_event(
            'admin_export_failed',
            'warning',
            'Admin CSV export failed',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
            metadata={'kind': kind, 'error': str(exc)[:300]},
        )
        return jsonify({'error': 'Export could not be created'}), 500
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(cfg['columns'])
    for row in rows:
        writer.writerow([_csv_safe(row[col] if col in row.keys() else '') for col in cfg['columns']])
    filename = f"adhya_{cfg['filename']}_{datetime.date.today().isoformat()}.csv"
    log_admin_action('data_exported', f'Exported {kind} CSV', {
        'entity_type': 'export',
        'entity_id': kind,
        'kind': kind,
        'rows': len(rows),
    })
    resp = Response(buf.getvalue(), mimetype='text/csv; charset=utf-8')
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    resp.headers['Cache-Control'] = 'no-store, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp


@app.route('/api/admin/health', methods=['GET'])
@admin_only_required
def admin_health():
    db = get_db()
    db_size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    backups_dir = os.path.join(os.path.dirname(__file__), 'backups')
    latest_backup_name = ''
    latest_backup_at = ''
    latest_backup_size = 0
    if os.path.isdir(backups_dir):
        try:
            backups = [
                os.path.join(backups_dir, name)
                for name in os.listdir(backups_dir)
                if name.endswith('.zip') and os.path.isfile(os.path.join(backups_dir, name))
            ]
            if backups:
                latest = max(backups, key=os.path.getmtime)
                latest_backup_name = os.path.basename(latest)
                latest_backup_at = datetime.datetime.fromtimestamp(
                    os.path.getmtime(latest), datetime.timezone.utc
                ).isoformat(timespec='seconds').replace('+00:00', 'Z')
                latest_backup_size = os.path.getsize(latest)
        except Exception:
            latest_backup_name = latest_backup_at = ''
            latest_backup_size = 0
    backups_count = len(backups) if 'backups' in locals() else 0
    uploads_count = 0
    uploads_size = 0
    try:
        for root, _dirs, files in os.walk(UPLOAD_FOLDER):
            uploads_count += len(files)
            for name in files:
                fp = os.path.join(root, name)
                if os.path.isfile(fp):
                    uploads_size += os.path.getsize(fp)
    except Exception:
        uploads_count = uploads_size = 0
    private_bills_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'private_bills'))
    private_bills_count = 0
    private_bills_size = 0
    try:
        for root, _dirs, files in os.walk(private_bills_dir):
            private_bills_count += len(files)
            for name in files:
                fp = os.path.join(root, name)
                if os.path.isfile(fp):
                    private_bills_size += os.path.getsize(fp)
    except Exception:
        private_bills_count = private_bills_size = 0
    try:
        db_integrity = db.execute('PRAGMA integrity_check').fetchone()[0]
        fk_issue_count = len(db.execute('PRAGMA foreign_key_check').fetchall())
    except Exception:
        db_integrity = 'unknown'
        fk_issue_count = -1
    integrity_report = build_integrity_report(db)

    row = db.execute("""
        SELECT
          (SELECT COUNT(*) FROM products WHERE is_active=1) AS active_products,
          (SELECT COUNT(*) FROM products WHERE is_active=1 AND stock<=5) AS low_stock_products,
          (SELECT COUNT(*) FROM orders WHERE status='pending') AS pending_orders,
          (SELECT COUNT(*) FROM orders WHERE status='return_requested') AS return_requests,
          (SELECT COUNT(*) FROM security_events WHERE severity IN ('warning','critical') AND created_at >= datetime('now','-24 hours')) AS warning_24h,
          (SELECT COUNT(*) FROM security_events WHERE severity='critical' AND created_at >= datetime('now','-7 days')) AS critical_7d,
          (SELECT COUNT(*) FROM admin_audit_log WHERE created_at >= datetime('now','-7 days')) AS audit_7d,
          (SELECT MAX(created_at) FROM orders) AS last_order_at
    """).fetchone()
    return jsonify({
        'generated_at': datetime.datetime.utcnow().isoformat(timespec='seconds') + 'Z',
        'database_size_mb': round(db_size / (1024 * 1024), 2),
        'uploads_count': uploads_count,
        'uploads_size_mb': round(uploads_size / (1024 * 1024), 2),
        'private_bills_count': private_bills_count,
        'private_bills_size_mb': round(private_bills_size / (1024 * 1024), 2),
        'backups_count': backups_count,
        'latest_backup_name': latest_backup_name,
        'latest_backup_at': latest_backup_at,
        'latest_backup_size_mb': round(latest_backup_size / (1024 * 1024), 2),
        'database_integrity': db_integrity,
        'foreign_key_issue_count': fk_issue_count,
        'integrity_critical_count': integrity_report['critical_count'],
        'integrity_warning_count': integrity_report['warning_count'],
        'active_products': row['active_products'],
        'low_stock_products': row['low_stock_products'],
        'pending_orders': row['pending_orders'],
        'return_requests': row['return_requests'],
        'warning_24h': row['warning_24h'],
        'critical_7d': row['critical_7d'],
        'audit_7d': row['audit_7d'],
        'last_order_at': row['last_order_at'],
    })


BACKUP_FILENAME_RE = re.compile(r'^adhya_backup_\d{8}_\d{6}\.zip$')
BACKUP_RETENTION_DAYS = 7


def _backup_dir():
    return os.path.join(os.path.dirname(__file__), 'backups')


def _backup_paths():
    root = _backup_dir()
    if not os.path.isdir(root):
        return []
    backup_root = os.path.abspath(root)
    paths = []
    for name in os.listdir(root):
        if not BACKUP_FILENAME_RE.fullmatch(name):
            continue
        path = os.path.abspath(os.path.join(root, name))
        try:
            if os.path.commonpath([backup_root, path]) != backup_root:
                continue
        except ValueError:
            continue
        if os.path.isfile(path):
            paths.append(path)
    return paths


def _safe_backup_path(filename):
    name = clean_text(filename, 80)
    if not BACKUP_FILENAME_RE.fullmatch(name):
        return None
    path = os.path.abspath(os.path.join(_backup_dir(), name))
    backup_root = os.path.abspath(_backup_dir())
    if not path.startswith(backup_root + os.sep):
        return None
    if not os.path.isfile(path):
        return None
    return path


def _read_backup_manifest(path):
    try:
        with zipfile.ZipFile(path, 'r') as zf:
            if 'backup_manifest.json' in zf.namelist():
                with zf.open('backup_manifest.json') as f:
                    return json.loads(f.read().decode('utf-8'))
    except Exception:
        return None
    return None


def _cleanup_old_backups(retention_days=BACKUP_RETENTION_DAYS):
    retention_days = max(1, int(retention_days or BACKUP_RETENTION_DAYS))
    now = datetime.datetime.now(datetime.timezone.utc)
    cutoff = now - datetime.timedelta(days=retention_days)
    result = {
        'retention_days': retention_days,
        'cutoff_utc': cutoff.isoformat(timespec='seconds').replace('+00:00', 'Z'),
        'deleted': [],
        'errors': [],
        'kept_newest': '',
        'total_before': 0,
        'total_after': 0,
    }
    paths = _backup_paths()
    result['total_before'] = len(paths)
    result['total_after'] = len(paths)
    if not paths:
        return result

    newest = max(paths, key=os.path.getmtime)
    result['kept_newest'] = os.path.basename(newest)
    backup_root = os.path.abspath(_backup_dir())

    for path in paths:
        if os.path.abspath(path) == os.path.abspath(newest):
            continue
        try:
            modified_at = datetime.datetime.fromtimestamp(os.path.getmtime(path), datetime.timezone.utc)
            if modified_at >= cutoff:
                continue
            abs_path = os.path.abspath(path)
            if os.path.commonpath([backup_root, abs_path]) != backup_root:
                continue
            size = os.path.getsize(abs_path)
            filename = os.path.basename(abs_path)
            os.remove(abs_path)
            result['deleted'].append({
                'filename': filename,
                'bytes': size,
                'size_mb': round(size / (1024 * 1024), 2),
                'modified_at': modified_at.isoformat(timespec='seconds').replace('+00:00', 'Z'),
            })
        except Exception as exc:
            result['errors'].append({
                'filename': os.path.basename(path),
                'error': str(exc)[:300],
            })

    result['total_after'] = max(0, result['total_before'] - len(result['deleted']))
    return result


def _log_backup_cleanup(cleanup):
    deleted = cleanup.get('deleted') or []
    if not deleted:
        return
    log_admin_action('backup_retention_cleaned', 'Old backups removed by 7-day retention policy', {
        'entity_type': 'backup',
        'entity_id': 'retention',
        'retention_days': cleanup.get('retention_days'),
        'cutoff_utc': cleanup.get('cutoff_utc'),
        'deleted_count': len(deleted),
        'deleted_files': [item.get('filename') for item in deleted],
        'kept_newest': cleanup.get('kept_newest'),
        'errors': cleanup.get('errors') or [],
    })


def _backup_payload(path, include_manifest=False):
    stat = os.stat(path)
    manifest = _read_backup_manifest(path)
    payload = {
        'filename': os.path.basename(path),
        'bytes': stat.st_size,
        'size_mb': round(stat.st_size / (1024 * 1024), 2),
        'modified_at': datetime.datetime.fromtimestamp(stat.st_mtime, datetime.timezone.utc).isoformat(timespec='seconds').replace('+00:00', 'Z'),
        'has_json_manifest': bool(manifest),
        'created_utc': (manifest or {}).get('created_utc', ''),
        'database_sha256': ((manifest or {}).get('database') or {}).get('sha256', ''),
        'uploads_count': ((manifest or {}).get('uploads') or {}).get('count', None),
        'private_bills_count': ((manifest or {}).get('private_bills') or {}).get('count', None),
        'source_files_count': ((manifest or {}).get('source_files') or {}).get('count', None),
    }
    if include_manifest:
        payload['manifest'] = manifest
    return payload


@app.route('/api/admin/backups', methods=['GET'])
@admin_only_required
def admin_list_backups():
    cleanup = _cleanup_old_backups()
    _log_backup_cleanup(cleanup)
    backups = []
    for path in _backup_paths():
        backups.append(_backup_payload(path))
    backups.sort(key=lambda b: b['modified_at'], reverse=True)
    return jsonify({
        'backups': backups[:50],
        'count': len(backups),
        'retention_days': BACKUP_RETENTION_DAYS,
        'retention_cleanup': cleanup,
    })


@app.route('/api/admin/backups', methods=['POST'])
@limiter.limit("3 per hour")
@admin_only_required
def admin_create_backup():
    try:
        import backup_database
        result = backup_database.create_backup(
            db=DB_PATH,
            out_dir=_backup_dir(),
            uploads=UPLOAD_FOLDER,
            private_bills=os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'private_bills')),
            verify=True,
            keep=0,
            retention_days=0,
        )
    except Exception as exc:
        log_security_event(
            'backup_failed',
            'critical',
            'Admin backup creation failed',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
            metadata={'error': str(exc)[:500]},
        )
        return jsonify({'error': 'Backup could not be created: ' + str(exc)}), 500
    log_admin_action('backup_created', 'Backup created and verified', {
        'entity_type': 'backup',
        'entity_id': result['filename'],
        'filename': result['filename'],
        'bytes': result['bytes'],
        'sha256': result['sha256'],
        'verified': result['verified'],
    })
    cleanup = _cleanup_old_backups()
    _log_backup_cleanup(cleanup)
    payload = _backup_payload(result['zip_path'], include_manifest=True)
    payload['sha256'] = result['sha256']
    payload['verified'] = result['verified']
    return jsonify({
        'message': 'Backup created and verified',
        'backup': payload,
        'retention_days': BACKUP_RETENTION_DAYS,
        'retention_cleanup': cleanup,
    })


@app.route('/api/admin/backups/<filename>/verify', methods=['POST'])
@limiter.limit("10 per hour")
@admin_only_required
def admin_verify_backup(filename):
    path = _safe_backup_path(filename)
    if not path:
        return jsonify({'error': 'Backup not found'}), 404
    try:
        import backup_database
        backup_database.verify_backup_zip(path)
    except Exception as exc:
        log_security_event(
            'backup_verify_failed',
            'critical',
            'Backup verification failed',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
            metadata={'filename': filename, 'error': str(exc)[:500]},
        )
        return jsonify({'error': 'Backup verification failed: ' + str(exc)}), 400
    log_admin_action('backup_verified', 'Backup verified', {
        'entity_type': 'backup',
        'entity_id': os.path.basename(path),
        'filename': os.path.basename(path),
    })
    return jsonify({'message': 'Backup verified', 'backup': _backup_payload(path, include_manifest=True)})


@app.route('/api/admin/backups/<filename>/restore-drill', methods=['POST'])
@limiter.limit("6 per hour")
@admin_only_required
def admin_backup_restore_drill(filename):
    path = _safe_backup_path(filename)
    if not path:
        return jsonify({'error': 'Backup not found'}), 404
    try:
        import backup_database
        report = backup_database.restore_drill_report(path)
    except Exception as exc:
        log_security_event(
            'backup_restore_drill_failed',
            'critical',
            'Backup restore drill could not run',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
            metadata={'filename': filename, 'error': str(exc)[:500]},
        )
        return jsonify({'error': 'Backup restore drill failed: ' + str(exc)}), 400

    if not report.get('ok'):
        log_security_event(
            'backup_restore_drill_failed',
            'critical',
            'Backup restore drill found problems',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
            metadata={
                'filename': os.path.basename(path),
                'checks': report.get('checks', [])[:20],
                'archive': report.get('archive', {}),
                'database': {
                    'integrity_check': (report.get('database') or {}).get('integrity_check'),
                    'missing_tables': (report.get('database') or {}).get('missing_tables'),
                    'foreign_key_issue_count': (report.get('database') or {}).get('foreign_key_issue_count'),
                },
            },
        )
        return jsonify({'message': 'Backup restore drill found problems', 'report': report}), 200

    log_admin_action('backup_restore_drill', 'Backup restore drill passed', {
        'entity_type': 'backup',
        'entity_id': os.path.basename(path),
        'filename': os.path.basename(path),
        'zip_bytes': report.get('zip_bytes'),
        'archive': report.get('archive', {}),
        'database': {
            'integrity_check': (report.get('database') or {}).get('integrity_check'),
            'table_counts': (report.get('database') or {}).get('table_counts', {}),
        },
    })
    return jsonify({'message': 'Backup restore drill passed', 'report': report})


@app.route('/api/admin/backups/<filename>/download', methods=['GET'])
@admin_only_required
def admin_download_backup(filename):
    path = _safe_backup_path(filename)
    if not path:
        return jsonify({'error': 'Backup not found'}), 404
    log_admin_action('backup_downloaded', 'Backup downloaded', {
        'entity_type': 'backup',
        'entity_id': os.path.basename(path),
        'filename': os.path.basename(path),
        'bytes': os.path.getsize(path),
    })
    resp = send_from_directory(
        _backup_dir(),
        os.path.basename(path),
        as_attachment=True,
        download_name=os.path.basename(path),
        mimetype='application/zip',
    )
    return secure_upload_headers(resp, attachment=True)


# ─── Dashboard Stats ──────────────────────────────────────────────────────────

@app.route('/api/admin/stats', methods=['GET'])
@admin_only_required
def get_stats():
    db = get_db()
    total_orders = db.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
    total_revenue = db.execute("SELECT COALESCE(SUM(total),0) FROM orders WHERE payment_status='paid'").fetchone()[0]
    total_products = db.execute("SELECT COUNT(*) FROM products WHERE is_active=1").fetchone()[0]
    total_customers = db.execute("SELECT COUNT(*) FROM users WHERE role='customer'").fetchone()[0]
    recent_orders = rows_to_list(db.execute(
        "SELECT id,order_number,customer_name,total,status,created_at FROM orders ORDER BY created_at DESC LIMIT 5"
    ).fetchall())
    order_stats = rows_to_list(db.execute(
        "SELECT status, COUNT(*) as count FROM orders GROUP BY status"
    ).fetchall())
    pending_orders = db.execute("SELECT COUNT(*) FROM orders WHERE status='pending'").fetchone()[0]
    processing_orders = db.execute("SELECT COUNT(*) FROM orders WHERE status='processing'").fetchone()[0]
    return_requests = db.execute("SELECT COUNT(*) FROM orders WHERE status='return_requested'").fetchone()[0]
    old_pending_orders = db.execute(
        "SELECT COUNT(*) FROM orders WHERE status='pending' AND created_at <= datetime('now','-2 days')"
    ).fetchone()[0]
    missing_tracking_orders = db.execute(
        "SELECT COUNT(*) FROM orders WHERE status='shipped' AND IFNULL(TRIM(tracking_number),'')=''"
    ).fetchone()[0]
    low_stock = rows_to_list(db.execute(
        "SELECT id,name,stock,images FROM products WHERE is_active=1 AND stock<=5 ORDER BY stock ASC LIMIT 8"
    ).fetchall())
    for p in low_stock:
        p['images'] = json.loads(p['images']) if p['images'] else []
    return jsonify({
        'total_orders': total_orders, 'total_revenue': total_revenue,
        'total_products': total_products, 'total_customers': total_customers,
        'recent_orders': recent_orders, 'order_stats': order_stats,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'return_requests': return_requests,
        'old_pending_orders': old_pending_orders,
        'missing_tracking_orders': missing_tracking_orders,
        'low_stock': low_stock,
    })


# ─── Public Order Tracking ────────────────────────────────────────────────────

@app.route('/api/orders/track', methods=['POST'])
@limiter.limit("10 per minute; 60 per hour")
def track_order():
    data = request.json or {}
    order_number = clean_text(data.get('order_number'), 40).upper()
    raw_email    = (data.get('email') or '').strip().lower()
    email        = normalize_public_email(raw_email)
    limited = public_abuse_guard(
        'order_track',
        email=email or raw_email,
        fingerprint=order_number,
        ip_limit=16,
        email_limit=8,
        fingerprint_limit=8,
        window_seconds=900,
    )
    if limited:
        return limited
    if not order_number or not email:
        return jsonify({'error': 'Order number and email are required'}), 400
    if not re.fullmatch(r'[A-Z0-9_-]{4,40}', order_number):
        return jsonify({'error': 'No order found. Please check your order number and email address.'}), 404
    db = get_db()
    o = row_to_dict(db.execute(
        "SELECT * FROM orders WHERE order_number=? AND LOWER(customer_email)=?",
        (order_number, email)
    ).fetchone())
    if not o:
        log_security_event(
            'order_track_failed',
            'warning',
            'Public order tracking lookup failed',
            email=email,
            metadata={'order_number': order_number[:12]},
        )
        return jsonify({'error': 'No order found. Please check your order number and email address.'}), 404
    return jsonify(public_order_payload(o))


# ─── Bulk Orders ──────────────────────────────────────────────────────────────

@app.route('/api/bulk-order', methods=['POST'])
@limiter.limit("5 per hour")
def create_bulk_order():
    data = request.json or {}
    if data.get('website'):  # honeypot
        log_security_event('public_honeypot_triggered', 'warning', 'Bulk inquiry honeypot was filled', metadata={'action': 'bulk_order'})
        return jsonify({'message': 'Bulk order request received'}), 200
    name  = clean_text(data.get('name'), 100)
    raw_email = clean_text(data.get('email'), 140).lower()
    email = normalize_public_email(raw_email)
    raw_phone = clean_text(data.get('phone'), 30)
    phone = normalize_public_phone(raw_phone) if raw_phone else ''
    product_type = clean_text(data.get('product_type'), 80)
    quantity = clean_text(data.get('quantity'), 40)
    needed_by = clean_text(data.get('needed_by'), 40)
    message = clean_text(data.get('message'), 2000)
    limited = public_abuse_guard('bulk_order', email=email or raw_email, ip_limit=6, email_limit=3, window_seconds=3600)
    if limited:
        return limited
    if not name or not email:
        return jsonify({'error': 'Name and email are required'}), 400
    if raw_phone and not phone:
        return jsonify({'error': 'Please enter a valid phone number'}), 400
    if message and len(message) < 10:
        return jsonify({'error': 'Please add a little more detail so we can prepare an accurate quote.'}), 400
    if public_text_looks_spammy(name, product_type, quantity, needed_by, message):
        log_security_event('bulk_order_rejected', 'warning', 'Bulk inquiry looked unsafe or spammy', email=email)
        return jsonify({'error': 'Please remove unsupported text and try again.'}), 400
    file_url = clean_text(data.get('file_url'), 260)
    if file_url and not is_safe_public_upload_url(file_url):
        return jsonify({'error': 'Invalid uploaded file reference'}), 400
    db = get_db()
    request_id = str(uuid.uuid4())
    db.execute(
        "INSERT INTO bulk_orders (id,name,business_name,email,phone,product_type,quantity,needed_by,message,file_url) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (request_id, name, clean_text(data.get('business_name'), 120), email,
         phone, product_type, quantity, needed_by, message, file_url)
    )
    db.commit()
    email_bulk_order_notification(
        name, clean_text(data.get('business_name'), 120), email, phone,
        product_type, quantity, needed_by, message
    )
    return jsonify({'message': 'Bulk order request received', 'reference': request_id[:8].upper()})


# ─── Contact Form ─────────────────────────────────────────────────────────────

@app.route('/api/contact', methods=['POST'])
@limiter.limit("5 per hour")
def contact_form():
    """Stores contact messages in the DB so admin can review them."""
    data = request.json or {}
    if data.get('website'):  # honeypot — bots fill hidden fields, humans don't
        log_security_event('public_honeypot_triggered', 'warning', 'Contact form honeypot was filled', metadata={'action': 'contact'})
        return jsonify({'message': 'Message received'}), 200
    name         = clean_text(data.get('name'), 100)
    raw_email    = clean_text(data.get('email'), 140).lower()
    email        = normalize_public_email(raw_email)
    raw_phone    = clean_text(data.get('phone'), 30)
    phone        = normalize_public_phone(raw_phone) if raw_phone else ''
    message      = clean_text(data.get('message'), 2000)
    inquiry_type = clean_text(data.get('inquiry_type'), 80)
    order_number = clean_text(data.get('order_number'), 40)
    limited = public_abuse_guard('contact', email=email or raw_email, ip_limit=6, email_limit=3, window_seconds=3600)
    if limited:
        return limited
    if not name or not email or not message:
        return jsonify({'error': 'Name, email and message are required'}), 400
    if raw_phone and not phone:
        return jsonify({'error': 'Please enter a valid phone number'}), 400
    if len(message) < 10:
        return jsonify({'error': 'Please add a little more detail so we can help properly.'}), 400
    if public_text_looks_spammy(name, message, inquiry_type, order_number):
        log_security_event('contact_rejected', 'warning', 'Contact form looked unsafe or spammy', email=email)
        return jsonify({'error': 'Please remove unsupported text and try again.'}), 400
    db = get_db()
    message_id = str(uuid.uuid4())
    db.execute(
        "INSERT INTO contact_messages (id,name,email,phone,message,inquiry_type,order_number) VALUES (?,?,?,?,?,?,?)",
        (message_id, name, email, phone, message, inquiry_type, order_number)
    )
    db.commit()
    email_contact_notification(name, email, phone, message, inquiry_type, order_number)
    return jsonify({'message': 'Message received', 'reference': message_id[:8].upper()})


# ─── Upload ───────────────────────────────────────────────────────────────────

def _bulk_import_cleanup():
    cutoff = datetime.datetime.now() - datetime.timedelta(hours=24)
    try:
        for name in os.listdir(BULK_IMPORT_FOLDER):
            if not re.fullmatch(r'[0-9a-f]{32}', name):
                continue
            path = os.path.join(BULK_IMPORT_FOLDER, name)
            if os.path.isdir(path) and datetime.datetime.fromtimestamp(os.path.getmtime(path)) < cutoff:
                shutil.rmtree(path, ignore_errors=True)
    except Exception as exc:
        app.logger.warning('bulk import cleanup failed: %s', exc)


def _bulk_job_path(job_id):
    if not re.fullmatch(r'[0-9a-f]{32}', job_id or ''):
        raise ValueError('Invalid import job')
    path = os.path.abspath(os.path.join(BULK_IMPORT_FOLDER, job_id))
    if os.path.commonpath([os.path.abspath(BULK_IMPORT_FOLDER), path]) != os.path.abspath(BULK_IMPORT_FOLDER):
        raise ValueError('Invalid import job')
    return path


def _bulk_read_job(job_id):
    job_dir = _bulk_job_path(job_id)
    meta_path = os.path.join(job_dir, 'job.json')
    if not os.path.isfile(meta_path):
        raise ValueError('Import preview expired. Please upload the files again.')
    with open(meta_path, 'r', encoding='utf-8') as fh:
        return json.load(fh), job_dir


def _bulk_bool(value, default=False):
    if value is None:
        return default
    return str(value).strip().lower() in ('1', 'true', 'yes', 'on')


def _bulk_header_key(value):
    return re.sub(r'[^a-z0-9]+', '_', clean_text(value, 80).lower()).strip('_')


def _bulk_number(value, required=False):
    text = clean_text(value, 50)
    if not text:
        return None if not required else None
    try:
        amount = float(text)
    except (TypeError, ValueError):
        return None
    return round(amount, 2) if amount >= 0 else None


def _bulk_stock(value, force_zero=True):
    if force_zero:
        return 0
    try:
        return max(0, min(999999, int(float(value or 0))))
    except (TypeError, ValueError):
        return 0


def _bulk_sanitize_zip_name(name):
    base = os.path.basename((name or '').replace('\\', '/'))
    base = re.sub(r'[\x00-\x1f\x7f]', '', base).strip()
    base = re.sub(r'[^A-Za-z0-9._() \-]+', '-', base)
    base = re.sub(r'\s+', ' ', base)[:180].strip(' .')
    return base


class _BulkZipImageUpload:
    def __init__(self, filename, data):
        self.filename = filename
        self.stream = io.BytesIO(data or b'')
        self.mimetype = 'application/octet-stream'


def _bulk_extract_images(zip_file, images_dir):
    image_by_name = {}
    image_by_stem = {}
    warnings = []
    total_uncompressed = 0
    file_count = 0
    os.makedirs(images_dir, exist_ok=True)
    with zipfile.ZipFile(zip_file) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            if file_count >= 600:
                warnings.append('Only the first 600 image files were scanned from the ZIP.')
                break
            original = _bulk_sanitize_zip_name(info.filename)
            ext = os.path.splitext(original)[1].lower()
            if not original or ext not in SAFE_IMAGE_EXTENSIONS:
                continue
            total_uncompressed += int(info.file_size or 0)
            if total_uncompressed > 350 * 1024 * 1024:
                raise ValueError('Image ZIP is too large after extraction.')
            if int(info.file_size or 0) > PUBLIC_UPLOAD_MAX_BYTES:
                raise ValueError(f'Image ZIP contains a file larger than {PUBLIC_UPLOAD_MAX_BYTES // (1024 * 1024)} MB: {original}')
            target = os.path.abspath(os.path.join(images_dir, original))
            if os.path.commonpath([os.path.abspath(images_dir), target]) != os.path.abspath(images_dir):
                continue
            if os.path.exists(target):
                stem, suffix = os.path.splitext(original)
                original = f"{stem}-{file_count + 1}{suffix}"
                target = os.path.abspath(os.path.join(images_dir, original))
            with zf.open(info) as src:
                data = src.read(PUBLIC_UPLOAD_MAX_BYTES + 1)
            if len(data) > PUBLIC_UPLOAD_MAX_BYTES:
                raise ValueError(f'Image ZIP contains a file larger than {PUBLIC_UPLOAD_MAX_BYTES // (1024 * 1024)} MB: {original}')
            try:
                meta = validate_upload(_BulkZipImageUpload(original, data), SAFE_IMAGE_EXTENSIONS, PUBLIC_UPLOAD_MAX_BYTES)
            except UploadSecurityError as exc:
                raise ValueError(f'Image ZIP contains an unsafe image file ({original}): {exc}') from exc
            if meta.get('ext') and meta['ext'] != ext:
                stem, _old_ext = os.path.splitext(original)
                original = f"{stem}{meta['ext']}"
                target = os.path.abspath(os.path.join(images_dir, original))
                if os.path.exists(target):
                    original = f"{stem}-{file_count + 1}{meta['ext']}"
                    target = os.path.abspath(os.path.join(images_dir, original))
            with open(target, 'wb') as dst:
                dst.write(data)
            file_count += 1
            image_by_name[original.lower()] = original
            image_by_stem[os.path.splitext(original)[0].lower()] = original
    return image_by_name, image_by_stem, file_count, warnings


def _bulk_match_image(ref, image_by_name, image_by_stem):
    raw = clean_text(ref, 180)
    if not raw:
        return ''
    base = _bulk_sanitize_zip_name(raw)
    key = base.lower()
    return image_by_name.get(key) or image_by_stem.get(os.path.splitext(key)[0] if os.path.splitext(key)[1] else key) or ''


def _bulk_load_workbook_rows(xlsx_path, image_by_name, image_by_stem, force_stock_zero=True):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('Excel support is not installed. Run: pip3 install openpyxl') from exc

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = None
    for ws in wb.worksheets:
        headers = [clean_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        header_keys = {_bulk_header_key(h) for h in headers}
        if 'name' in header_keys and 'price' in header_keys:
            worksheet = ws
            break
    if worksheet is None:
        raise ValueError('Excel file must have a sheet with at least name and price columns.')

    raw_headers = [clean_text(v) for v in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    headers = [_bulk_header_key(h) for h in raw_headers]
    aliases = {
        'product_name': 'name',
        'title': 'name',
        'selling_price': 'price',
        'sale_price': 'price',
        'regular_price': 'compare_price',
        'compare_at_price': 'compare_price',
        'image_filename_1': 'image_1',
        'image_filename_2': 'image_2',
        'image_filename_3': 'image_3',
        'image1': 'image_1',
        'image2': 'image_2',
        'image3': 'image_3',
        'bestseller': 'is_bestseller',
        'best_seller': 'is_bestseller',
    }
    headers = [aliases.get(h, h) for h in headers]
    image_cols = [i for i, h in enumerate(headers) if h.startswith('image')][:8]
    idx = {h: i for i, h in enumerate(headers) if h}
    rows = []

    for excel_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v not in (None, '') for v in values):
            continue
        get = lambda key: values[idx[key]] if key in idx and idx[key] < len(values) else ''
        name = clean_text(get('name'), 160)
        description = clean_text(get('description'), 5000)
        price = _bulk_number(get('price'), required=True)
        compare_price = _bulk_number(get('compare_price'))
        stock = _bulk_stock(get('stock'), force_zero=force_stock_zero)
        sku = clean_text(get('sku'), 80)
        is_bestseller = _bulk_bool(get('is_bestseller'), False)
        errors = []
        if not name:
            errors.append('Missing product name')
        if price is None or price <= 0:
            errors.append('Missing or invalid price')
        image_names = []
        missing_images = []
        for col in image_cols:
            ref = values[col] if col < len(values) else ''
            if not clean_text(ref):
                continue
            match = _bulk_match_image(ref, image_by_name, image_by_stem)
            if match:
                image_names.append(match)
            else:
                missing_images.append(clean_text(ref, 120))
        if missing_images:
            errors.append('Missing image file(s): ' + ', '.join(missing_images[:4]))
        if not image_names:
            errors.append('At least one matching image is required')
        rows.append({
            'row': excel_row,
            'name': name,
            'description': description,
            'price': price,
            'compare_price': compare_price,
            'stock': stock,
            'sku': sku,
            'is_bestseller': 1 if is_bestseller else 0,
            'image_files': image_names,
            'errors': errors,
        })
    if not rows:
        raise ValueError('No product rows were found in the Excel file.')
    return rows, worksheet.title


def _bulk_existing_products(db):
    by_name = {}
    by_sku = {}
    for row in db.execute("SELECT id,name,sku,price,compare_price,images,stock FROM products").fetchall():
        item = dict(row)
        if item.get('name'):
            by_name[item['name'].casefold()] = item
        if item.get('sku'):
            by_sku[item['sku'].casefold()] = item
    return {'by_name': by_name, 'by_sku': by_sku}


def _bulk_find_existing(row, existing):
    sku = (row.get('sku') or '').casefold()
    name = (row.get('name') or '').casefold()
    if sku and existing['by_sku'].get(sku):
        return existing['by_sku'][sku], 'sku'
    if name and existing['by_name'].get(name):
        return existing['by_name'][name], 'name'
    return None, ''


def _bulk_product_preview_rows(rows, existing, job_id=''):
    out = []
    seen_names = {}
    seen_skus = {}
    for row in rows:
        errors = list(row.get('errors') or [])
        warnings = []
        name_key = (row.get('name') or '').casefold()
        sku_key = (row.get('sku') or '').casefold()
        if name_key and name_key in seen_names:
            errors.append(f'Duplicate product name in Excel file; first seen on row {seen_names[name_key]}')
        elif name_key:
            seen_names[name_key] = row['row']
        if sku_key and sku_key in seen_skus:
            errors.append(f'Duplicate SKU in Excel file; first seen on row {seen_skus[sku_key]}')
        elif sku_key:
            seen_skus[sku_key] = row['row']
        exists, match = _bulk_find_existing(row, existing)
        if exists:
            warnings.append(f'Existing product match by {match}: {exists.get("name") or ""}')
        status = 'error' if errors else ('existing' if exists else 'ready')
        image_files = row.get('image_files') or []
        out.append({
            'row': row['row'],
            'name': row['name'],
            'sku': row.get('sku') or '',
            'price': row['price'],
            'compare_price': row.get('compare_price'),
            'stock': row['stock'],
            'images': len(image_files),
            'thumbs': [
                f'/api/admin/bulk-products/{job_id}/image/{quote(image_name)}'
                for image_name in image_files[:3]
            ] if job_id else [],
            'status': status,
            'existing_id': exists.get('id') if exists else '',
            'existing_match': match,
            'existing_stock': exists.get('stock') if exists else None,
            'errors': errors,
            'warnings': warnings,
        })
    return out


def _bulk_backup_database():
    try:
        import backup_database
        result = backup_database.create_backup(
            db=DB_PATH,
            out_dir=_backup_dir(),
            verify=True,
            keep=30,
            retention_days=BACKUP_RETENTION_DAYS,
        )
        return result['zip_path']
    except Exception as exc:
        log_security_event(
            'backup_failed',
            'critical',
            'Automatic backup before bulk import failed',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
            metadata={'error': str(exc)[:300], 'source': 'bulk_product_import'},
        )
        raise


def _bulk_insert_product(db, row, category_id, image_urls):
    pid = str(uuid.uuid4())
    db.execute(
        """INSERT INTO products
           (id,name,description,price,compare_price,category_id,stock,sku,images,variations,
            is_active,allow_custom_print,is_bestseller,cost_price,low_stock_threshold)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            pid,
            row['name'],
            row.get('description') or '',
            row['price'],
            row.get('compare_price'),
            category_id,
            int(row.get('stock') or 0),
            row.get('sku') or '',
            json.dumps(image_urls),
            json.dumps([]),
            1,
            0,
            1 if row.get('is_bestseller') else 0,
            0,
            5,
        )
    )
    return pid


@app.route('/api/admin/bulk-products/preview', methods=['POST'])
@limiter.limit("6 per hour")
@admin_only_required
def admin_bulk_products_preview():
    if request.content_length and request.content_length > ADMIN_BULK_UPLOAD_MAX_BYTES:
        return jsonify({'error': 'Bulk upload is too large. Keep the Excel and image ZIP under 150 MB.'}), 413
    _bulk_import_cleanup()
    sheet = request.files.get('sheet')
    images_zip = request.files.get('images_zip')
    category_id = clean_text(request.form.get('category_id'), 80)
    force_stock_zero = _bulk_bool(request.form.get('force_stock_zero'), True)
    if not sheet or not images_zip:
        return jsonify({'error': 'Excel file and image ZIP are required'}), 400
    if not sheet.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Please upload an .xlsx Excel file'}), 400
    if not images_zip.filename.lower().endswith('.zip'):
        return jsonify({'error': 'Please upload images as a .zip file'}), 400
    db = get_db()
    if not db.execute("SELECT id FROM categories WHERE id=? AND IFNULL(is_active,1)=1", (category_id,)).fetchone():
        return jsonify({'error': 'Please select an active category'}), 400

    job_id = secrets.token_hex(16)
    job_dir = _bulk_job_path(job_id)
    images_dir = os.path.join(job_dir, 'images')
    os.makedirs(job_dir, exist_ok=True)
    xlsx_path = os.path.join(job_dir, 'products.xlsx')
    zip_path = os.path.join(job_dir, 'images.zip')
    sheet.save(xlsx_path)
    images_zip.save(zip_path)

    try:
        image_by_name, image_by_stem, image_count, warnings = _bulk_extract_images(zip_path, images_dir)
        rows, sheet_name = _bulk_load_workbook_rows(xlsx_path, image_by_name, image_by_stem, force_stock_zero=force_stock_zero)
        existing = _bulk_existing_products(db)
        preview_rows = _bulk_product_preview_rows(rows, existing, job_id)
        errors = sum(1 for r in preview_rows if r['status'] == 'error')
        ready = sum(1 for r in preview_rows if r['status'] == 'ready')
        existing_count = sum(1 for r in preview_rows if r['status'] == 'existing')
        meta = {
            'job_id': job_id,
            'created_at': datetime.datetime.now().isoformat(timespec='seconds'),
            'sheet_name': sheet_name,
            'category_id': category_id,
            'force_stock_zero': force_stock_zero,
            'rows': rows,
            'image_count': image_count,
            'warnings': warnings,
        }
        with open(os.path.join(job_dir, 'job.json'), 'w', encoding='utf-8') as fh:
            json.dump(meta, fh)
        log_admin_action('bulk_product_previewed', 'Bulk product import preview created', {
            'job_id': job_id,
            'entity_type': 'product',
            'rows': len(rows),
            'ready': ready,
            'existing': existing_count,
            'errors': errors,
        })
        return jsonify({
            'job_id': job_id,
            'summary': {
                'rows': len(rows),
                'ready': ready,
                'existing': existing_count,
                'errors': errors,
                'images': image_count,
                'sheet': sheet_name,
                'force_stock_zero': force_stock_zero,
            },
            'warnings': warnings,
            'rows': preview_rows[:500],
        })
    except Exception:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise


@app.route('/api/admin/bulk-products/<job_id>/image/<path:filename>', methods=['GET'])
@admin_only_required
def admin_bulk_product_preview_image(job_id, filename):
    try:
        meta, job_dir = _bulk_read_job(job_id)
    except ValueError:
        return jsonify({'error': 'Preview job not found'}), 404
    safe_name = _bulk_sanitize_zip_name(filename)
    if not safe_name or safe_name != filename:
        return jsonify({'error': 'Image not found'}), 404
    allowed = {name for row in (meta.get('rows') or []) for name in (row.get('image_files') or [])}
    if safe_name not in allowed:
        return jsonify({'error': 'Image not found'}), 404
    images_dir = os.path.abspath(os.path.join(job_dir, 'images'))
    path = os.path.abspath(os.path.join(images_dir, safe_name))
    if os.path.commonpath([images_dir, path]) != images_dir or not os.path.isfile(path):
        return jsonify({'error': 'Image not found'}), 404
    return send_from_directory(images_dir, safe_name)


@app.route('/api/admin/bulk-products/commit', methods=['POST'])
@limiter.limit("6 per hour")
@admin_only_required
def admin_bulk_products_commit():
    data = request.json or {}
    job_id = clean_text(data.get('job_id'), 40)
    update_existing_images = _bulk_bool(data.get('update_existing_images'), True)
    mode = clean_text(data.get('mode'), 40).lower() or 'create_update'
    if mode not in ('create_update', 'update_price_images'):
        mode = 'create_update'
    try:
        meta, job_dir = _bulk_read_job(job_id)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    images_dir = os.path.join(job_dir, 'images')
    db = get_db()
    category_id = clean_text(meta.get('category_id'), 80)
    if not db.execute("SELECT id FROM categories WHERE id=? AND IFNULL(is_active,1)=1", (category_id,)).fetchone():
        return jsonify({'error': 'The selected category is no longer active'}), 400
    existing = _bulk_existing_products(db)
    rows = meta.get('rows') or []
    preview_rows = _bulk_product_preview_rows(rows, existing, job_id)
    valid_row_numbers = {r['row'] for r in preview_rows if r.get('status') != 'error'}
    valid_rows = [r for r in rows if r.get('row') in valid_row_numbers]
    if not valid_rows:
        return jsonify({'error': 'There are no valid products to import'}), 400

    try:
        backup_path = _bulk_backup_database()
    except Exception as exc:
        return jsonify({'error': 'Bulk import stopped because the automatic backup could not be created: ' + str(exc)}), 500

    imported = 0
    updated_images = 0
    updated_price_images = 0
    skipped_existing = 0
    failed = []

    try:
        db.execute("BEGIN IMMEDIATE")
        for row in valid_rows:
            exists, match = _bulk_find_existing(row, existing)
            image_urls = []
            for image_name in row.get('image_files') or []:
                src = os.path.abspath(os.path.join(images_dir, image_name))
                if os.path.commonpath([os.path.abspath(images_dir), src]) != os.path.abspath(images_dir) or not os.path.isfile(src):
                    failed.append({'name': row.get('name'), 'error': f'Missing image {image_name}'})
                    image_urls = []
                    break
                fname, _meta = _save_public_image_path(src)
                image_urls.append(f'/uploads/{fname}')
            if not image_urls:
                continue
            if exists:
                if mode == 'update_price_images':
                    db.execute(
                        "UPDATE products SET price=?, compare_price=?, images=? WHERE id=?",
                        (row['price'], row.get('compare_price'), json.dumps(image_urls), exists['id'])
                    )
                    updated_price_images += 1
                    updated_images += 1
                elif update_existing_images:
                    db.execute("UPDATE products SET images=? WHERE id=?", (json.dumps(image_urls), exists['id']))
                    updated_images += 1
                else:
                    skipped_existing += 1
                continue
            if mode == 'update_price_images':
                skipped_existing += 1
                continue
            _bulk_insert_product(db, row, category_id, image_urls)
            imported += 1
        db.commit()
    except Exception as exc:
        db.rollback()
        log_security_event('bulk_product_import_failed', 'critical', 'Bulk product import failed', user_id=g.user.get('id'), email=g.user.get('email'), metadata={'job_id': job_id, 'error': str(exc)[:300]})
        return jsonify({'error': 'Bulk import failed before it could finish. The database backup is still available.'}), 500

    log_admin_action('bulk_product_imported', 'Bulk product import committed', {
        'job_id': job_id,
        'entity_type': 'product',
        'imported': imported,
        'updated_images': updated_images,
        'updated_price_images': updated_price_images,
        'skipped_existing': skipped_existing,
        'failed': len(failed),
        'backup': os.path.basename(backup_path),
        'mode': mode,
    })
    return jsonify({
        'message': 'Bulk import complete',
        'imported': imported,
        'updated_images': updated_images,
        'updated_price_images': updated_price_images,
        'skipped_existing': skipped_existing,
        'failed': failed,
        'backup': os.path.basename(backup_path),
        'mode': mode,
    })

def _prepare_public_image(image):
    image = ImageOps.exif_transpose(image)
    max_pixels = int(os.environ.get('UPLOAD_IMAGE_MAX_PIXELS', '20000000') or 20000000)
    width, height = image.size
    if width <= 0 or height <= 0 or (width * height) > max_pixels:
        raise UploadSecurityError('Image dimensions are too large')
    if image.mode not in ('RGB', 'RGBA'):
        image = image.convert('RGBA' if 'A' in image.getbands() else 'RGB')
    max_side = int(os.environ.get('UPLOAD_IMAGE_MAX_SIDE', '1600') or 1600)
    if max(image.size) > max_side:
        image.thumbnail((max_side, max_side), Image.Resampling.LANCZOS)
    return image


def _store_public_image_object(image):
    base = os.path.abspath(UPLOAD_FOLDER)
    save_attempts = (
        ('.webp', 'WEBP', {'quality': 82, 'method': 6}),
        ('.png', 'PNG', {'optimize': True}),
    )
    for ext, fmt, kwargs in save_attempts:
        fname = random_stored_name(ext)
        path = os.path.abspath(os.path.join(base, fname))
        try:
            if os.path.commonpath([base, path]) != base:
                raise UploadSecurityError('Unsafe upload path')
            image.save(path, fmt, **kwargs)
            return fname, {
                'stored_ext': ext,
                'stored_size': os.path.getsize(path),
                'compressed': True,
                'width': image.size[0],
                'height': image.size[1],
            }
        except UploadSecurityError:
            raise
        except Exception:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except OSError:
                pass  # nosec B110
            continue
    raise UploadSecurityError('Image could not be safely processed')


def _save_public_upload_image(file_storage, meta):
    if not _has_pillow:
        raise UploadSecurityError('Image processing is unavailable. Please try again later.')
    try:
        file_storage.stream.seek(0)
        image = Image.open(file_storage.stream)
        image.verify()
        file_storage.stream.seek(0)
        image = Image.open(file_storage.stream)
        image.load()
        image = _prepare_public_image(image)
        return _store_public_image_object(image)
    except UploadSecurityError:
        raise
    except Exception as exc:
        raise UploadSecurityError('Image could not be safely processed') from exc


def _save_public_image_path(path):
    if not _has_pillow:
        raise UploadSecurityError('Image processing is unavailable. Please try again later.')
    try:
        with Image.open(path) as image:
            image.verify()
        with Image.open(path) as image:
            image.load()
            image = _prepare_public_image(image)
            return _store_public_image_object(image)
    except UploadSecurityError:
        raise
    except Exception as exc:
        raise UploadSecurityError('Image could not be safely processed') from exc


@app.route('/api/upload', methods=['POST'])
@limiter.limit("20 per minute; 100 per hour")
@token_required
def upload_file():
    from flask import request as req
    if 'file' not in req.files:
        log_security_event(
            'upload_rejected',
            'warning',
            'Upload request did not include a file',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
        )
        return jsonify({'error': 'No file'}), 400
    f = req.files['file']
    try:
        meta = validate_upload(f, SAFE_IMAGE_EXTENSIONS, PUBLIC_UPLOAD_MAX_BYTES)
    except UploadSecurityError as exc:
        log_security_event(
            'upload_rejected',
            'warning',
            'Blocked unsafe public upload',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
            metadata={'filename': f.filename, 'reason': str(exc)},
        )
        return jsonify({'error': str(exc)}), 400
    try:
        fname, stored_meta = _save_public_upload_image(f, meta)
    except UploadSecurityError as exc:
        log_security_event(
            'upload_rejected',
            'warning',
            'Blocked unsafe public upload during image processing',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
            metadata={'filename': f.filename, 'reason': str(exc)},
        )
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        log_security_event(
            'upload_failed',
            'critical',
            'Public upload failed while storing file',
            user_id=g.user.get('id'),
            email=g.user.get('email'),
            metadata={'filename': f.filename, 'error': str(exc)[:300]},
        )
        return jsonify({'error': 'Upload failed. Please try again.'}), 500
    log_security_event(
        'upload_accepted',
        'info',
        'Public upload accepted',
        user_id=g.user.get('id'),
        email=g.user.get('email'),
        metadata={
            'stored_name': fname,
            'role': g.user.get('role'),
            'ext': meta['ext'],
            'kind': meta.get('kind'),
            'size': meta['size'],
            **stored_meta,
        },
    )
    return jsonify({'url': f'/uploads/{fname}'})


@app.route('/uploads/<filename>')
def serve_upload(filename):
    if not is_safe_stored_filename(filename) or not is_safe_public_upload_url(f'/uploads/{filename}'):
        abort(404)
    ext = os.path.splitext(filename)[1].lower()
    mimetype = {
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.webp': 'image/webp',
    }.get(ext, 'application/octet-stream')
    resp = send_from_directory(UPLOAD_FOLDER, filename, mimetype=mimetype)
    return secure_upload_headers(resp)


# ─── SPA Fallback ─────────────────────────────────────────────────────────────
#
# Cache-busting: CSS/JS files keep the same filename on every deploy, so
# Cloudflare (and browsers) can keep serving a stale cached copy after a fix
# ships even though the origin has the new file. Local CSS/JS <link>/<script>
# tags in the HTML shells get a "?v=<timestamp>" suffix computed from the
# newest file mtime under client/js, client/css, and client/accounts — any code
# change bumps the version automatically, so the URL changes and old cached
# copies are bypassed without anyone needing to manually purge Cloudflare.

def get_asset_version():
    latest = 0.0
    css_path = os.path.join(CLIENT_DIR, 'css', 'style.css')
    if os.path.isfile(css_path):
        latest = max(latest, os.path.getmtime(css_path))
    js_dir = os.path.join(CLIENT_DIR, 'js')
    for root, _dirs, files in os.walk(js_dir):
        for f in files:
            latest = max(latest, os.path.getmtime(os.path.join(root, f)))
    accounts_dir = os.path.join(CLIENT_DIR, 'accounts')
    for root, _dirs, files in os.walk(accounts_dir):
        for f in files:
            latest = max(latest, os.path.getmtime(os.path.join(root, f)))
    return str(int(latest))


def inject_asset_version(html, version):
    return re.sub(
        r'(href|src)="(/(?:(?:css|js)|accounts/(?:css|js))/[^"]+)"',
        rf'\1="\2?v={version}"',
        html
    )


@app.route('/admin')
@app.route('/admin/')
@app.route('/admin/<path:admin_path>')
def serve_admin_spa(admin_path=''):
    with open(os.path.join(CLIENT_DIR, 'admin.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    html = inject_asset_version(html, get_asset_version())
    resp = Response(html, mimetype='text/html')
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


def _looks_sensitive_static_probe(path):
    p = (path or '').replace('\\', '/').strip('/').lower()
    if not p:
        return False
    parts = [x for x in p.split('/') if x]
    if any(x in ('.git', '.hg', '.svn', '__pycache__') for x in parts):
        return True
    if p in ('private', 'private_bills') or p.startswith(('server/', 'private/', 'private_bills/', 'instance/')):
        return True
    if p in ('.env', 'server/.env'):
        return True
    sensitive_exts = (
        '.env', '.db', '.sqlite', '.sqlite3', '.py', '.pyc', '.pyo',
        '.zip', '.tar', '.gz', '.bak', '.sql', '.pem', '.key', '.crt',
    )
    return p.endswith(sensitive_exts)


@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_spa(path):
    if path.startswith('api/') or path.startswith('uploads/'):
        return jsonify({'error': 'Not found'}), 404
    if _looks_sensitive_static_probe(path):
        return jsonify({'error': 'Not found'}), 404
    if path:
        file_path = os.path.join(CLIENT_DIR, path)
        if os.path.isfile(file_path):
            resp = send_from_directory(CLIENT_DIR, path)
            if path.startswith('css/') or path.startswith('js/'):
                # Safe to cache long-term — any future change produces a new
                # ?v= URL, so a stale cached copy of THIS exact URL is never wrong.
                resp.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
            return resp
    shell_name = 'admin.html' if (path == 'admin' or path.startswith('admin/')) else 'index.html'
    with open(os.path.join(CLIENT_DIR, shell_name), 'r', encoding='utf-8') as f:
        html = f.read()
    html = inject_asset_version(html, get_asset_version())
    resp = Response(html, mimetype='text/html')
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


# ─── Accounts / bookkeeping module (API under /api/acc, UI at /accounts) ──────
ACC_BILLS_DIR = os.path.join(os.path.dirname(__file__), '..', 'private_bills')
ACC_CLIENT_DIR = os.path.join(CLIENT_DIR, 'accounts')
import accounts_module
accounts_module.register(app, {
    'get_db': get_db, 'rows_to_list': rows_to_list, 'row_to_dict': row_to_dict,
    'admin_required': admin_required, 'admin_only_required': admin_only_required,
    'bills_dir': ACC_BILLS_DIR, 'db_path': DB_PATH,
    'log_security_event': log_security_event,
    'log_audit_event': log_audit_event,
    'notify_back_in_stock': notify_back_in_stock_for_product,
})


@app.route('/accounts')
@app.route('/accounts/<path:path>')
def serve_accounts(path=''):
    if path:
        fp = os.path.join(ACC_CLIENT_DIR, path)
        if os.path.isfile(fp):
            resp = send_from_directory(ACC_CLIENT_DIR, path)
            if path.startswith('css/') or path.startswith('js/'):
                resp.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
            return resp
    with open(os.path.join(ACC_CLIENT_DIR, 'index.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    html = inject_asset_version(html, get_asset_version())
    resp = Response(html, mimetype='text/html')
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


if __name__ == '__main__':
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    print(f"Adhya Shakti Shop server starting on http://localhost:5000  [debug={debug}]")
    app.run(debug=debug, port=5000)
